# -*- coding: utf-8 -*-
"""
create_chara_excel.py - SourceChara TSV auto-generation

Reads character definitions from tools/data/charas.py and generates
CWL-format TSV files for JP and EN.
"""

import csv
import os
import re
import sys

# Path setup
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)

sys.path.insert(0, TOOLS_DIR)
from data.charas import CUSTOM_CHARAS
from data.elements import CUSTOM_SPELLS

# Vanilla SourceGame.xlsx (for element alias validation)
VANILLA_SOURCE_GAME_XLSX = os.path.join(
    PROJECT_ROOT, "..", "SourceExcels", "SourceGame.xlsx"
)

# Output paths
OUTPUT_JP_TSV = os.path.join(PROJECT_ROOT, "LangMod", "JP", "Chara.tsv")
OUTPUT_EN_TSV = os.path.join(PROJECT_ROOT, "LangMod", "EN", "Chara.tsv")
OUTPUT_CN_TSV = os.path.join(PROJECT_ROOT, "LangMod", "CN", "Chara.tsv")

# SourceChara column headers (50 columns, matching vanilla SourceChara.CreateRow)
HEADERS = [
    "id", "_id", "name_JP", "name", "aka_JP", "aka", "idActor", "sort",
    "size", "_idRenderData", "tiles", "tiles_snow", "colorMod", "components",
    "defMat", "LV", "chance", "quality", "hostility", "biome", "tag", "trait",
    "race", "job", "tactics", "aiIdle", "aiParam", "actCombat", "mainElement",
    "elements", "equip", "loot", "category", "filter", "gachaFilter", "tone",
    "actIdle", "lightData", "idExtra", "bio", "faith", "works", "hobbies",
    "idText", "moveAnime", "factory", "components", "recruitItems",
    "detail_JP", "detail",
]

# Type info (row 2)
TYPES = [
    "string", "int", "string", "string", "string", "string", "string[]", "int",
    "int[]", "string", "int[]", "int[]", "int", "string[]", "string", "int",
    "int", "int", "string", "string", "string[]", "string[]", "string",
    "string", "string", "string", "int[]", "string[]", "string[]", "elements",
    "string", "string[]", "string", "string[]", "string[]", "string",
    "string[]", "string", "string", "string", "string", "string[]", "string[]",
    "string", "string", "string[]", "string[]", "string[]", "string", "string",
]

# Default values (row 3)
# Default values (row 3) - must match vanilla SourceChara.xlsx
# Col indices: 0=id, 1=_id(552), 9=_idRenderData(chara), 10=tiles(0),
# 13=components(log/1), 15=LV(1), 16=chance(100), 22=race(norland),
# 23=job(none), 32=category(chara)
DEFAULTS = [
    "", "552", "", "", "", "", "", "", "", "chara", "0", "", "0", "log/1",
    "", "1", "100", "", "", "", "", "", "norland", "none", "", "", "", "", "",
    "", "", "", "chara", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
    "", "", "",
]

# Column name -> index (use first occurrence for duplicates)
HEADER_MAP = {}
for idx, name in enumerate(HEADERS):
    if name not in HEADER_MAP:
        HEADER_MAP[name] = idx


def set_cell(row, column_name, value):
    """Set value at column."""
    if column_name in HEADER_MAP:
        row[HEADER_MAP[column_name]] = value


# ============================================================================
# Validation
# ============================================================================


def load_element_aliases(xlsx_path):
    """SourceGame.xlsx の Element シートから alias 一覧を読み込む"""
    aliases = set()
    if not os.path.exists(xlsx_path):
        return aliases

    try:
        import openpyxl
    except ImportError:
        return aliases

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "Element" not in wb.sheetnames:
            return aliases

        ws = wb["Element"]
        headers = list(
            next(ws.iter_rows(min_row=1, max_row=1, max_col=80, values_only=True))
        )
        while headers and headers[-1] is None:
            headers.pop()
        if "id" not in headers or "alias" not in headers:
            return aliases

        id_idx = headers.index("id")
        alias_idx = headers.index("alias")

        empty_run = 0
        for row in ws.iter_rows(min_row=4, max_col=alias_idx + 1, values_only=True):
            row_id = row[id_idx] if id_idx < len(row) else None
            alias = row[alias_idx] if alias_idx < len(row) else None

            if row_id is None and alias in (None, ""):
                empty_run += 1
                if empty_run >= 200:
                    break
                continue

            empty_run = 0
            if isinstance(alias, str) and alias:
                aliases.add(alias)
    finally:
        wb.close()

    return aliases


def load_mod_element_aliases():
    """Mod の CUSTOM_SPELLS から alias 一覧を取得"""
    return {spell["alias"] for spell in CUSTOM_SPELLS if spell.get("alias")}


def validate_act_combat_entry(chara_id, act_combat, known_aliases):
    """actCombat エントリの構文・IDを検証してエラー一覧を返す"""
    errors = []
    if not act_combat:
        return errors

    # パラメータ化プレフィックス（バニラ全種）
    PARAMETERIZED_PREFIXES = (
        "breathe_", "hand_", "arrow_",
        "miasma_", "bolt_", "ball_", "funnel_",
        "sword_", "weapon_", "puddle_", "bit_", "flare_",
    )

    # 特殊アクション（プレフィックスなし、weight不要）
    BARE_ACTIONS = {"Wait"}

    for idx, entry in enumerate(act_combat):
        entry = entry.strip()

        # Wait などの単独アクション（weight なし可）
        if entry in BARE_ACTIONS:
            continue

        # "/" で分割: action/weight または action/weight/flags
        parts = entry.split("/")
        action = parts[0].strip()

        if not action:
            errors.append(
                f"{chara_id}: actCombat[{idx}] has empty action in '{entry}'"
            )
            continue

        # weight 検証（2番目の要素）
        if len(parts) >= 2:
            weight_str = parts[1].strip()
            if not weight_str.isdigit() or int(weight_str) <= 0:
                errors.append(
                    f"{chara_id}: actCombat[{idx}] has invalid weight in '{entry}'"
                )
            # 3番目以降はフラグ（pt など）→ 形式チェックのみ
            for fi, flag in enumerate(parts[2:], start=2):
                flag = flag.strip()
                if not re.match(r"^[a-zA-Z]+$", flag):
                    errors.append(
                        f"{chara_id}: actCombat[{idx}] invalid flag '{flag}' in '{entry}'"
                    )
        else:
            # weight なしは BARE_ACTIONS 以外では不正
            errors.append(
                f"{chara_id}: actCombat[{idx}] '{entry}' is missing weight "
                "(expected '<Action>/<Weight>')"
            )
            continue

        # Hallucination/typo 検知: 既知の act*/Act*/Sp*/St* alias と突合
        if action.startswith(("Act", "Sp", "St", "act")):
            if action not in known_aliases:
                errors.append(
                    f"{chara_id}: actCombat[{idx}] unknown action '{action}' "
                    "(not found in Element aliases)"
                )
            continue

        # パラメータ化された攻撃IDは形式のみ検証
        if action.startswith(PARAMETERIZED_PREFIXES):
            if not re.match(r"^[A-Za-z0-9_]+$", action):
                errors.append(
                    f"{chara_id}: actCombat[{idx}] invalid token '{action}'"
                )
            continue

        # 未知フォーマット
        errors.append(
            f"{chara_id}: actCombat[{idx}] unsupported action format '{action}'"
        )

    return errors


def validate_ai_param(chara_id, ai_param):
    """aiParam の要素数・値を検証してエラー一覧を返す"""
    errors = []
    if not ai_param:
        return errors

    if len(ai_param) != 2:
        errors.append(
            f"{chara_id}: aiParam has {len(ai_param)} elements "
            "(expected 2: [destDist, chanceMove])"
        )
        return errors

    labels = ["destDist", "chanceMove"]
    for i, (val, label) in enumerate(zip(ai_param, labels)):
        if not isinstance(val, int) or val < 0:
            errors.append(
                f"{chara_id}: aiParam[{i}] ({label}) = {val!r} "
                "(expected non-negative integer)"
            )

    return errors


def validate_bio_entry(chara_id, bio):
    """bio 文字列の形式を検証してエラー一覧を返す"""
    errors = []
    if not bio:
        return errors

    gender_set = {"m", "f", "n"}
    value = bio.strip()

    # 短縮形式 ("f", "m", "n")
    if value in gender_set:
        return errors

    parts = value.split("/")
    if len(parts) != 4:
        errors.append(
            f"{chara_id}: bio '{bio}' is invalid "
            "(expected 'gender/version/height/weight' or short gender)"
        )
        return errors

    gender, version, height, weight = [p.strip() for p in parts]
    if gender not in gender_set:
        errors.append(f"{chara_id}: bio has invalid gender '{gender}' in '{bio}'")
    if not version.isdigit():
        errors.append(f"{chara_id}: bio has invalid version '{version}' in '{bio}'")
    if not re.match(r"^-?\d+$", height):
        errors.append(f"{chara_id}: bio has invalid height '{height}' in '{bio}'")
    if not re.match(r"^-?\d+$", weight):
        errors.append(f"{chara_id}: bio has invalid weight '{weight}' in '{bio}'")

    return errors


def validate_all_chara_fields():
    """全キャラの actCombat/aiParam/bio を検証。エラーがあれば終了"""
    known_aliases = set()
    known_aliases.update(load_element_aliases(VANILLA_SOURCE_GAME_XLSX))
    known_aliases.update(load_mod_element_aliases())

    if not known_aliases:
        print(
            "WARNING: actCombat validation skipped "
            "(no Element alias sources found)"
        )
        return

    errors = []
    for chara_id, chara in CUSTOM_CHARAS.items():
        errors.extend(
            validate_act_combat_entry(chara_id, chara.act_combat, known_aliases)
        )
        errors.extend(validate_ai_param(chara_id, chara.ai_param))
        errors.extend(validate_bio_entry(chara_id, chara.bio))

    if errors:
        print("\n[ERROR] chara field validation failed:")
        for e in errors:
            print(f"  - {e}")
        print(
            "\nFix invalid actCombat/aiParam/bio entries in tools/data/charas.py"
        )
        sys.exit(1)

    print("  Chara field validation passed.")


def generate_rows(lang=None):
    """Generate TSV rows from chara definitions.

    Args:
        lang: None for JP/EN, "cn" for CN (swaps name/detail to CN with EN fallback)
    """
    rows = [HEADERS, TYPES, DEFAULTS]

    for chara_id, chara in CUSTOM_CHARAS.items():
        row = [""] * len(HEADERS)

        set_cell(row, "id", chara.id)
        set_cell(row, "name_JP", chara.name_jp)
        set_cell(row, "_idRenderData", chara.render_data)
        set_cell(row, "tiles", chara.tiles)
        set_cell(row, "LV", chara.lv)
        set_cell(row, "chance", chara.chance)
        set_cell(row, "hostility", chara.hostility)
        set_cell(row, "race", chara.race)
        set_cell(row, "bio", chara.bio)

        if lang == "cn":
            set_cell(row, "name", chara.name_cn or chara.name_en)
            set_cell(row, "detail", chara.detail_cn or chara.detail_en)
        else:
            set_cell(row, "name", chara.name_en)
            set_cell(row, "detail", chara.detail_en)

        if chara.job:
            set_cell(row, "job", chara.job)
        if chara.trait_name:
            set_cell(row, "trait", chara.trait_name)
        if chara.tag:
            set_cell(row, "tag", ",".join(chara.tag))
        if chara.tactics:
            set_cell(row, "tactics", chara.tactics)
        if chara.elements:
            set_cell(row, "elements", ",".join(chara.elements))
        if chara.act_combat:
            set_cell(row, "actCombat", ",".join(chara.act_combat))
        if chara.ai_param:
            set_cell(row, "aiParam", ",".join(str(x) for x in chara.ai_param))
        if chara.faith:
            set_cell(row, "faith", chara.faith)

        set_cell(row, "detail_JP", chara.detail_jp)

        rows.append(row)

    return rows


def write_tsv(path, row_data):
    """Write TSV file."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
        writer.writerows(row_data)
    print(f"  Created TSV: {path}")


def main():
    print(f"Generating Chara TSV from {len(CUSTOM_CHARAS)} character definition(s)...")

    validate_all_chara_fields()

    rows = generate_rows()
    write_tsv(OUTPUT_JP_TSV, rows)
    write_tsv(OUTPUT_EN_TSV, rows)

    rows_cn = generate_rows(lang="cn")
    write_tsv(OUTPUT_CN_TSV, rows_cn)

    print(f"Generated {len(CUSTOM_CHARAS)} character(s) (JP + EN + CN)")


if __name__ == "__main__":
    main()
