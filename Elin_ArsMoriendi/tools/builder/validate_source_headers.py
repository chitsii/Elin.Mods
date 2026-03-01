# -*- coding: utf-8 -*-
"""
validate_source_headers.py - MODビルダーのHEADERS/TYPES/DEFAULTSをバニラExcelと照合

MOD側の create_*_excel.py に定義された HEADERS, TYPES, DEFAULTS が
バニラのSource Excelファイルの先頭3行と一致するかを検証する。

Usage:
    # 比較（キャッシュがあればキャッシュを使用、なければExcelから読み取り）
    python validate_source_headers.py compare

    # 静かモード（ビルド統合向け、エラー時のみ出力）
    python validate_source_headers.py compare -q

    # 特定シートのみ
    python validate_source_headers.py compare --sheets Thing,Element

    # バニラExcelからキャッシュ生成
    python validate_source_headers.py extract

    # Google Docsからダウンロードしてキャッシュ再生成
    python validate_source_headers.py sync

Exit codes:
    0 - 全てパス（またはextract/sync成功）
    1 - 検証エラー
"""

import argparse
import importlib.util
import io
import json
import os
import sys
import urllib.request

# ---------------------------------------------------------------------------
# Path setup
# ---------------------------------------------------------------------------
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)

# バニラSourceExcelsのパス（プロジェクトルートから ../SourceExcels）
SOURCE_EXCELS_PATH = os.path.join(PROJECT_ROOT, "..", "SourceExcels")

# JSONキャッシュのパス
CACHE_PATH = os.path.join(BUILDER_DIR, "vanilla_defaults.json")

# ---------------------------------------------------------------------------
# Google Docs IDs（syncコマンド用）
# ---------------------------------------------------------------------------
GOOGLE_DOCS_IDS = {
    "SourceCard.xlsx": "175DaEeB-8qU3N4iBTnaal1ZcP5SU6S_Z",
    "SourceGame.xlsx": "16-LkHtVqjuN9U0rripjBn-nYwyqqSGg_",
}

# ---------------------------------------------------------------------------
# MOD builder <-> バニラExcel マッピング
# ---------------------------------------------------------------------------
# (シート名, バニラExcelファイル名, MODビルダーモジュール名)
SHEET_MAPPINGS = [
    ("Thing", "SourceCard.xlsx", "create_thing_excel"),
    ("Chara", "SourceChara.xlsx", "create_chara_excel"),
    ("Element", "SourceGame.xlsx", "create_element_excel"),
    ("Stat", "SourceGame.xlsx", "create_stat_excel"),
]


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------
def _trim_trailing_empty(lst: list[str]) -> list[str]:
    """末尾の空文字列を除去したリストを返す。"""
    result = list(lst)
    while result and result[-1] == "":
        result.pop()
    return result


def _normalize(value) -> str:
    """None や数値を文字列に正規化する。"""
    if value is None:
        return ""
    return str(value)


# ---------------------------------------------------------------------------
# Read vanilla Excel
# ---------------------------------------------------------------------------
def read_vanilla_excel(filepath: str, sheet_name: str) -> dict:
    """
    バニラExcelからヘッダー3行を読み取る。

    Returns:
        {"headers": [...], "types": [...], "defaults": [...]}
    """
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {filepath}")

    ws = wb[sheet_name]
    max_col = ws.max_column

    rows = []
    for row_num in range(1, 4):  # rows 1-3
        row_data = []
        for col in range(1, max_col + 1):
            val = ws.cell(row=row_num, column=col).value
            row_data.append(_normalize(val))
        rows.append(_trim_trailing_empty(row_data))

    wb.close()
    return {
        "headers": rows[0],
        "types": rows[1],
        "defaults": rows[2],
    }


# ---------------------------------------------------------------------------
# Read MOD builder definitions
# ---------------------------------------------------------------------------
def load_builder_module(module_name: str):
    """
    importlib.util でビルダーモジュールを読み込む。
    stdout と sys.argv の副作用を抑制する。
    """
    module_path = os.path.join(BUILDER_DIR, f"{module_name}.py")
    if not os.path.exists(module_path):
        raise FileNotFoundError(f"Builder module not found: {module_path}")

    spec = importlib.util.spec_from_file_location(module_name, module_path)
    mod = importlib.util.module_from_spec(spec)

    # sys.argv を退避（モジュールが argparse 等を使う場合の安全策）
    saved_argv = sys.argv
    sys.argv = [module_path]

    # stdout を抑制
    saved_stdout = sys.stdout
    sys.stdout = io.StringIO()

    # sys.path にTOOLS_DIRを追加（data.* インポート用）
    path_added = False
    if TOOLS_DIR not in sys.path:
        sys.path.insert(0, TOOLS_DIR)
        path_added = True

    try:
        spec.loader.exec_module(mod)
    finally:
        sys.argv = saved_argv
        sys.stdout = saved_stdout
        if path_added and TOOLS_DIR in sys.path:
            sys.path.remove(TOOLS_DIR)

    return mod


def read_mod_definitions(module_name: str) -> dict:
    """
    ビルダーモジュールからHEADERS, TYPES, DEFAULTSを読み取る。

    Returns:
        {"headers": [...], "types": [...], "defaults": [...]}
    """
    mod = load_builder_module(module_name)

    headers = list(mod.HEADERS)
    types = list(mod.TYPES)

    # DEFAULTS: dict の場合はリストに変換、list の場合はそのまま
    raw_defaults = mod.DEFAULTS
    if isinstance(raw_defaults, dict):
        defaults = [str(raw_defaults.get(h, "")) for h in headers]
    else:
        defaults = [str(v) if v is not None else "" for v in raw_defaults]

    return {
        "headers": headers,
        "types": types,
        "defaults": defaults,
    }


# ---------------------------------------------------------------------------
# JSON cache
# ---------------------------------------------------------------------------
def load_cache() -> dict | None:
    """JSONキャッシュを読み込む。存在しなければNone。"""
    if not os.path.exists(CACHE_PATH):
        return None
    with open(CACHE_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def save_cache(data: dict):
    """JSONキャッシュに書き込む。"""
    with open(CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  キャッシュ保存: {CACHE_PATH}")


# ---------------------------------------------------------------------------
# Compare
# ---------------------------------------------------------------------------
def compare_rows(
    vanilla: dict,
    mod: dict,
    sheet_name: str,
    module_name: str,
) -> tuple[list[str], bool]:
    """
    バニラとMODの定義を比較する。

    Returns:
        (出力行リスト, OK かどうか)
    """
    lines = []
    ok = True

    v_headers = vanilla["headers"]
    m_headers = mod["headers"]
    v_types = vanilla["types"]
    m_types = mod["types"]
    v_defaults = vanilla["defaults"]
    m_defaults = mod["defaults"]

    lines.append(f"--- {sheet_name} ({module_name}.py) ---")
    lines.append(f"  バニラ: {len(v_headers)} カラム, MOD: {len(m_headers)} カラム")

    diffs_headers = []
    diffs_types = []
    diffs_defaults = []

    max_cols = max(len(v_headers), len(m_headers))

    for i in range(max_cols):
        vh = v_headers[i] if i < len(v_headers) else ""
        mh = m_headers[i] if i < len(m_headers) else ""
        if vh != mh:
            diffs_headers.append(f"    Col {i}: バニラ='{vh}', MOD='{mh}'")

    for i in range(max_cols):
        vt = v_types[i] if i < len(v_types) else ""
        mt = m_types[i] if i < len(m_types) else ""
        if vt != mt:
            col_name = v_headers[i] if i < len(v_headers) else f"Col{i}"
            diffs_types.append(f"    {col_name} (Col {i}): バニラ='{vt}', MOD='{mt}'")

    for i in range(max_cols):
        vd = v_defaults[i] if i < len(v_defaults) else ""
        md = m_defaults[i] if i < len(m_defaults) else ""
        if vd != md:
            col_name = v_headers[i] if i < len(v_headers) else f"Col{i}"
            diffs_defaults.append(f"    {col_name} (Col {i}): バニラ='{vd}', MOD='{md}'")

    if diffs_headers:
        ok = False
        lines.append("  [カラム名の差異]")
        lines.extend(diffs_headers)

    if diffs_types:
        ok = False
        lines.append("  [型の差異]")
        lines.extend(diffs_types)

    if diffs_defaults:
        ok = False
        lines.append("  [デフォルト値の差異]")
        lines.extend(diffs_defaults)

    if ok:
        lines.append("  OK")

    return lines, ok


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------
def cmd_extract():
    """バニラExcelからJSONキャッシュを生成する。"""
    print("バニラExcelからキャッシュを抽出中...")

    cache = {}
    for sheet_name, excel_file, _module in SHEET_MAPPINGS:
        excel_path = os.path.join(SOURCE_EXCELS_PATH, excel_file)
        if not os.path.exists(excel_path):
            print(f"  [SKIP] {excel_file} が見つかりません")
            continue

        print(f"  読み取り中: {excel_file} -> {sheet_name}")
        data = read_vanilla_excel(excel_path, sheet_name)
        cache[sheet_name] = data

    save_cache(cache)
    print(f"抽出完了: {len(cache)} シート")
    return 0


def cmd_sync():
    """Google Docsからダウンロードしてキャッシュを再生成する。"""
    print("Google Docsからダウンロード中...")

    os.makedirs(SOURCE_EXCELS_PATH, exist_ok=True)

    for filename, doc_id in GOOGLE_DOCS_IDS.items():
        url = f"https://drive.google.com/uc?export=download&id={doc_id}"
        dest = os.path.join(SOURCE_EXCELS_PATH, filename)
        print(f"  ダウンロード中: {filename} ...")

        try:
            urllib.request.urlretrieve(url, dest)
            print(f"  保存: {dest}")
        except Exception as e:
            print(f"  [ERROR] {filename}: {e}")
            return 1

    print("ダウンロード完了。キャッシュを生成します...")
    return cmd_extract()


def cmd_compare(quiet: bool = False, sheets_filter: list[str] | None = None):
    """MODビルダー定義をバニラと比較する。"""
    # キャッシュまたはExcelからバニラデータを取得
    cache = load_cache()
    vanilla_from_cache = cache is not None

    all_ok = True
    output_lines = []

    for sheet_name, excel_file, module_name in SHEET_MAPPINGS:
        # シートフィルタ
        if sheets_filter and sheet_name not in sheets_filter:
            continue

        # バニラデータ取得
        vanilla_data = None
        if vanilla_from_cache and sheet_name in cache:
            vanilla_data = cache[sheet_name]
        else:
            # キャッシュにないのでExcelから直接読み取り
            excel_path = os.path.join(SOURCE_EXCELS_PATH, excel_file)
            if not os.path.exists(excel_path):
                if not quiet:
                    output_lines.append(f"--- {sheet_name} ({module_name}.py) ---")
                    output_lines.append(f"  [SKIP] バニラExcelが見つかりません: {excel_file}")
                continue
            try:
                import openpyxl  # noqa: F401 - ensure available
                vanilla_data = read_vanilla_excel(excel_path, sheet_name)
            except Exception as e:
                output_lines.append(f"--- {sheet_name} ({module_name}.py) ---")
                output_lines.append(f"  [ERROR] バニラExcel読み取りエラー: {e}")
                all_ok = False
                continue

        # MODビルダーデータ取得
        try:
            mod_data = read_mod_definitions(module_name)
        except Exception as e:
            output_lines.append(f"--- {sheet_name} ({module_name}.py) ---")
            output_lines.append(f"  [ERROR] MODビルダー読み取りエラー: {e}")
            all_ok = False
            continue

        # 比較
        lines, ok = compare_rows(vanilla_data, mod_data, sheet_name, module_name)
        if not ok:
            all_ok = False
        output_lines.extend(lines)

    # 出力
    if quiet:
        # エラー時のみ出力
        if not all_ok:
            for line in output_lines:
                print(line)
    else:
        for line in output_lines:
            print(line)

    return 0 if all_ok else 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="MODビルダーのHEADERS/TYPES/DEFAULTSをバニラExcelと照合"
    )
    subparsers = parser.add_subparsers(dest="command")

    # compare (default)
    p_compare = subparsers.add_parser("compare", help="バニラExcelとMOD定義を比較")
    p_compare.add_argument(
        "-q", "--quiet",
        action="store_true",
        help="静かモード（エラー時のみ出力）",
    )
    p_compare.add_argument(
        "--sheets",
        type=str,
        default=None,
        help="比較対象シート（カンマ区切り、例: Thing,Element）",
    )

    # extract
    subparsers.add_parser("extract", help="バニラExcelからJSONキャッシュを生成")

    # sync
    subparsers.add_parser("sync", help="Google Docsからダウンロードしてキャッシュ再生成")

    args = parser.parse_args()

    # デフォルトはcompare
    command = args.command or "compare"

    if command == "compare":
        quiet = getattr(args, "quiet", False)
        sheets_raw = getattr(args, "sheets", None)
        sheets_filter = [s.strip() for s in sheets_raw.split(",")] if sheets_raw else None
        sys.exit(cmd_compare(quiet=quiet, sheets_filter=sheets_filter))

    elif command == "extract":
        sys.exit(cmd_extract())

    elif command == "sync":
        sys.exit(cmd_sync())


if __name__ == "__main__":
    main()
