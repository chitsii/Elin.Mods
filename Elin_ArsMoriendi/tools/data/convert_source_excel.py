"""SourceExcel (xlsx) を CSV に変換する。

使い方:
    python tools/data/convert_source_excel.py

出力先: tools/data/csv/<ファイル名>_<シート名>.csv
ゲーム更新時に再実行すること。
"""

import csv
import os
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.join(SCRIPT_DIR, "..", "..")
SOURCE_EXCELS_DIR = os.path.join(PROJECT_ROOT, "..", "SourceExcels")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "csv")

MAX_ROWS = 100_000  # これを超えるシートはプレースホルダとみなしスキップ


def convert():
    if not os.path.isdir(SOURCE_EXCELS_DIR):
        print(f"SourceExcels が見つかりません: {SOURCE_EXCELS_DIR}", file=sys.stderr)
        sys.exit(1)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for xlsx_file in sorted(os.listdir(SOURCE_EXCELS_DIR)):
        if not xlsx_file.endswith(".xlsx"):
            continue
        filepath = os.path.join(SOURCE_EXCELS_DIR, xlsx_file)
        prefix = xlsx_file.replace(".xlsx", "")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            out_path = os.path.join(OUTPUT_DIR, f"{prefix}_{sheet_name}.csv")

            row_count = 0
            with open(out_path, "w", newline="", encoding="utf-8") as csvfile:
                writer = csv.writer(csvfile)
                for row in ws.iter_rows(values_only=True):
                    if all(v is None for v in row):
                        continue
                    row_count += 1
                    if row_count > MAX_ROWS:
                        break
                    writer.writerow(row)

            if row_count > MAX_ROWS:
                os.remove(out_path)
                print(f"  SKIP  {prefix}_{sheet_name} (>{MAX_ROWS} rows)")
            else:
                size = os.path.getsize(out_path)
                print(f"  OK    {prefix}_{sheet_name}.csv ({row_count} rows, {size:,} bytes)")

        wb.close()

    print(f"\n出力先: {os.path.abspath(OUTPUT_DIR)}")


if __name__ == "__main__":
    convert()
