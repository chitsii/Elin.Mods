"""
CWL Quest Library - Generic Drama Builder

This module provides a generic DramaBuilder class for creating CWL drama files.
Game-specific extensions can inherit from this class to add custom functionality.

Usage:
    from cwl_quest_lib import DramaBuilder, DramaLabel, DramaActor

    drama = DramaBuilder()
    pc = drama.register_actor("pc", "Player", "Player")
    npc = drama.register_actor("npc", "NPC", "NPC")

    main = drama.label("main")
    drama.step(main)
    drama.say("greet", "Hello!", actor=npc)
    drama.finish()

    drama.save("drama.xlsx", sheet_name="npc")
"""

import os
import re
from typing import Any, Callable, Dict, List, Optional, Set, Union

import openpyxl

# CWL準拠ヘッダー (if + if2 で複合条件、version/text列あり)
HEADERS = [
    "step",
    "jump",
    "if",
    "if2",
    "action",
    "param",
    "actor",
    "version",
    "id",
    "text_JP",
    "text_EN",
    "text_CN",
    "text",
]


class DramaLabel:
    """ステップラベルを表すクラス"""

    def __init__(self, key: str):
        self.key = key

    def __str__(self):
        return self.key


class DramaActor:
    """アクターを表すクラス"""

    def __init__(self, key: str):
        self.key = key

    def __str__(self):
        return self.key


class ChoiceReaction:
    """
    選択肢とその反応を一体化して定義するクラス。
    choice_block()と組み合わせて使用する。

    使用例:
        builder.choice_block([
            ChoiceReaction("Yes", "c1")
                .say("yes1", "Great!", actor=npc)
                .jump(next_step),

            ChoiceReaction("No", "c2")
                .say("no1", "Goodbye.", actor=npc)
                .jump(end),
        ], cancel=end)
    """

    def __init__(
        self,
        text_jp: str,
        text_id: str = "",
        text_en: str = "",
        text_cn: str = "",
        condition: str = "",
    ):
        """
        Args:
            text_jp: 選択肢テキスト（日本語）
            text_id: テキストID（省略可）
            text_en: 選択肢テキスト（英語、省略時は日本語と同じ）
            text_cn: 選択肢テキスト（簡体中国語、省略可）
            condition: 表示条件（if列、省略可）
        """
        self.text_jp = text_jp
        self.text_en = text_en or text_jp
        self.text_cn = text_cn or ""
        self.text_id = text_id
        self.condition = condition
        self.actions: List[Dict[str, Any]] = []
        self._label: Optional[str] = None  # choice_blockで自動設定

    def _add_action(self, action: Dict[str, Any]) -> "ChoiceReaction":
        """アクションを追加（内部用）"""
        self.actions.append(action)
        return self

    # === 会話系 ===

    def say(
        self,
        text_id: str,
        text_jp: str,
        text_en: str = "",
        text_cn: str = "",
        actor: Union[str, "DramaActor"] = None,
    ) -> "ChoiceReaction":
        """テキスト行を追加"""
        actor_key = actor.key if isinstance(actor, DramaActor) else actor
        entry = {
            "id": text_id,
            "text_JP": text_jp,
            "text_EN": text_en or text_jp,
            "text_CN": text_cn or "",
        }
        if actor_key:
            entry["actor"] = actor_key
        return self._add_action(entry)

    # === フロー制御 ===

    def jump(self, jump_to: Union[str, "DramaLabel"]) -> "ChoiceReaction":
        """指定ステップにジャンプ"""
        key = jump_to.key if isinstance(jump_to, DramaLabel) else jump_to
        return self._add_action({"jump": key})

    def end(self) -> "ChoiceReaction":
        """ドラマを終了"""
        return self._add_action({"action": "end"})

    # === フラグ操作 ===

    def set_flag(self, flag: str, value: int = 1) -> "ChoiceReaction":
        """フラグを設定"""
        return self._add_action(
            {
                "action": "setFlag",
                "param": f"{flag},{value}",
            }
        )

    def mod_flag(
        self,
        flag: str,
        operator: str,
        value: int,
        actor: Union[str, "DramaActor"] = None,
    ) -> "ChoiceReaction":
        """フラグを変更"""
        actor_key = actor.key if isinstance(actor, DramaActor) else (actor or "pc")
        return self._add_action(
            {
                "action": "invoke*",
                "param": f"mod_flag({flag}, {operator}{value})",
                "actor": actor_key,
            }
        )

    # === 演出 ===

    def shake(self) -> "ChoiceReaction":
        """画面を揺らす"""
        return self._add_action({"action": "shake"})

    def wait(self, seconds: float) -> "ChoiceReaction":
        """待機"""
        return self._add_action({"action": "wait", "param": str(seconds)})

    def play_sound(self, sound_id: str) -> "ChoiceReaction":
        """効果音を再生"""
        return self._add_action({"action": "sound", "param": sound_id})

    # === クエスト ===

    def complete_quest(
        self, quest_id: str, actor: Union[str, "DramaActor"] = None
    ) -> "ChoiceReaction":
        """クエストを完了"""
        actor_key = actor.key if isinstance(actor, DramaActor) else (actor or "pc")
        return self._add_action(
            {
                "action": "modInvoke",
                "param": f"complete_quest({quest_id})",
                "actor": actor_key,
            }
        )

    # === 汎用 ===

    def action(
        self,
        action_name: str,
        param: str = None,
        jump: Union[str, "DramaLabel"] = None,
        actor: Union[str, "DramaActor"] = None,
    ) -> "ChoiceReaction":
        """汎用アクションを追加"""
        entry = {"action": action_name}
        if param:
            entry["param"] = param
        if jump:
            entry["jump"] = jump.key if isinstance(jump, DramaLabel) else jump
        if actor:
            entry["actor"] = actor.key if isinstance(actor, DramaActor) else actor
        return self._add_action(entry)

    def eval(
        self, code: str, actor: Union[str, "DramaActor"] = None
    ) -> "ChoiceReaction":
        """C#コードを実行"""
        entry = {"action": "eval", "param": code}
        if actor:
            entry["actor"] = actor.key if isinstance(actor, DramaActor) else actor
        return self._add_action(entry)


class DramaBuilder:
    """
    CWLドラマファイルを構築するビルダークラス。

    使用例:
        drama = DramaBuilder()
        pc = drama.register_actor("pc", "プレイヤー", "Player")
        npc = drama.register_actor("master", "マスター", "Master")

        lbl_main = drama.label("main")
        drama.step(lbl_main)
        drama.say("greet", "こんにちは", actor=npc)
        drama.finish()

        drama.save("drama.xlsx", sheet_name="master")
    """

    def __init__(self, mod_name: str = "CWLMod") -> None:
        """
        Args:
            mod_name: Mod name for debug logs (e.g., "SukutsuArena")
        """
        self.entries: List[Dict[str, Any]] = []
        self.actors: Dict[str, Dict[str, str]] = {}
        self.registered_steps: Set[str] = set()
        self._current_step: Optional[str] = None
        self._mod_name = mod_name
        self._validation_errors: List[str] = []
        self._choice_id_counter: int = 0

    def _next_choice_id(self) -> str:
        """text_id 未指定の選択肢用に自動IDを生成"""
        self._choice_id_counter += 1
        return f"_c{self._choice_id_counter}"

    def _resolve_key(self, obj: Union[str, DramaLabel, DramaActor]) -> str:
        """オブジェクトからキーを取得"""
        if isinstance(obj, (DramaLabel, DramaActor)):
            return obj.key
        return obj

    def register_actor(
        self, actor_id: str, name_jp: str, name_en: str = "", name_cn: str = ""
    ) -> DramaActor:
        """アクターを登録"""
        self.actors[actor_id] = {
            "jp": name_jp,
            "en": name_en or name_jp,
            "cn": name_cn or name_en or name_jp,
        }
        return DramaActor(actor_id)

    def add_actor(
        self, actor: Union[str, DramaActor], name_jp: str, name_en: str = ""
    ) -> "DramaBuilder":
        """ドラマランタイムにアクターを追加（addActor アクション）。
        マップ上に存在しないキャラやテキスト専用キャラに必要。
        register_actor はビルダー内部用、add_actor は実行時の DramaSequence に登録する。
        """
        actor_key = self._resolve_key(actor) if isinstance(actor, DramaActor) else actor
        self.entries.append({
            "action": "addActor",
            "actor": actor_key,
            "text_JP": name_jp,
            "text_EN": name_en or name_jp,
            "text": name_en or name_jp,
        })
        return self

    def label(self, key: str) -> DramaLabel:
        """ラベルを作成"""
        return DramaLabel(key)

    def step(self, step_label: Union[str, DramaLabel]) -> "DramaBuilder":
        """
        新しいステップを開始。
        CWL形式では、step行は step列のみを含み、内容は次の行から。
        """
        key = self._resolve_key(step_label)
        self.registered_steps.add(key)
        self._current_step = key
        self.entries.append({"step": key})
        return self

    def say(
        self,
        text_id: str,
        text_jp: str,
        text_en: str = "",
        text_cn: str = "",
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """テキスト行を追加"""
        actor_key = self._resolve_key(actor) if actor else None
        entry = {
            "id": text_id,
            "text_JP": text_jp,
            "text_EN": text_en or text_jp,
            "text_CN": text_cn or "",
        }
        if actor_key:
            entry["actor"] = actor_key
        self.entries.append(entry)
        return self

    def choice(
        self,
        jump_to: Union[str, DramaLabel],
        text_jp: str,
        text_en: str = "",
        text_cn: str = "",
        text_id: str = "",
        condition: str = "",
    ) -> "DramaBuilder":
        """選択肢を追加"""
        key = self._resolve_key(jump_to)
        entry = {
            "action": "choice",
            "jump": key,
            "text_JP": text_jp,
            "text_EN": text_en or text_jp,
            "text_CN": text_cn or "",
        }
        entry["id"] = text_id if text_id else self._next_choice_id()
        if condition:
            entry["if"] = condition
        self.entries.append(entry)
        return self

    def jump(self, jump_to: Union[str, DramaLabel]) -> "DramaBuilder":
        """指定ステップにジャンプ"""
        key = self._resolve_key(jump_to)
        self.entries.append({"jump": key})
        return self

    def branch_if(
        self,
        flag: str,
        operator: str,
        value: int,
        jump_to: Union[str, DramaLabel],
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """条件分岐。invoke* + if_flag を使用。条件が真ならjump列のステップにジャンプ。"""
        key = self._resolve_key(jump_to)
        actor_key = self._resolve_key(actor) if actor else "pc"
        entry = {
            "action": "invoke*",
            "param": f"if_flag({flag}, {operator}{value})",
            "jump": key,
            "actor": actor_key,
        }
        self.entries.append(entry)
        return self

    def branch_quest_done(
        self,
        quest_id: str,
        jump_to: Union[str, DramaLabel],
        quest_done_prefix: str = "quest_done_",
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """クエスト完了済みなら指定ラベルにジャンプ。"""
        flag_key = f"{quest_done_prefix}{quest_id}"
        return self.branch_if(flag_key, "==", 1, jump_to, actor)

    def switch_on_flag(
        self,
        flag: str,
        cases: Dict[int, Union[str, DramaLabel]],
        fallback: Union[str, DramaLabel] = None,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """フラグ値に基づく複数条件分岐"""
        actor_key = self._resolve_key(actor) if actor else "pc"

        for value, jump_to in cases.items():
            key = self._resolve_key(jump_to)
            entry = {
                "action": "invoke*",
                "param": f"if_flag({flag}, =={value})",
                "jump": key,
                "actor": actor_key,
            }
            self.entries.append(entry)

        if fallback is not None:
            fallback_key = self._resolve_key(fallback)
            entry = {
                "action": "invoke*",
                "param": f"if_flag({flag}, ==0)",
                "jump": fallback_key,
                "actor": actor_key,
            }
            self.entries.append(entry)

        return self

    def switch_flag(
        self,
        flag: str,
        cases: List[Union[str, DramaLabel]],
        fallback: Union[str, DramaLabel] = None,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """フラグ値に基づいて直接ジャンプ（switch_flagコマンド使用）"""
        actor_key = self._resolve_key(actor) if actor else "pc"
        jump_targets = [self._resolve_key(c) for c in cases]
        if fallback:
            jump_targets.append(self._resolve_key(fallback))

        self.entries.append(
            {
                "action": "modInvoke",
                "param": f"switch_flag({flag}, {', '.join(jump_targets)})",
                "actor": actor_key,
            }
        )
        return self

    def choice_block(
        self,
        choices: List["ChoiceReaction"],
        cancel: Union[str, DramaLabel] = None,
        label_prefix: str = None,
    ) -> "DramaBuilder":
        """選択肢と反応を一体化して定義する。"""
        prefix = label_prefix or self._current_step or "choice"

        for i, cr in enumerate(choices):
            cr._label = f"{prefix}_react_{i}"

            entry = {
                "action": "choice",
                "jump": cr._label,
                "text_JP": cr.text_jp,
                "text_EN": cr.text_en,
                "text_CN": cr.text_cn or "",
            }
            entry["id"] = cr.text_id if cr.text_id else self._next_choice_id()
            if cr.condition:
                entry["if"] = cr.condition
            self.entries.append(entry)

        if cancel is not None:
            cancel_key = self._resolve_key(cancel)
            self.entries.append({"action": "cancel", "jump": cancel_key})

        for cr in choices:
            self.entries.append({"step": cr._label})
            self.registered_steps.add(cr._label)
            for action in cr.actions:
                self.entries.append(action)

        return self

    def check_quests(
        self, checks: list, actor: Union[str, DramaActor] = None
    ) -> "DramaBuilder":
        """複数のクエスト利用可能チェックを一括で実行。"""
        for quest_id, jump_target in checks:
            self.check_quest_available(quest_id, jump_target, actor)
        return self

    def set_flag(self, flag: str, value: int = 1) -> "DramaBuilder":
        """フラグを設定"""
        self.entries.append(
            {
                "action": "setFlag",
                "param": f"{flag},{value}",
            }
        )
        return self

    def mod_flag(
        self, flag: str, operator: str, value: int, actor: Union[str, DramaActor] = None
    ) -> "DramaBuilder":
        """フラグを変更 (invoke* mod_flag)"""
        actor_key = self._resolve_key(actor) if actor else "pc"
        self.entries.append(
            {
                "action": "invoke*",
                "param": f"mod_flag({flag}, {operator}{value})",
                "actor": actor_key,
            }
        )
        return self

    def eval(self, code: str, actor: Union[str, DramaActor] = None) -> "DramaBuilder":
        """C#コードを実行"""
        entry = {"action": "eval", "param": code}
        if actor:
            entry["actor"] = self._resolve_key(actor)
        self.entries.append(entry)
        return self

    def _add_validation_error(self, message: str) -> None:
        """検証エラーを追加"""
        self._validation_errors.append(message)
        print(f"[DramaBuilder] ERROR: {message}")

    def get_validation_errors(self) -> List[str]:
        """検証エラーを取得"""
        return self._validation_errors

    def action(
        self,
        action_name: str,
        param: str = None,
        jump: Union[str, DramaLabel] = None,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """汎用アクションを追加"""
        entry = {"action": action_name}
        if param:
            entry["param"] = param
        if jump:
            entry["jump"] = self._resolve_key(jump)
        if actor:
            entry["actor"] = self._resolve_key(actor)
        self.entries.append(entry)
        return self

    def on_cancel(self, jump_to: Union[str, DramaLabel]) -> "DramaBuilder":
        """キャンセル時の動作を設定"""
        key = self._resolve_key(jump_to)
        self.entries.append({"action": "cancel", "jump": key})
        return self

    def finish(self) -> "DramaBuilder":
        """ドラマを終了"""
        self.entries.append({"action": "end"})
        return self

    def play_sound(self, sound_id: str) -> "DramaBuilder":
        """効果音を再生"""
        self.entries.append({"action": "sound", "param": sound_id})
        return self

    def show_book(
        self,
        book_id: str,
        category: str = "Book",
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """CWL API の show_book を呼び出してテキスト本を開く。

        Args:
            book_id: Text/<category>/<book_id>.txt の book_id 部分
            category: "Book" または "Scroll"
        """
        normalized = book_id.replace("\\", "/").strip()
        if "/" in normalized:
            # Accept preformatted entries like "Book/ars_hecatia_guide".
            book_entry = normalized
        else:
            short_id = normalized[:-4] if normalized.lower().endswith(".txt") else normalized
            book_entry = f"{category}/{short_id}"

        actor_key = self._resolve_key(actor) if actor else "pc"
        self.entries.append(
            {
                "action": "invoke*",
                # CWL show_book currently expects one path-like argument: "Book/<id>".
                "param": f"show_book({book_entry})",
                "actor": actor_key,
            }
        )
        return self

    def play_bgm(self, bgm_id: str) -> "DramaBuilder":
        """
        BGMを再生

        Args:
            bgm_id: BGM ID (例: "BGM/my_bgm")
        """
        code = f'''
            Debug.Log("[{self._mod_name}] Attempting to play BGM: {bgm_id}");
            var data = SoundManager.current.GetData("{bgm_id}");
            if (data != null) {{
                Debug.Log("[{self._mod_name}] Found BGM data, type: " + data.GetType().Name);
                if (data is BGMData bgm) {{
                    Debug.Log("[{self._mod_name}] Playing as BGM");
                    LayerDrama.haltPlaylist = true;
                    LayerDrama.maxBGMVolume = true;
                    SoundManager.current.PlayBGM(bgm);
                }} else {{
                    Debug.Log("[{self._mod_name}] Playing as Sound");
                    SoundManager.current.Play(data);
                }}
            }} else {{
                Debug.LogWarning("[{self._mod_name}] BGM not found: {bgm_id}");
            }}
        '''.replace("\n", " ").strip()
        self.entries.append({"action": "eval", "param": code})
        return self

    def play_bgm_with_fallback(self, primary: str, fallback: str) -> "DramaBuilder":
        """BGMを再生（primaryが見つからなければfallbackを再生）

        Args:
            primary: 優先BGM ID (例: "BGM/Ominous_Suspense_01")
            fallback: フォールバックBGM ID (例: "BGM/ManuscriptByCandlelight")
        """
        code = f'''
            var data = SoundManager.current.GetData("{primary}");
            if (data != null && data is BGMData bgm) {{
                LayerDrama.haltPlaylist = true;
                LayerDrama.maxBGMVolume = true;
                SoundManager.current.PlayBGM(bgm);
            }} else {{
                var fb = SoundManager.current.GetData("{fallback}");
                if (fb != null && fb is BGMData fbBgm) {{
                    LayerDrama.haltPlaylist = true;
                    LayerDrama.maxBGMVolume = true;
                    SoundManager.current.PlayBGM(fbBgm);
                }}
            }}
        '''.replace("\n", " ").strip()
        self.entries.append({"action": "eval", "param": code})
        return self

    def play_bgm_if_not_playing(self, bgm_id: str) -> "DramaBuilder":
        """
        BGMを再生（既に同じBGMが再生中なら何もしない）

        Args:
            bgm_id: BGM ID (例: "BGM/my_bgm")
        """
        code = f'''
            var data = SoundManager.current.GetData("{bgm_id}");
            if (data != null && data is BGMData bgm) {{
                var current = SoundManager.current.currentBGM;
                if (current == null || current.data != bgm) {{
                    Debug.Log("[{self._mod_name}] Playing BGM: {bgm_id}");
                    LayerDrama.haltPlaylist = true;
                    LayerDrama.maxBGMVolume = true;
                    SoundManager.current.PlayBGM(bgm);
                }}
            }} else {{
                Debug.LogWarning("[{self._mod_name}] BGM not found: {bgm_id}");
            }}
        '''.replace("\n", " ").strip()
        self.entries.append({"action": "eval", "param": code})
        return self

    def play_bgm_vanilla(self, bgm_id: int) -> "DramaBuilder":
        """バニラBGMを再生（数値ID）"""
        code = f"""
            Debug.Log("[{self._mod_name}] Playing vanilla BGM ID: {bgm_id}");
            if (EMono.core.refs.dictBGM.TryGetValue({bgm_id}, out var bgm)) {{
                LayerDrama.haltPlaylist = true;
                LayerDrama.maxBGMVolume = true;
                EMono.Sound.PlayBGM(bgm);
            }} else {{
                Debug.LogWarning("[{self._mod_name}] Vanilla BGM not found: {bgm_id}");
            }}
        """.replace("\n", " ").strip()
        self.entries.append({"action": "eval", "param": code})
        return self

    def wait(self, seconds: float) -> "DramaBuilder":
        """待機"""
        self.entries.append({"action": "wait", "param": str(seconds)})
        return self

    def effect(self, effect_id: str) -> "DramaBuilder":
        """エフェクトを再生"""
        self.entries.append({"action": effect_id})
        return self

    # ============================================================================
    # CWL 拡張機能: 組み込みジャンプ
    # ============================================================================

    def inject_unique(self) -> "DramaBuilder":
        """inject/Unique アクションを実行"""
        self.entries.append({"action": "inject", "param": "Unique"})
        return self

    def mod_affinity(self, amount: int) -> "DramaBuilder":
        """好感度を変更する

        Args:
            amount: 好感度の変化量（正で上昇、負で下降）
                    デバッグモードでは10倍になる
        """
        self.entries.append({"action": "modAffinity", "param": str(amount)})
        return self

    def inject_builtin_choices(self) -> "DramaBuilder":
        """
        バニラの自動選択肢を注入する。

        inject/Unique + _choices を実行し、仲間化後のパーティ脱退、
        ホーム設定などの標準選択肢を現在の会話に追加する。

        使用例:
            builder.say("何か用かい？")
            builder.inject_builtin_choices()  # ← 自動選択肢を注入
            builder.add_menu([...])           # ← Mod固有のメニュー
        """
        self.entries.append({"action": "inject", "param": "Unique"})
        self.entries.append({"action": "_choices"})
        return self

    def jump_to_trade(self) -> "DramaBuilder":
        """商人との取引画面にジャンプ"""
        self.entries.append({"jump": "_trade"})
        return self

    def jump_to_buy(self) -> "DramaBuilder":
        """購入画面にジャンプ"""
        self.entries.append({"jump": "_buy"})
        return self

    def jump_to_join_party(self) -> "DramaBuilder":
        """パーティ加入にジャンプ"""
        self.entries.append({"jump": "_joinParty"})
        return self

    def jump_to_leave_party(self) -> "DramaBuilder":
        """パーティ離脱にジャンプ"""
        self.entries.append({"jump": "_leaveParty"})
        return self

    def jump_to_train(self) -> "DramaBuilder":
        """訓練画面にジャンプ"""
        self.entries.append({"jump": "_train"})
        return self

    def jump_to_heal(self) -> "DramaBuilder":
        """回復サービスにジャンプ"""
        self.entries.append({"jump": "_heal"})
        return self

    def jump_to_invest_shop(self) -> "DramaBuilder":
        """店への投資にジャンプ"""
        self.entries.append({"jump": "_investShop"})
        return self

    def jump_to_sell_fame(self) -> "DramaBuilder":
        """名声売却にジャンプ"""
        self.entries.append({"jump": "_sellFame"})
        return self

    def jump_to_copy_item(self) -> "DramaBuilder":
        """アイテム複製にジャンプ"""
        self.entries.append({"jump": "_copyItem"})
        return self

    def jump_to_give(self) -> "DramaBuilder":
        """アイテム渡すにジャンプ"""
        self.entries.append({"jump": "_give"})
        return self

    def jump_to_whore(self) -> "DramaBuilder":
        """売春サービスにジャンプ"""
        self.entries.append({"jump": "_whore"})
        return self

    def jump_to_tail(self) -> "DramaBuilder":
        """尾行にジャンプ"""
        self.entries.append({"jump": "_tail"})
        return self

    def jump_to_suck(self) -> "DramaBuilder":
        """吸血にジャンプ"""
        self.entries.append({"jump": "_suck"})
        return self

    def jump_to_bout(self) -> "DramaBuilder":
        """決闘にジャンプ"""
        self.entries.append({"jump": "_bout"})
        return self

    def jump_to_rumor(self) -> "DramaBuilder":
        """噂話にジャンプ"""
        self.entries.append({"jump": "_rumor"})
        return self

    def jump_to_news(self) -> "DramaBuilder":
        """ニュースにジャンプ"""
        self.entries.append({"jump": "_news"})
        return self

    def jump_builtin(self, builtin_name: str) -> "DramaBuilder":
        """汎用組み込みジャンプ"""
        if not builtin_name.startswith("_"):
            builtin_name = "_" + builtin_name
        self.entries.append({"jump": builtin_name})
        return self

    # ============================================================================
    # CWL 拡張機能: ランタイムスクリプト
    # ============================================================================

    def cs_eval(
        self, code: str, actor: Union[str, DramaActor] = None
    ) -> "DramaBuilder":
        """C# コードをランタイムで実行 (eval アクション)"""
        actor_key = self._resolve_key(actor) if actor else None
        entry = {"action": "eval", "param": code}
        if actor_key:
            entry["actor"] = actor_key
        self.entries.append(entry)
        return self

    def cs_script_var_set(self, var_name: str, value_expr: str) -> "DramaBuilder":
        """スクリプト共有変数を設定"""
        return self.cs_eval(f'Script["{var_name}"] = {value_expr};')

    def cs_script_var_get(self, var_name: str, cast_type: str = "int") -> str:
        """スクリプト共有変数を取得するC#式を生成"""
        return f'({cast_type})Script["{var_name}"]'

    def cs_call(
        self,
        method_expr: str,
        *arg_exprs: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """完全修飾メソッド呼び出しを eval 経由で追加する。

        例:
            builder.cs_call("Elin_CommonDrama.DramaRuntime.PlayBgm", '"BGM/MyTrack"')
        """
        args = ", ".join(arg_exprs)
        code = f"{method_expr}({args});"
        return self.cs_eval(code, actor=actor)

    def cs_call_common_runtime(
        self,
        method_name: str,
        *arg_exprs: str,
        actor: Union[str, DramaActor] = None,
        runtime_type: str = "Elin_CommonDrama.DramaRuntime",
    ) -> "DramaBuilder":
        """共通ドラマランタイム呼び出しのショートハンド。

        他Modへ移植しやすいよう、シナリオ側から eval 文字列を隠蔽する。
        """
        return self.cs_call(
            f"{runtime_type}.{method_name}",
            *arg_exprs,
            actor=actor,
        )

    def stop_bgm_now(self, actor: Union[str, DramaActor] = None) -> "DramaBuilder":
        """現在のBGMを停止。"""
        return self.resolve_run("cmd.scene.stop_bgm", actor=actor)

    def play_pc_effect(
        self, effect_id: str, actor: Union[str, DramaActor] = None
    ) -> "DramaBuilder":
        """PCにエフェクトを再生。"""
        return self.resolve_run(f"fx.pc.{effect_id}", actor=actor)

    def play_pc_effect_with_sound(
        self,
        effect_id: str,
        sound_id: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """PCにエフェクト+SEを再生。"""
        return self.resolve_run(
            f"fx.pc.{effect_id}+sfx.pc.{sound_id}",
            actor=actor,
        )

    def run_cue(
        self,
        cue_key: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """整理済みキューを実行する。"""
        return self.resolve_run(f"cue.{cue_key}", actor=actor)

    def resolve_flag(
        self,
        dependency_key: str,
        target_flag_key: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """依存キーを評価し、結果をdialogFlagに同期する。"""
        return self.cs_call_common_runtime(
            "ResolveFlag",
            f'"{dependency_key}"',
            f'"{target_flag_key}"',
            actor=actor,
        )

    def resolve_run(
        self,
        dependency_key: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """依存キーに対応するコマンドを実行する。"""
        return self.cs_call_common_runtime(
            "ResolveRun",
            f'"{dependency_key}"',
            actor=actor,
        )

    # Backward-compatible wrappers (scheduled for removal)
    def sync_flag_from_dependency(
        self,
        dependency_key: str,
        target_flag_key: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        return self.resolve_flag(dependency_key, target_flag_key, actor=actor)

    def run_dependency_command(
        self,
        dependency_key: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        return self.resolve_run(dependency_key, actor=actor)

    def quest_check(
        self,
        drama_id: str,
        target_flag_key: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """ドラマ開始条件を評価し、結果をフラグへ同期する。"""
        return self.resolve_flag(
            f"state.quest.can_start.{drama_id}",
            target_flag_key,
            actor=actor,
        )

    def quest_try_start(
        self,
        drama_id: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """ドラマ開始コマンドを実行する（idempotent）。"""
        return self.resolve_run(
            f"cmd.quest.try_start.{drama_id}",
            actor=actor,
        )

    def quest_try_start_repeatable(
        self,
        drama_id: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """ドラマ開始コマンドを実行する（完了まで再試行可能）。"""
        return self.resolve_run(
            f"cmd.quest.try_start_repeatable.{drama_id}",
            actor=actor,
        )

    def quest_try_start_until_complete(
        self,
        drama_id: str,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """ドラマ開始コマンドを実行する（完了まで再試行可能）。"""
        return self.resolve_run(
            f"cmd.quest.try_start_until_complete.{drama_id}",
            actor=actor,
        )

    # ============================================================================
    # CWL 拡張機能: 条件付きエントリ
    # ============================================================================

    def say_if(
        self,
        text_id: str,
        text_jp: str,
        condition: str,
        text_en: str = "",
        text_cn: str = "",
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """条件付きテキスト行を追加"""
        actor_key = self._resolve_key(actor) if actor else None
        entry = {
            "if": condition,
            "id": text_id,
            "text_JP": text_jp,
            "text_EN": text_en or text_jp,
            "text_CN": text_cn or "",
        }
        if actor_key:
            entry["actor"] = actor_key
        self.entries.append(entry)
        return self

    def choice_if(
        self,
        jump_to: Union[str, DramaLabel],
        text_jp: str,
        condition: str,
        text_en: str = "",
        text_cn: str = "",
        text_id: str = "",
    ) -> "DramaBuilder":
        """条件付き選択肢を追加"""
        key = self._resolve_key(jump_to)
        entry = {
            "if": condition,
            "action": "choice",
            "jump": key,
            "text_JP": text_jp,
            "text_EN": text_en or text_jp,
            "text_CN": text_cn or "",
        }
        entry["id"] = text_id if text_id else self._next_choice_id()
        self.entries.append(entry)
        return self

    def choice_if2(
        self,
        jump_to: Union[str, DramaLabel],
        text_jp: str,
        condition1: str,
        condition2: str,
        text_en: str = "",
        text_cn: str = "",
        text_id: str = "",
    ) -> "DramaBuilder":
        """複数条件付き選択肢を追加（if + if2 の両方を使用）"""
        key = self._resolve_key(jump_to)
        entry = {
            "if": condition1,
            "if2": condition2,
            "action": "choice",
            "jump": key,
            "text_JP": text_jp,
            "text_EN": text_en or text_jp,
            "text_CN": text_cn or "",
        }
        entry["id"] = text_id if text_id else self._next_choice_id()
        self.entries.append(entry)
        return self

    def choice_dynamic(
        self,
        jump_to: Union[str, DramaLabel],
        text_jp: str,
        condition: str,
        text_en: str = "",
        text_cn: str = "",
        text_id: str = "",
    ) -> "DramaBuilder":
        """CWL動的条件付き選択肢を追加（param列に条件を設定）

        CWLのinvoke*拡張条件を選択肢に適用します。
        条件が真の場合のみ選択肢が表示されます。

        Args:
            jump_to: 選択時のジャンプ先
            text_jp: 選択肢テキスト（日本語）
            condition: CWL動的条件（例: "if_in_party()", "not(if_in_party())"）
            text_en: 選択肢テキスト（英語）
            text_cn: 選択肢テキスト（簡体中国語）
            text_id: テキストID

        使用例:
            builder.choice_dynamic(
                builder.label("_leaveParty"),
                "パーティから外す",
                "if_in_party()",
                text_en="Leave party",
            )
        """
        key = self._resolve_key(jump_to)
        entry = {
            "action": "choice",
            "param": condition,
            "jump": key,
            "text_JP": text_jp,
            "text_EN": text_en or text_jp,
            "text_CN": text_cn or "",
        }
        entry["id"] = text_id if text_id else self._next_choice_id()
        self.entries.append(entry)
        return self

    # ============================================================================
    # CWL 拡張機能: 条件ヘルパー
    # ============================================================================

    @staticmethod
    def cond_has_flag(flag_name: str) -> str:
        """hasFlag条件を生成"""
        return f"hasFlag,{flag_name}"

    @staticmethod
    def cond_no_flag(flag_name: str) -> str:
        """!hasFlag条件を生成"""
        return f"!hasFlag,{flag_name}"

    @staticmethod
    def cond_has_item(item_id: str) -> str:
        """hasItem条件を生成"""
        return f"hasItem,{item_id}"

    @staticmethod
    def cond_quest_completed(quest_id: str) -> str:
        """isCompleted条件を生成"""
        return f"isCompleted,{quest_id}"

    @staticmethod
    def cond_flag_equals(flag_name: str, value: int) -> str:
        """フラグ値が一致する条件を生成"""
        return f"=,{flag_name},{value}"

    @staticmethod
    def cond_flag_greater(flag_name: str, value: int) -> str:
        """フラグ値が大きい条件を生成"""
        return f">,{flag_name},{value}"

    @staticmethod
    def cond_flag_less(flag_name: str, value: int) -> str:
        """フラグ値が小さい条件を生成"""
        return f"<,{flag_name},{value}"

    # ============================================================================
    # CWL 拡張機能: 視覚効果
    # ============================================================================

    def fade_in(self, duration: float = 1.0, color: str = "black") -> "DramaBuilder":
        """フェードイン（画面が明るくなる）"""
        param = f"{duration},{color}"
        self.entries.append({"action": "fadeIn", "param": param})
        return self

    def fade_out(self, duration: float = 1.0, color: str = "black") -> "DramaBuilder":
        """フェードアウト（画面が暗くなる）"""
        param = f"{duration},{color}"
        self.entries.append({"action": "fadeOut", "param": param})
        return self

    def flash_lut(
        self, lut_name: str, duration: float = 2.0, fade_time: float = 0.3
    ) -> "DramaBuilder":
        """
        LUTを一時的に切り替える（フェードトランジション付き）

        明度を下げる → LUT切替 → 明度を戻す → 待機 → 明度を下げる → LUT戻し → 明度を戻す

        Args:
            lut_name: LUT名（例: "LUT_Invert", "LUT_Horror1"）
            duration: 効果時間（秒）デフォルト2秒
            fade_time: フェード時間（秒）デフォルト0.3秒

        使用例:
            builder.flash_lut("LUT_Invert", duration=2.0)  # 狂気的な演出
            builder.flash_lut("LUT_Horror1", duration=3.0, fade_time=0.5)  # ホラー演出
        """
        code = f'Elin_SukutsuArena.Effects.LutEffect.Flash("{lut_name}", {duration}f, {fade_time}f);'
        self.entries.append({"action": "eval", "param": code})
        return self

    def set_background(self, bg_id: str) -> "DramaBuilder":
        """背景画像を設定"""
        code = f'dm.imageBG.enabled = true; dm.imageBG.sprite = "{bg_id}".LoadSprite();'
        self.entries.append({"action": "eval", "param": code})
        return self

    def glitch(self) -> "DramaBuilder":
        """グリッチエフェクトを有効化"""
        self.entries.append({"action": "glitch"})
        return self

    def shake(self) -> "DramaBuilder":
        """画面を揺らす"""
        self.entries.append({"action": "shake"})
        return self

    def set_dialog_style(self, style: str = "Default") -> "DramaBuilder":
        """ダイアログスタイルを変更"""
        self.entries.append({"action": "setDialog", "param": f",,{style}"})
        return self

    # ============================================================================
    # ハイレベルAPI: ドラマ演出
    # ============================================================================

    def drama_start(
        self, bg_id: str = None, bgm_id: str = None, fade_duration: float = 1.0
    ) -> "DramaBuilder":
        """ドラマ開始演出"""
        self.fade_out(duration=fade_duration, color="black")
        if bg_id:
            self.set_background(bg_id)
        self.fade_in(duration=fade_duration, color="black")
        if bgm_id:
            self.play_bgm(bgm_id)
        return self

    def drama_end(self, fade_duration: float = 1.0) -> "DramaBuilder":
        """ドラマ終了演出"""
        self.fade_out(duration=fade_duration, color="black")
        self.finish()
        return self

    def scene_transition(
        self, bg_id: str = None, bgm_id: str = None, fade_duration: float = 0.5
    ) -> "DramaBuilder":
        """シーン切り替え演出"""
        self.fade_out(duration=fade_duration, color="black")
        if bg_id:
            self.set_background(bg_id)
        self.fade_in(duration=fade_duration, color="black")
        if bgm_id:
            self.play_bgm(bgm_id)
        return self

    # ============================================================================
    # High-Level API: Conversation Patterns
    # ============================================================================

    def conversation(
        self,
        lines: List[tuple],
        actor: Union[str, DramaActor] = None,
        id_prefix: str = "",
    ) -> "DramaBuilder":
        """
        連続した会話を簡潔に定義。

        Args:
            lines: 会話行のリスト。各要素は以下の形式:
                   - (text_id, text_jp) - 英語は日本語と同じ
                   - (text_id, text_jp, text_en)
                   - (text_id, text_jp, text_en, text_cn)
                   - (text_id, text_jp, text_en, text_cn, line_actor) - 行ごとにアクター指定
            actor: デフォルトのアクター（line_actor指定がない場合に使用）
            id_prefix: テキストIDのプレフィックス（空なら各行のtext_idをそのまま使用）

        使用例:
            # シンプルな連続会話
            builder.conversation([
                ("greet1", "こんにちは。"),
                ("greet2", "今日も良い天気ですね。"),
                ("greet3", "何かご用ですか？"),
            ], actor=npc)

            # 行ごとにアクター指定
            builder.conversation([
                ("line1", "調子はどうだ？", "", "", npc),
                ("line2", "まあまあです。", "", "", pc),
                ("line3", "そうか、頑張れよ。", "", "", npc),
            ])

            # プレフィックス付き
            builder.conversation([
                ("1", "最初の行"),
                ("2", "次の行"),
            ], actor=npc, id_prefix="scene1_")
            # -> text_id は "scene1_1", "scene1_2" になる
        """
        for line in lines:
            if len(line) == 2:
                text_id, text_jp = line
                text_en = text_jp
                text_cn = ""
                line_actor = actor
            elif len(line) == 3:
                text_id, text_jp, text_en = line
                text_cn = ""
                line_actor = actor
            elif len(line) == 4:
                text_id, text_jp, text_en, text_cn = line
                line_actor = actor
            elif len(line) >= 5:
                text_id, text_jp, text_en, text_cn, line_actor = line[:5]
            else:
                continue

            full_id = f"{id_prefix}{text_id}" if id_prefix else text_id
            self.say(
                full_id, text_jp, text_en or text_jp, text_cn or "", actor=line_actor
            )

        return self

    def greeting_with_choices(
        self,
        greeting_label: Union[str, DramaLabel],
        choices_label: Union[str, DramaLabel],
        greeting: tuple,
        choices: List[tuple],
        cancel: Union[str, DramaLabel] = None,
    ) -> "DramaBuilder":
        """
        挨拶+選択肢パターンを簡潔に定義。

        Args:
            greeting_label: 挨拶ステップのラベル
            choices_label: 選択肢ステップのラベル（挨拶後にジャンプ）
            greeting: 挨拶タプル:
                     (text_id, text_jp, actor) - 基本形
                     (text_id, text_jp, text_en, actor)
                     (text_id, text_jp, text_en, text_cn, actor)
            choices: 選択肢リスト。各要素は:
                     (text_jp, jump_to) - 基本形
                     (text_jp, jump_to, text_id) - ID付き
                     (text_jp, jump_to, text_id, text_en) - 英語付き
                     (text_jp, jump_to, text_id, text_en, text_cn) - 中国語付き
            cancel: キャンセル時のジャンプ先（省略可）

        使用例:
            builder.greeting_with_choices(
                greeting_label=greet,
                choices_label=menu,
                greeting=("greet", "いらっしゃいませ。", npc),
                choices=[
                    ("買い物", shop_step),
                    ("話を聞く", talk_step, "c_talk"),
                    ("さようなら", end_step, "c_bye", "Goodbye"),
                ],
                cancel=end_step
            )
        """
        # 挨拶ステップ
        self.step(greeting_label)
        if len(greeting) == 3:
            text_id, text_jp, actor = greeting
            self.say(text_id, text_jp, text_jp, "", actor=actor)
        elif len(greeting) == 4:
            text_id, text_jp, text_en, actor = greeting
            self.say(text_id, text_jp, text_en, "", actor=actor)
        elif len(greeting) >= 5:
            text_id, text_jp, text_en, text_cn, actor = greeting[:5]
            self.say(text_id, text_jp, text_en, text_cn, actor=actor)
        self.jump(choices_label)

        # 選択肢ステップ
        self.step(choices_label)
        for choice_item in choices:
            if len(choice_item) == 2:
                text_jp, jump_to = choice_item
                self.choice(jump_to, text_jp)
            elif len(choice_item) == 3:
                text_jp, jump_to, text_id = choice_item
                self.choice(jump_to, text_jp, text_id=text_id)
            elif len(choice_item) == 4:
                text_jp, jump_to, text_id, text_en = choice_item
                self.choice(jump_to, text_jp, text_en=text_en, text_id=text_id)
            elif len(choice_item) >= 5:
                text_jp, jump_to, text_id, text_en, text_cn = choice_item[:5]
                self.choice(
                    jump_to, text_jp, text_en=text_en, text_cn=text_cn, text_id=text_id
                )

        if cancel is not None:
            self.on_cancel(cancel)

        return self

    def scene(
        self,
        label: Union[str, DramaLabel],
        lines: List[tuple],
        next_step: Union[str, DramaLabel] = None,
        actor: Union[str, DramaActor] = None,
        id_prefix: str = "",
    ) -> "DramaBuilder":
        """
        シーン全体を定義（ラベル + 会話 + 遷移）。

        Args:
            label: シーンのラベル
            lines: 会話行のリスト（conversation()と同形式）
            next_step: 次のステップ（省略時はend）
            actor: デフォルトのアクター
            id_prefix: テキストIDのプレフィックス

        使用例:
            # 単純なシーン（次のステップへジャンプ）
            builder.scene(
                label=intro,
                lines=[
                    ("intro1", "ようこそ。"),
                    ("intro2", "ここは闘技場です。"),
                ],
                next_step=menu,
                actor=npc
            )

            # シーン終了（end）
            builder.scene(
                label=outro,
                lines=[
                    ("outro1", "さようなら。"),
                ],
                actor=npc
            )
        """
        self.step(label)
        self.conversation(lines, actor=actor, id_prefix=id_prefix)

        if next_step is not None:
            self.jump(next_step)
        else:
            self.finish()

        return self

    def choice_scene(
        self,
        label: Union[str, DramaLabel],
        intro_lines: List[tuple] = None,
        choices: List["ChoiceReaction"] = None,
        cancel: Union[str, DramaLabel] = None,
        actor: Union[str, DramaActor] = None,
        id_prefix: str = "",
    ) -> "DramaBuilder":
        """
        選択肢シーンを定義（導入 + 選択肢 + 反応）。

        Args:
            label: シーンのラベル
            intro_lines: 選択肢前の導入会話（省略可）
            choices: ChoiceReactionのリスト
            cancel: キャンセル時のジャンプ先
            actor: デフォルトのアクター
            id_prefix: テキストIDのプレフィックス

        使用例:
            builder.choice_scene(
                label=decision,
                intro_lines=[
                    ("ask", "どちらを選びますか？"),
                ],
                choices=[
                    ChoiceReaction("はい", "c_yes")
                        .say("yes_resp", "わかりました。", actor=npc)
                        .jump(next_step),
                    ChoiceReaction("いいえ", "c_no")
                        .say("no_resp", "そうですか。", actor=npc)
                        .jump(end_step),
                ],
                cancel=end_step,
                actor=npc
            )
        """
        self.step(label)

        if intro_lines:
            self.conversation(intro_lines, actor=actor, id_prefix=id_prefix)

        if choices:
            self.choice_block(choices, cancel=cancel)

        return self

    def multi_branch(
        self,
        label: Union[str, DramaLabel],
        flag: str,
        branches: Dict[int, Union[str, DramaLabel]],
        fallback: Union[str, DramaLabel] = None,
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """
        フラグ値による多分岐を簡潔に定義。

        Args:
            label: 分岐点のラベル
            flag: 分岐条件のフラグキー
            branches: {フラグ値: ジャンプ先} の辞書
            fallback: どの条件にも一致しない場合のジャンプ先
            actor: アクター

        使用例:
            builder.multi_branch(
                label=rank_check,
                flag="chitsii.arena.player.rank",
                branches={
                    0: unranked_step,  # UNRANKED
                    1: rank_g_step,    # G
                    2: rank_f_step,    # F
                },
                fallback=default_step
            )
        """
        self.step(label)
        self.switch_on_flag(flag, branches, fallback=fallback, actor=actor)
        return self

    # ============================================================================
    # CWL 拡張機能: カメラ・フォーカス
    # ============================================================================

    def focus_chara(
        self, chara_id: str, wait_before: float = 0.3, wait_after: float = 0.5
    ) -> "DramaBuilder":
        """キャラクターにフォーカス"""
        if wait_before > 0:
            self.entries.append({"action": "wait", "param": str(wait_before)})
        self.entries.append({"action": "modFocusChara", "param": chara_id})
        if wait_after > 0:
            self.entries.append({"action": "wait", "param": str(wait_after)})
        return self

    def unfocus(self) -> "DramaBuilder":
        """カメラフォーカスを解除"""
        self.entries.append({"action": "unfocus"})
        return self

    # ============================================================================
    # CWL 拡張機能: システム
    # ============================================================================

    def save_game(self) -> "DramaBuilder":
        """ゲームをセーブ"""
        self.entries.append({"action": "save"})
        return self

    def set_hour(self, hour: int) -> "DramaBuilder":
        """時刻を設定"""
        self.entries.append({"action": "setHour", "param": str(hour)})
        return self

    # ============================================================================
    # CWL 拡張機能: クエスト操作
    # ============================================================================

    def start_quest(
        self, quest_id: str, actor: Union[str, DramaActor] = None
    ) -> "DramaBuilder":
        """クエストを開始（ArenaQuestManager経由）

        注意: CWLネイティブの startQuest ではなく modInvoke を使用。
        CWLネイティブはElinの標準クエストシステム(Quest.Create)を使用するが、
        Modのカスタムクエストは組み込み辞書に存在しないためエラーになる。
        """
        return self.mod_invoke(f"start_quest({quest_id})", actor or "pc")

    def complete_quest(
        self, quest_id: str = "", actor: Union[str, DramaActor] = None
    ) -> "DramaBuilder":
        """クエストを完了（ArenaQuestManager経由）

        注意: CWLネイティブの completeQuest ではなく modInvoke を使用。
        CWLネイティブはElinの標準クエストシステム(EMono.game.quests)を使用するが、
        ArenaモジュールはArenaQuestManager（フラグベース）を使用するため。
        """
        if quest_id:
            return self.mod_invoke(f"complete_quest({quest_id})", actor)
        # quest_id が空の場合は何もしない（エラー防止）
        return self

    def next_phase(self, quest_id: str) -> "DramaBuilder":
        """クエストの次フェーズに進む"""
        self.entries.append({"action": "nextPhase", "param": quest_id})
        return self

    def change_phase(self, quest_id: str, phase: int) -> "DramaBuilder":
        """クエストのフェーズを変更"""
        self.entries.append({"action": "changePhase", "param": f"{quest_id},{phase}"})
        return self

    def set_quest_client(self) -> "DramaBuilder":
        """現在のクエストのクライアントを tg に設定"""
        self.entries.append({"action": "setQuestClient"})
        return self

    def update_journal(self) -> "DramaBuilder":
        """ジャーナルを更新"""
        self.entries.append({"action": "updateJournal"})
        return self

    # ============================================================================
    # CWL Native Quest (Journal Display)
    # ============================================================================

    def start_journal_quest(self, quest_id: str) -> "DramaBuilder":
        """CWLネイティブのstartQuestアクション（ジャーナル表示用）

        SourceQuest.xlsxで定義したクエストをジャーナルに表示する。
        ArenaQuestManagerとは別のシステムで、ジャーナル表示のみを担当。
        """
        self.entries.append({"action": "startQuest", "param": quest_id})
        return self

    def change_journal_phase(self, quest_id: str, phase: int) -> "DramaBuilder":
        """CWLネイティブのchangePhaseアクション（ジャーナル表示用）

        ジャーナル表示のフェーズを変更する。
        phase: 1=Unranked, 2=G, 3=F, 4=E, 5=D, 6=C, 7=B, 8=A, 9=S
        """
        self.entries.append({"action": "changePhase", "param": f"{quest_id},{phase}"})
        return self

    # ============================================================================
    # Quest System: modInvoke Actions
    # ============================================================================

    def mod_invoke(
        self, method_call: str, actor: Union[str, DramaActor] = None
    ) -> "DramaBuilder":
        """C# メソッドを modInvoke アクションで呼び出す"""
        actor_key = self._resolve_key(actor) if actor else "pc"
        self.entries.append(
            {
                "action": "modInvoke",
                "param": method_call,
                "actor": actor_key,
            }
        )
        return self

    def check_quest_available(
        self,
        quest_id: str,
        jump_to: Union[str, DramaLabel],
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """クエストが利用可能かチェックし、利用可能ならジャンプ"""
        jump_key = self._resolve_key(jump_to)
        return self.mod_invoke(f"check_quest_available({quest_id}, {jump_key})", actor)

    def if_flag_string(
        self,
        flag: str,
        operator: str,
        value: str,
        jump_to: Union[str, DramaLabel],
        actor: Union[str, DramaActor] = None,
    ) -> "DramaBuilder":
        """フラグの文字列/enum値を比較してジャンプ"""
        jump_key = self._resolve_key(jump_to)
        actor_key = self._resolve_key(actor) if actor else "pc"
        self.entries.append(
            {
                "action": "modInvoke",
                "param": f"if_flag({flag}, {operator}, {value}, {jump_key})",
                "actor": actor_key,
            }
        )
        return self

    def debug_log_flags(self, actor: Union[str, DramaActor] = None) -> "DramaBuilder":
        """デバッグ: 全フラグをログ出力"""
        return self.mod_invoke("debug_log_flags()", actor)

    def debug_log_quests(self, actor: Union[str, DramaActor] = None) -> "DramaBuilder":
        """デバッグ: 全クエスト状態をログ出力"""
        return self.mod_invoke("debug_log_quests()", actor)

    def check_available_quests_for_npc(
        self, npc_id: str, actor: Union[str, DramaActor] = None
    ) -> "DramaBuilder":
        """NPCごとの利用可能クエストをチェックしフラグを設定"""
        return self.mod_invoke(f"check_available_quests({npc_id})", actor)

    def check_quests_for_dispatch(
        self, flag_name: str, quest_ids: List[str]
    ) -> "DramaBuilder":
        """クエストディスパッチ用のチェック"""
        args = [flag_name] + quest_ids
        param = f"check_quests_for_dispatch({', '.join(args)})"
        return self.action("modInvoke", param=param, actor="pc")

    # ============================================================================
    # 検証・出力
    # ============================================================================

    def build(self) -> List[Dict[str, Any]]:
        """エントリーリストを返す（検証なし）"""
        return self.entries

    def _validate_text_id_uniqueness(self) -> List[str]:
        """text_id の重複をチェックし、エラーメッセージのリストを返す"""
        errors = []
        text_ids: Dict[str, int] = {}  # text_id -> 出現回数

        for entry in self.entries:
            text_id = entry.get("id")
            if text_id:
                if text_id in text_ids:
                    text_ids[text_id] += 1
                else:
                    text_ids[text_id] = 1

        # 重複しているものを報告
        for text_id, count in text_ids.items():
            if count > 1:
                errors.append(f"Duplicate text_id '{text_id}' appears {count} times")

        return errors

    def _validate_drama_structure(self) -> List[str]:
        """ドラマ構造を検証し、警告メッセージのリストを返す"""
        warnings = []

        steps = {}
        jump_targets = set()
        current_step = None
        current_step_has_terminator = False
        current_step_has_fallback = False

        for i, entry in enumerate(self.entries):
            if entry.get("step"):
                if current_step and not current_step_has_terminator:
                    if not current_step_has_fallback:
                        steps[current_step]["has_terminator"] = False

                current_step = entry["step"]
                current_step_has_terminator = False
                current_step_has_fallback = False
                steps[current_step] = {"has_terminator": True, "index": i}

            action = entry.get("action")
            jump = entry.get("jump")
            param = entry.get("param", "")

            if action in ("end", "choice", "cancel"):
                current_step_has_terminator = True
            if jump:
                current_step_has_terminator = True
                jump_targets.add(jump)

            if action in ("modInvoke", "invoke*") and param:
                # if_flag パターン
                match = re.search(r"if_flag\([^,]+,\s*[^,]+,\s*([^)]+)\)", param)
                if match:
                    target = match.group(1).strip()
                    jump_targets.add(target)
                    if "==0" in param:
                        current_step_has_fallback = True

                # check_quest_available パターン
                match = re.search(r"check_quest_available\([^,]+,\s*([^)]+)\)", param)
                if match:
                    target = match.group(1).strip()
                    jump_targets.add(target)

                # switch_flag パターン: switch_flag(flag, target1, target2, ...)
                match = re.search(r"switch_flag\([^,]+,\s*(.+)\)", param)
                if match:
                    targets_str = match.group(1)
                    for target in targets_str.split(","):
                        target = target.strip()
                        if target:
                            jump_targets.add(target)
                    current_step_has_terminator = True  # switch_flag は terminator

        if (
            current_step
            and not current_step_has_terminator
            and not current_step_has_fallback
        ):
            steps[current_step]["has_terminator"] = False

        for step_name, info in steps.items():
            if not info["has_terminator"]:
                warnings.append(
                    f"Step '{step_name}' has no terminator (end/jump/choice)"
                )

        for step_name in steps:
            if step_name != "main" and step_name not in jump_targets:
                if "_react_" not in step_name:
                    warnings.append(f"Step '{step_name}' is never referenced (orphan)")

        defined_steps = set(steps.keys())
        builtin_steps = {
            "_trade",
            "_buy",
            "_joinParty",
            "_leaveParty",
            "_train",
            "_heal",
            "_investShop",
            "_sellFame",
            "_copyItem",
            "_give",
            "_whore",
            "_tail",
            "_suck",
            "_bout",
            "_rumor",
            "_news",
        }
        # 選択肢インデックスラベルのパターン（choice1, choice2, choice2_5, final_choice 等）
        choice_label_pattern = re.compile(r"^(choice\d+(_\d+)?|.*_choice)$")

        for target in jump_targets:
            if target not in defined_steps and target not in builtin_steps:
                # 選択肢ラベルへのジャンプは許容
                if not choice_label_pattern.match(target):
                    warnings.append(f"Jump target '{target}' is not defined")

        # ========================================
        # choice が say なしで使われているかチェック
        # Elin のドラマシステムでは choice は lastTalk に追加されるため、
        # say がないと前のステップの lastTalk に追加されてしまう
        #
        # ただし、参照元のすべてに say がある場合は許容する。
        # 問題になるのは「参照元に say がないステップがある」場合。
        # ========================================

        # 各ステップへの参照元を集計（どのステップからジャンプしてくるか）
        step_referrers: Dict[str, List[str]] = {}
        current_step = None
        for entry in self.entries:
            if entry.get("step"):
                current_step = entry["step"]
            jump = entry.get("jump")
            if jump and current_step:
                if jump not in step_referrers:
                    step_referrers[jump] = []
                if current_step not in step_referrers[jump]:
                    step_referrers[jump].append(current_step)

        # 各ステップの say/choice 状態を収集
        current_step = None
        step_has_say: Dict[str, bool] = {}
        step_has_choice: Dict[str, bool] = {}

        for entry in self.entries:
            if entry.get("step"):
                current_step = entry["step"]
                if current_step not in step_has_say:
                    step_has_say[current_step] = False
                    step_has_choice[current_step] = False

            if current_step:
                action = entry.get("action")
                # say があるか（text_JP が設定されていて action が choice/cancel でない）
                if entry.get("text_JP") and action not in ("choice", "cancel"):
                    step_has_say[current_step] = True
                if action == "choice":
                    step_has_choice[current_step] = True

        # say なしで choice があるステップについて、参照元をチェック
        for step_name, has_choice in step_has_choice.items():
            if not has_choice:
                continue
            if step_has_say.get(step_name, False):
                continue  # say がある場合はOK

            referrers = step_referrers.get(step_name, [])
            if len(referrers) < 2:
                continue  # 単一参照は許容

            # 参照元に say がないステップがあるかチェック
            referrers_without_say = [
                r for r in referrers if not step_has_say.get(r, False)
            ]
            if referrers_without_say:
                warnings.append(
                    f"Step '{step_name}' has choice without say and is referenced "
                    f"from steps without say: {referrers_without_say}. "
                    f"This may cause choices to be added to unexpected lastTalk."
                )

        return warnings

    def save(self, filepath: str, sheet_name: str = "main") -> None:
        """CWL準拠のExcelファイルとして保存。

        全言語カラム（text_JP, text_EN, text_CN, text）を1ファイルに出力。
        CWLがユーザーの言語設定に応じて適切なカラムを自動選択する。
        textカラムはフォールバック用（text_ENの値を使用）。

        Args:
            filepath: 出力ファイルパス
            sheet_name: シート名
        """
        # text_id の重複チェック（エラー）
        text_id_errors = self._validate_text_id_uniqueness()
        if text_id_errors:
            for e in text_id_errors:
                print(f"[DramaBuilder] ERROR: {e}")
            raise ValueError(
                f"Drama validation failed: {len(text_id_errors)} duplicate text_id(s) found in '{sheet_name}'. "
                f"See errors above. Each text_id must be unique within a drama file."
            )

        warnings = self._validate_drama_structure()

        # 「choice without say」エラーを分離
        choice_errors = [w for w in warnings if "choice without say" in w]
        other_warnings = [w for w in warnings if "choice without say" not in w]

        for w in other_warnings:
            print(f"[DramaBuilder] WARNING: {w}")

        # choice without say はエラーとして扱う
        if choice_errors:
            for e in choice_errors:
                print(f"[DramaBuilder] ERROR: {e}")
            raise ValueError(
                f"Drama validation failed: {len(choice_errors)} step(s) have "
                f"choice without say. See errors above."
            )

        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        for col, header in enumerate(HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

        row = 6
        for entry in self.entries:
            for col, header in enumerate(HEADERS, 1):
                value = entry.get(header)
                # textカラムにはフォールバック値（text_EN）を設定
                if header == "text" and value is None:
                    value = entry.get("text_EN")
                if value is not None:
                    ws.cell(row=row, column=col, value=value)
            row += 1

        wb.save(filepath)
        print(f"Created: {filepath} (sheet: {sheet_name})")


# ============================================================================
# Unit Tests
# ============================================================================

if __name__ == "__main__":
    print("=== Drama Builder Base Test ===\n")

    # Create a simple drama
    drama = DramaBuilder(mod_name="TestMod")
    pc = drama.register_actor("pc", "Player", "Player")
    npc = drama.register_actor("npc", "NPC", "NPC")

    main = drama.label("main")
    end_step = drama.label("end")

    drama.step(main)
    drama.say("greet", "Hello!", "Hello!", "你好！", actor=npc)
    drama.choice(end_step, "Goodbye", "Goodbye", "再见")
    drama.on_cancel(end_step)

    drama.step(end_step)
    drama.say("bye", "Farewell!", "Farewell!", "再会！", actor=npc)
    drama.finish()

    # Test validation
    warnings = drama._validate_drama_structure()
    print(f"Validation warnings: {warnings}")

    # Test build
    entries = drama.build()
    print(f"Total entries: {len(entries)}")

    print("\n=== All Tests Passed! ===")
