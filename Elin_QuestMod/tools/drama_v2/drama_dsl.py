from __future__ import annotations

import hashlib
import os
from collections import OrderedDict
from dataclasses import dataclass, field
from typing import Any, Iterable, Optional, Union

import openpyxl


HEADERS = [
    "step",
    "jump",
    "if",
    "if2",
    "action",
    "param",
    "actor",
    "version",
    "id",
    "text_JP",
    "text_EN",
    "text_CN",
    "text",
]


@dataclass(frozen=True)
class Id:
    kind: str
    value: str


@dataclass(frozen=True)
class Chara:
    alias: str


@dataclass(frozen=True)
class NodeRef:
    name: str


@dataclass(frozen=True)
class Cond:
    op: str
    args: tuple[Any, ...] = ()


@dataclass(frozen=True)
class OptionSpec:
    label: str
    to: Union[NodeRef, str]
    cond: Optional[Cond] = None


@dataclass(frozen=True)
class StepSpec:
    kind: str
    payload: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class NodeSpec:
    ref: NodeRef
    steps: list[StepSpec]


@dataclass(frozen=True)
class StorySpec:
    start: NodeRef
    nodes: list[NodeSpec]
    meta: dict[str, Any] = field(default_factory=dict)


def _actor_alias(actor: Union[str, Chara, None]) -> Optional[str]:
    if actor is None:
        return None
    if isinstance(actor, Chara):
        return actor.alias
    return str(actor)


def _as_ref(target: Union[NodeRef, str]) -> NodeRef:
    if isinstance(target, NodeRef):
        return target
    return NodeRef(str(target))


def _auto_text_id(
    node_name: str,
    step_index: int,
    line_index: int,
    actor_alias: str,
    jp_text: str,
) -> str:
    seed = f"{node_name}|{step_index}|{line_index}|{actor_alias}|{jp_text}"
    digest = hashlib.blake2s(seed.encode("utf-8"), digest_size=6).hexdigest()
    return f"t_{digest}"


class DramaDsl:
    def __init__(self, mod_name: str, default_lang: str = "jp"):
        self.mod_name = mod_name
        self.default_lang = default_lang
        self._nodes: "OrderedDict[str, NodeSpec]" = OrderedDict()

    def reset(self) -> None:
        self._nodes.clear()

    def ref(self, name: str) -> NodeRef:
        return NodeRef(name)

    def node(self, name: str, *steps: StepSpec) -> NodeRef:
        if name in self._nodes:
            raise ValueError(f"Duplicate node name: {name}")
        normalized = [s for s in steps if isinstance(s, StepSpec)]
        self._nodes[name] = NodeSpec(ref=NodeRef(name), steps=normalized)
        return NodeRef(name)

    def story(
        self,
        start: Union[NodeRef, str],
        meta: Optional[dict[str, Any]] = None,
        nodes: Optional[list[Union[NodeSpec, NodeRef, str]]] = None,
    ) -> StorySpec:
        selected: list[NodeSpec] = []
        if nodes is None:
            selected = list(self._nodes.values())
        else:
            for n in nodes:
                if isinstance(n, NodeSpec):
                    selected.append(n)
                else:
                    key = _as_ref(n).name
                    if key not in self._nodes:
                        raise ValueError(f"Node is not registered: {key}")
                    selected.append(self._nodes[key])
        return StorySpec(start=_as_ref(start), nodes=selected, meta=meta or {})

    def id(self, kind: str, value: str) -> Id:
        return Id(kind=kind, value=value)

    def chara(self, alias: str) -> Chara:
        return Chara(alias=alias)

    def line(
        self,
        jp: str,
        actor: Union[Chara, str, None] = None,
        en: str = "",
        cn: str = "",
    ) -> StepSpec:
        return StepSpec(
            "line",
            {
                "jp": jp,
                "en": en or jp,
                "cn": cn or "",
                "actor": _actor_alias(actor),
            },
        )

    def say(
        self,
        jp: str,
        actor: Union[Chara, str],
        en: str = "",
        cn: str = "",
    ) -> StepSpec:
        return StepSpec(
            "say",
            {
                "jp": jp,
                "en": en or jp,
                "cn": cn or "",
                "actor": _actor_alias(actor),
            },
        )

    def dialog(
        self,
        lines: Optional[list[Union[StepSpec, tuple[Any, ...]]]] = None,
        prompt: Optional[StepSpec] = None,
        choices: Optional[list[OptionSpec]] = None,
        cancel: Union[NodeRef, str, None] = None,
    ) -> StepSpec:
        normalized_lines: list[StepSpec] = []
        for item in lines or []:
            if isinstance(item, StepSpec):
                normalized_lines.append(item)
                continue
            if isinstance(item, tuple):
                jp = str(item[0]) if len(item) > 0 else ""
                en = str(item[1]) if len(item) > 1 else ""
                cn = str(item[2]) if len(item) > 2 else ""
                actor = item[3] if len(item) > 3 else None
                normalized_lines.append(self.line(jp, actor=actor, en=en, cn=cn))

        return StepSpec(
            "dialog",
            {
                "lines": normalized_lines,
                "prompt": prompt,
                "choices": list(choices or []),
                "cancel": cancel,
            },
        )

    def option(
        self,
        label: str,
        to: Union[NodeRef, str],
        cond: Optional[Cond] = None,
    ) -> OptionSpec:
        return OptionSpec(label=label, to=to, cond=cond)

    def when(
        self,
        cond: Cond,
        then_to: Union[NodeRef, str],
        else_to: Union[NodeRef, str, None] = None,
    ) -> StepSpec:
        return StepSpec(
            "when",
            {"cond": cond, "then_to": then_to, "else_to": else_to},
        )

    def go(self, to: Union[NodeRef, str]) -> StepSpec:
        return StepSpec("go", {"to": to})

    def end(self) -> StepSpec:
        return StepSpec("end", {})

    def transition(
        self,
        bg: Optional[Id] = None,
        bgm: Optional[Id] = None,
        fade: float = 0.5,
        clear_bg: bool = False,
        stop_bgm: bool = False,
    ) -> StepSpec:
        return StepSpec(
            "transition",
            {
                "bg": bg,
                "bgm": bgm,
                "fade": fade,
                "clear_bg": clear_bg,
                "stop_bgm": stop_bgm,
            },
        )

    def wait(self, seconds: float) -> StepSpec:
        return StepSpec("wait", {"seconds": float(seconds)})

    def sound(self, snd: Id) -> StepSpec:
        return StepSpec("sound", {"snd": snd})

    def effect(self, fx: Id, actor: Union[Chara, str, None] = None) -> StepSpec:
        return StepSpec("effect", {"fx": fx, "actor": _actor_alias(actor)})

    def shake(self) -> StepSpec:
        return StepSpec("shake", {})

    def glitch(self) -> StepSpec:
        return StepSpec("glitch", {})

    def spawn(
        self,
        actor: Union[Chara, str],
        chara_id: Id,
        level: Optional[int] = None,
    ) -> StepSpec:
        return StepSpec(
            "spawn",
            {
                "actor": _actor_alias(actor),
                "chara_id": chara_id,
                "level": level,
            },
        )

    def move_to(self, actor: Union[Chara, str], x: int, y: int) -> StepSpec:
        return StepSpec(
            "move_to",
            {"actor": _actor_alias(actor), "x": int(x), "y": int(y)},
        )

    def move_tile(self, actor: Union[Chara, str], dx: int, dy: int) -> StepSpec:
        return StepSpec(
            "move_tile",
            {"actor": _actor_alias(actor), "dx": int(dx), "dy": int(dy)},
        )

    def move_next_to(
        self,
        actor: Union[Chara, str],
        target: Union[Chara, str],
    ) -> StepSpec:
        return StepSpec(
            "move_next_to",
            {"actor": _actor_alias(actor), "target": _actor_alias(target)},
        )

    def emote(
        self,
        actor: Union[Chara, str],
        emote_id: Id,
        duration: Optional[float] = None,
    ) -> StepSpec:
        return StepSpec(
            "emote",
            {
                "actor": _actor_alias(actor),
                "emote_id": emote_id,
                "duration": duration,
            },
        )

    def set_portrait(
        self,
        actor: Union[Chara, str],
        portrait_id: Optional[Id] = None,
    ) -> StepSpec:
        return StepSpec(
            "set_portrait",
            {"actor": _actor_alias(actor), "portrait_id": portrait_id},
        )

    def set_sprite(
        self,
        actor: Union[Chara, str],
        sprite_id: Optional[Id] = None,
    ) -> StepSpec:
        return StepSpec(
            "set_sprite",
            {"actor": _actor_alias(actor), "sprite_id": sprite_id},
        )

    def quest_begin(
        self,
        quest_id: Id,
        phase: int = 1,
        journal: bool = True,
    ) -> StepSpec:
        return StepSpec(
            "quest_begin",
            {"quest_id": quest_id, "phase": int(phase), "journal": bool(journal)},
        )

    def quest_update(
        self,
        quest_id: Id,
        phase: Optional[int] = None,
        journal: bool = False,
    ) -> StepSpec:
        return StepSpec(
            "quest_update",
            {"quest_id": quest_id, "phase": phase, "journal": bool(journal)},
        )

    def quest_finish(
        self,
        quest_id: Id,
        phase: Optional[int] = 999,
        journal: bool = True,
    ) -> StepSpec:
        return StepSpec(
            "quest_finish",
            {"quest_id": quest_id, "phase": phase, "journal": bool(journal)},
        )

    def has_flag(
        self,
        flag_id: Id,
        expr: str = ">=1",
        actor: Union[Chara, str] = "pc",
    ) -> Cond:
        return Cond("has_flag", (flag_id, expr, _actor_alias(actor)))

    def has_item(
        self,
        item_id: Id,
        expr: str = ">=1",
        actor: Union[Chara, str] = "pc",
    ) -> Cond:
        return Cond("has_item", (item_id, expr, _actor_alias(actor)))

    def has_condition(
        self,
        cond_id: Id,
        expr: str = ">=1",
        actor: Union[Chara, str] = "pc",
    ) -> Cond:
        return Cond("has_condition", (cond_id, expr, _actor_alias(actor)))

    def has_feat(
        self,
        feat_id: Id,
        expr: str = ">=1",
        actor: Union[Chara, str] = "pc",
    ) -> Cond:
        return Cond("has_feat", (feat_id, expr, _actor_alias(actor)))

    def has_keyitem(self, key_id: Id, expr: str = ">0") -> Cond:
        return Cond("has_keyitem", (key_id, expr))

    def in_zone(
        self,
        zone_id: Id,
        level: Optional[int] = None,
        actor: Union[Chara, str] = "pc",
    ) -> Cond:
        return Cond("in_zone", (zone_id, level, _actor_alias(actor)))

    def native_if(self, raw_if_expr: str) -> Cond:
        return Cond("native_if", (raw_if_expr,))

    def all_of(self, *conds: Cond) -> Cond:
        return Cond("all_of", conds)

    def any_of(self, *conds: Cond) -> Cond:
        return Cond("any_of", conds)

    def not_(self, cond: Cond) -> Cond:
        return Cond("not", (cond,))


def _ensure_id_kind(
    id_obj: Optional[Id],
    expected: str,
    step_name: str,
    field_name: str,
) -> None:
    if id_obj is None:
        return
    if not isinstance(id_obj, Id):
        raise ValueError(f"{step_name}.{field_name} must be Id, got: {type(id_obj)}")
    if id_obj.kind != expected:
        raise ValueError(
            f"{step_name}.{field_name} requires kind='{expected}', got '{id_obj.kind}'"
        )


def _cond_expr(cond: Cond) -> str:
    if cond.op == "native_if":
        return str(cond.args[0])
    if cond.op in {"has_flag", "has_item", "has_condition", "has_feat"}:
        id_obj, expr, actor = cond.args
        return f"{cond.op}({actor},{id_obj.value}){expr}"
    if cond.op == "has_keyitem":
        id_obj, expr = cond.args
        return f"has_keyitem({id_obj.value}){expr}"
    if cond.op == "in_zone":
        zone_id, level, actor = cond.args
        level_expr = f",{level}" if level is not None else ""
        return f"in_zone({actor},{zone_id.value}{level_expr})"
    if cond.op == "all_of":
        return "(" + " && ".join(_cond_expr(c) for c in cond.args) + ")"
    if cond.op == "any_of":
        return "(" + " || ".join(_cond_expr(c) for c in cond.args) + ")"
    if cond.op == "not":
        return f"!({_cond_expr(cond.args[0])})"
    return cond.op


def _target_name(target: Union[NodeRef, str, None]) -> Optional[str]:
    if target is None:
        return None
    return _as_ref(target).name


def _iter_targets(step: StepSpec) -> Iterable[str]:
    if step.kind == "go":
        name = _target_name(step.payload.get("to"))
        if name:
            yield name
    elif step.kind == "when":
        then_name = _target_name(step.payload.get("then_to"))
        else_name = _target_name(step.payload.get("else_to"))
        if then_name:
            yield then_name
        if else_name:
            yield else_name
    elif step.kind == "dialog":
        for opt in step.payload.get("choices", []):
            name = _target_name(opt.to)
            if name:
                yield name
        cancel = _target_name(step.payload.get("cancel"))
        if cancel:
            yield cancel


def _validate_story(spec: StorySpec, strict: bool) -> None:
    if not spec.nodes:
        raise ValueError("Story has no nodes")

    node_names: list[str] = [n.ref.name for n in spec.nodes]
    name_set = set(node_names)
    if len(name_set) != len(node_names):
        raise ValueError("Story contains duplicate node definitions")
    if spec.start.name not in name_set:
        raise ValueError(f"Story start node is not defined: {spec.start.name}")

    for node in spec.nodes:
        for step in node.steps:
            if step.kind == "dialog":
                choices = step.payload.get("choices")
                if choices is not None and len(choices) == 0:
                    raise ValueError(
                        f"dialog(choices=[]) is not allowed (node={node.ref.name})"
                    )
                if strict and choices and step.payload.get("prompt") is None:
                    raise ValueError(
                        "strict mode requires prompt when dialog choices are present"
                    )
            for target in _iter_targets(step):
                if target not in name_set:
                    raise ValueError(f"Undefined target node: {target}")

            if not strict:
                continue

            if step.kind == "transition":
                _ensure_id_kind(step.payload.get("bg"), "bg", "transition", "bg")
                _ensure_id_kind(step.payload.get("bgm"), "bgm", "transition", "bgm")
            elif step.kind == "sound":
                _ensure_id_kind(step.payload.get("snd"), "sound", "sound", "snd")
            elif step.kind == "effect":
                _ensure_id_kind(step.payload.get("fx"), "effect", "effect", "fx")
            elif step.kind == "spawn":
                _ensure_id_kind(
                    step.payload.get("chara_id"),
                    "chara",
                    "spawn",
                    "chara_id",
                )
            elif step.kind == "emote":
                _ensure_id_kind(
                    step.payload.get("emote_id"),
                    "emote",
                    "emote",
                    "emote_id",
                )
            elif step.kind == "set_portrait":
                _ensure_id_kind(
                    step.payload.get("portrait_id"),
                    "portrait",
                    "set_portrait",
                    "portrait_id",
                )
            elif step.kind == "set_sprite":
                _ensure_id_kind(
                    step.payload.get("sprite_id"),
                    "sprite",
                    "set_sprite",
                    "sprite_id",
                )
            elif step.kind in {"quest_begin", "quest_update", "quest_finish"}:
                _ensure_id_kind(
                    step.payload.get("quest_id"),
                    "quest",
                    step.kind,
                    "quest_id",
                )
            elif step.kind == "when":
                _validate_cond(step.payload.get("cond"))
            elif step.kind == "dialog":
                for opt in step.payload.get("choices", []):
                    if opt.cond is not None:
                        _validate_cond(opt.cond)


def _validate_cond(cond: Cond) -> None:
    if not isinstance(cond, Cond):
        raise ValueError(f"Condition must be Cond, got: {type(cond)}")
    if cond.op in {"has_flag", "has_item", "has_condition", "has_feat"}:
        id_obj, expr, _actor = cond.args
        if not isinstance(id_obj, Id):
            raise ValueError(f"{cond.op} id must be Id")
        if not isinstance(expr, str) or not expr:
            raise ValueError(f"{cond.op} expr must be non-empty string")
    elif cond.op == "has_keyitem":
        id_obj, expr = cond.args
        if not isinstance(id_obj, Id):
            raise ValueError("has_keyitem id must be Id")
        if not isinstance(expr, str) or not expr:
            raise ValueError("has_keyitem expr must be non-empty string")
    elif cond.op == "in_zone":
        zone_id, _level, _actor = cond.args
        if not isinstance(zone_id, Id):
            raise ValueError("in_zone zone_id must be Id")
    elif cond.op in {"all_of", "any_of"}:
        for c in cond.args:
            _validate_cond(c)
    elif cond.op == "not":
        _validate_cond(cond.args[0])
    elif cond.op == "native_if":
        if not isinstance(cond.args[0], str) or not cond.args[0]:
            raise ValueError("native_if expression must be non-empty string")


def _line_entry(
    *,
    node_name: str,
    step_index: int,
    line_index: int,
    jp: str,
    en: str,
    cn: str,
    actor: Optional[str],
) -> dict[str, Any]:
    entry: dict[str, Any] = {
        "id": _auto_text_id(node_name, step_index, line_index, actor or "", jp),
        "text_JP": jp,
        "text_EN": en or jp,
        "text_CN": cn or "",
    }
    if actor:
        entry["actor"] = actor
    return entry


def _compile_step(node_name: str, step_index: int, step: StepSpec) -> list[dict[str, Any]]:
    kind = step.kind
    p = step.payload

    if kind in {"line", "say"}:
        return [
            _line_entry(
                node_name=node_name,
                step_index=step_index,
                line_index=0,
                jp=p["jp"],
                en=p.get("en", ""),
                cn=p.get("cn", ""),
                actor=p.get("actor"),
            )
        ]

    if kind == "dialog":
        out: list[dict[str, Any]] = []
        line_index = 0
        for line_step in p.get("lines", []):
            lp = line_step.payload
            out.append(
                _line_entry(
                    node_name=node_name,
                    step_index=step_index,
                    line_index=line_index,
                    jp=lp["jp"],
                    en=lp.get("en", ""),
                    cn=lp.get("cn", ""),
                    actor=lp.get("actor"),
                )
            )
            line_index += 1

        prompt = p.get("prompt")
        if isinstance(prompt, StepSpec):
            pp = prompt.payload
            out.append(
                _line_entry(
                    node_name=node_name,
                    step_index=step_index,
                    line_index=line_index,
                    jp=pp["jp"],
                    en=pp.get("en", ""),
                    cn=pp.get("cn", ""),
                    actor=pp.get("actor"),
                )
            )
            line_index += 1

        for choice_idx, opt in enumerate(p.get("choices", [])):
            entry = {
                "action": "choice",
                "jump": _target_name(opt.to),
                "id": _auto_text_id(
                    node_name,
                    step_index,
                    line_index + choice_idx,
                    "",
                    f"choice:{opt.label}",
                ),
                "text_JP": opt.label,
                "text_EN": opt.label,
                "text_CN": "",
            }
            if opt.cond is not None:
                entry["if"] = _cond_expr(opt.cond)
            out.append(entry)

        cancel = _target_name(p.get("cancel"))
        if cancel:
            out.append({"action": "cancel", "jump": cancel})

        return out

    if kind == "go":
        return [{"jump": _target_name(p["to"])}]

    if kind == "when":
        out = [{"jump": _target_name(p["then_to"]), "if": _cond_expr(p["cond"])}]
        else_target = _target_name(p.get("else_to"))
        if else_target:
            out.append({"jump": else_target})
        return out

    if kind == "end":
        return [{"action": "end"}]

    if kind == "transition":
        out = []
        fade = p.get("fade")
        fade_sec = float(fade) if isinstance(fade, (int, float)) and fade > 0 else 0.0

        if p.get("clear_bg"):
            if fade_sec > 0:
                out.append({"action": "fadeOut", "param": f"{fade_sec},black"})
            out.append({"action": "setBG", "param": ""})
        elif isinstance(p.get("bg"), Id):
            out.append({"action": "setBG", "param": p["bg"].value})
            if fade_sec > 0:
                out.append({"action": "fadeIn", "param": str(fade_sec)})

        if p.get("stop_bgm"):
            out.append({"action": "stopBGM"})

        if isinstance(p.get("bgm"), Id):
            bgm_id = p["bgm"].value
            out.append(
                {
                    "action": "eval",
                    "param": (
                        f'var data = SoundManager.current.GetData("{bgm_id}"); '
                        "if (data is BGMData bgm) { "
                        "LayerDrama.haltPlaylist = true; "
                        "LayerDrama.maxBGMVolume = true; "
                        "SoundManager.current.PlayBGM(bgm); }"
                    ),
                }
            )

        return out

    if kind == "wait":
        return [{"action": "wait", "param": str(p["seconds"])}]

    if kind == "sound":
        return [{"action": "sound", "param": p["snd"].value}]

    if kind == "effect":
        entry = {"action": "invoke*", "param": f"play_effect_ext({p['fx'].value})"}
        if p.get("actor"):
            entry["actor"] = p["actor"]
        return [entry]

    if kind == "shake":
        return [{"action": "shake"}]

    if kind == "glitch":
        return [{"action": "glitch"}]

    if kind == "spawn":
        actor_alias = p["actor"]
        chara_id = p["chara_id"].value
        level = p.get("level")
        level_code = str(level) if level is not None else "1"
        code = (
            f'var spawned = CharaGen.Create("{chara_id}", {level_code}); '
            "if (spawned != null) { spawned.Spawn(Chara.FromFilter(EClass.pc).pos); "
            "var drama = LayerDrama.GetDrama(); "
            "if (drama != null && drama.sequence != null) { "
            f'drama.sequence.AddActor("{actor_alias}", new Person(spawned)); '
            "} }"
        )
        return [{"action": "eval", "param": code}]

    if kind == "move_to":
        entry = {
            "action": "invoke*",
            "param": f"move_to({p['x']}, {p['y']})",
        }
        if p.get("actor"):
            entry["actor"] = p["actor"]
        return [entry]

    if kind == "move_tile":
        entry = {
            "action": "invoke*",
            "param": f"move_tile({p['dx']}, {p['dy']})",
        }
        if p.get("actor"):
            entry["actor"] = p["actor"]
        return [entry]

    if kind == "move_next_to":
        entry = {"action": "invoke*", "param": f"move_next_to({p['target']})"}
        if p.get("actor"):
            entry["actor"] = p["actor"]
        return [entry]

    if kind == "emote":
        emote_id = p["emote_id"].value
        duration = p.get("duration")
        if duration is None:
            param = f"play_emote({emote_id})"
        else:
            param = f"play_emote({emote_id}, {duration})"
        entry = {"action": "invoke*", "param": param}
        if p.get("actor"):
            entry["actor"] = p["actor"]
        return [entry]

    if kind == "set_portrait":
        portrait_id = p.get("portrait_id")
        param = "set_portrait()" if portrait_id is None else f"set_portrait({portrait_id.value})"
        entry = {"action": "invoke*", "param": param}
        if p.get("actor"):
            entry["actor"] = p["actor"]
        return [entry]

    if kind == "set_sprite":
        sprite_id = p.get("sprite_id")
        param = "set_sprite()" if sprite_id is None else f"set_sprite({sprite_id.value})"
        entry = {"action": "invoke*", "param": param}
        if p.get("actor"):
            entry["actor"] = p["actor"]
        return [entry]

    if kind == "quest_begin":
        quest_id = p["quest_id"].value
        out = [
            {"action": "startQuest", "param": quest_id},
            {"action": "changePhase", "param": f"{quest_id},{p['phase']}"},
        ]
        if p.get("journal"):
            out.append({"action": "updateJournal"})
        return out

    if kind == "quest_update":
        quest_id = p["quest_id"].value
        out = []
        if p.get("phase") is not None:
            out.append({"action": "changePhase", "param": f"{quest_id},{p['phase']}"})
        if p.get("journal"):
            out.append({"action": "updateJournal"})
        return out

    if kind == "quest_finish":
        quest_id = p["quest_id"].value
        out = []
        if p.get("phase") is not None:
            out.append({"action": "changePhase", "param": f"{quest_id},{p['phase']}"})
        out.append({"action": "completeQuest", "param": quest_id})
        if p.get("journal"):
            out.append({"action": "updateJournal"})
        return out

    raise ValueError(f"Unsupported step kind: {kind}")


def compile_xlsx(spec: StorySpec, strict: bool = True) -> openpyxl.Workbook:
    _validate_story(spec, strict=strict)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(spec.meta.get("sheet_name", "main"))

    for col, header in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=header)

    row = 6
    for node in spec.nodes:
        ws.cell(row=row, column=1, value=node.ref.name)
        row += 1

        for step_index, step in enumerate(node.steps):
            entries = _compile_step(node.ref.name, step_index, step)
            for entry in entries:
                for col, header in enumerate(HEADERS, 1):
                    value = entry.get(header)
                    if header == "text" and value is None:
                        value = entry.get("text_EN")
                    if value is not None:
                        ws.cell(row=row, column=col, value=value)
                row += 1

    return wb


def save_xlsx(workbook: openpyxl.Workbook, path: str, sheet: str = "main") -> None:
    if workbook.active is not None and sheet:
        workbook.active.title = sheet
    os.makedirs(os.path.dirname(path), exist_ok=True)
    workbook.save(path)


__all__ = [
    "Chara",
    "Cond",
    "DramaDsl",
    "HEADERS",
    "Id",
    "NodeRef",
    "NodeSpec",
    "OptionSpec",
    "StepSpec",
    "StorySpec",
    "compile_xlsx",
    "save_xlsx",
]
