import tempfile
import unittest
from pathlib import Path

import openpyxl

from tools.drama_v2.drama_dsl import DramaDsl, compile_xlsx, save_xlsx


class DramaDslV51Tests(unittest.TestCase):
    def test_story_auto_collects_nodes_in_registration_order(self):
        d = DramaDsl(mod_name="QuestMod")
        d.node("main", d.go("end"))
        d.node("end", d.end())

        spec = d.story(start="main")
        self.assertEqual("main", spec.start.name)
        self.assertEqual(["main", "end"], [n.ref.name for n in spec.nodes])

    def test_story_resolves_forward_targets(self):
        d = DramaDsl(mod_name="QuestMod")
        d.node("main", d.go("later"))
        d.node("later", d.end())

        wb = compile_xlsx(d.story(start="main"), strict=True)
        ws = wb.active

        # Row 6: step main, Row 7: jump to later
        self.assertEqual("main", ws.cell(row=6, column=1).value)
        self.assertEqual("later", ws.cell(row=7, column=2).value)

    def test_compile_rejects_undefined_target(self):
        d = DramaDsl(mod_name="QuestMod")
        d.node("main", d.go("missing"))

        with self.assertRaises(ValueError):
            compile_xlsx(d.story(start="main"), strict=True)

    def test_duplicate_node_name_is_rejected(self):
        d = DramaDsl(mod_name="QuestMod")
        d.node("main", d.end())
        with self.assertRaises(ValueError):
            d.node("main", d.end())

    def test_dialog_with_empty_choices_is_rejected(self):
        d = DramaDsl(mod_name="QuestMod")
        d.node("main", d.dialog(prompt=d.line("選んでください"), choices=[]))
        with self.assertRaises(ValueError):
            compile_xlsx(d.story(start="main"), strict=True)

    def test_text_ids_are_auto_generated_and_stable(self):
        d = DramaDsl(mod_name="QuestMod")
        pc = d.chara("pc")
        end_ref = d.node("end", d.end())
        d.node(
            "main",
            d.line("はじめまして。", actor=pc),
            d.say("続けます。", actor=pc),
            d.dialog(
                prompt=d.line("行動を選択", actor=pc),
                choices=[d.option("終了", to=end_ref)],
                cancel=end_ref,
            ),
        )
        spec = d.story(start="main")

        wb1 = compile_xlsx(spec, strict=True)
        wb2 = compile_xlsx(spec, strict=True)
        ws1 = wb1.active
        ws2 = wb2.active

        ids_1 = []
        ids_2 = []
        for row in range(6, ws1.max_row + 1):
            v1 = ws1.cell(row=row, column=9).value
            v2 = ws2.cell(row=row, column=9).value
            if isinstance(v1, str) and v1:
                ids_1.append(v1)
            if isinstance(v2, str) and v2:
                ids_2.append(v2)

        self.assertGreaterEqual(len(ids_1), 4)  # line/say/prompt/choice
        self.assertEqual(ids_1, ids_2)

    def test_strict_rejects_kind_mismatch(self):
        d = DramaDsl(mod_name="QuestMod")
        wrong_bg_id = d.id("sound", "base.ok")
        d.node("main", d.transition(bg=wrong_bg_id), d.end())
        with self.assertRaises(ValueError):
            compile_xlsx(d.story(start="main"), strict=True)

    def test_save_writes_headers_and_starts_from_row_6(self):
        d = DramaDsl(mod_name="QuestMod")
        d.node("main", d.say("テスト", actor=d.chara("pc")), d.end())
        wb = compile_xlsx(d.story(start="main"), strict=True)

        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "drama_v5_1.xlsx"
            save_xlsx(wb, str(out))
            self.assertTrue(out.exists())

            saved = openpyxl.load_workbook(out)
            ws = saved.active
            self.assertEqual("step", ws.cell(row=1, column=1).value)
            self.assertEqual("jump", ws.cell(row=1, column=2).value)
            self.assertEqual("id", ws.cell(row=1, column=9).value)
            self.assertEqual("main", ws.cell(row=6, column=1).value)

    def test_reset_clears_registered_nodes(self):
        d = DramaDsl(mod_name="QuestMod")
        d.node("main", d.end())
        d.reset()
        d.node("intro", d.end())

        spec = d.story(start="intro")
        self.assertEqual(["intro"], [n.ref.name for n in spec.nodes])

    def test_quest_lifecycle_compiles_to_cwl_actions(self):
        d = DramaDsl(mod_name="QuestMod")
        q = d.id("quest", "quest_demo")
        d.node(
            "main",
            d.quest_begin(q, phase=1, journal=True),
            d.quest_update(q, phase=2, journal=False),
            d.quest_finish(q, phase=999, journal=True),
            d.end(),
        )
        wb = compile_xlsx(d.story(start="main"), strict=True)
        ws = wb.active

        rows = []
        for row in range(6, ws.max_row + 1):
            action = ws.cell(row=row, column=5).value
            param = ws.cell(row=row, column=6).value
            if action is not None:
                rows.append((action, param))

        self.assertIn(("startQuest", "quest_demo"), rows)
        self.assertIn(("changePhase", "quest_demo,1"), rows)
        self.assertIn(("changePhase", "quest_demo,2"), rows)
        self.assertIn(("completeQuest", "quest_demo"), rows)
        self.assertIn(("updateJournal", None), rows)

    def test_when_compiles_conditional_jump_and_fallback(self):
        d = DramaDsl(mod_name="QuestMod")
        cond = d.has_flag(d.id("flag", "tmp.ready"), expr=">=1")
        d.node("main", d.when(cond, then_to="ok", else_to="ng"))
        d.node("ok", d.end())
        d.node("ng", d.end())

        wb = compile_xlsx(d.story(start="main"), strict=True)
        ws = wb.active

        self.assertEqual("ok", ws.cell(row=7, column=2).value)
        self.assertEqual("has_flag(pc,tmp.ready)>=1", ws.cell(row=7, column=3).value)
        self.assertEqual("ng", ws.cell(row=8, column=2).value)


if __name__ == "__main__":
    unittest.main()
