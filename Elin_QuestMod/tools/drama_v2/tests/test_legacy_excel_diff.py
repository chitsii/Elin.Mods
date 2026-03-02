import tempfile
import unittest
from pathlib import Path

import openpyxl

from tools.drama.drama_builder import DramaBuilder as LegacyDramaBuilder
from tools.drama.scenarios.quest_drama_placeholder import define_quest_drama_placeholder
from tools.drama_v2.drama_builder import DramaBuilder, compile_xlsx, save_xlsx


HEADERS = [
    "step",
    "jump",
    "if",
    "if2",
    "action",
    "param",
    "actor",
    "version",
    "id",
    "text_JP",
    "text_EN",
    "text_CN",
    "text",
]


def _read_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row_idx in range(6, ws.max_row + 1):
        row = {}
        for col_idx, header in enumerate(HEADERS, start=1):
            row[header] = ws.cell(row=row_idx, column=col_idx).value
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows


class LegacyExcelDiffTests(unittest.TestCase):
    def test_placeholder_output_matches_legacy_except_text_id(self):
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            legacy_path = td_path / "legacy.xlsx"
            v2_path = td_path / "v2.xlsx"

            legacy_builder = LegacyDramaBuilder(mod_name="QuestMod")
            define_quest_drama_placeholder(legacy_builder)
            legacy_builder.save(str(legacy_path), sheet_name="quest_drama_replace_me")

            d = DramaBuilder(mod_name="QuestMod")
            narrator = d.chara("narrator")
            d.node(
                "main",
                d.say(
                    "QuestMod placeholder drama. Replace this scenario with your own quest flow.",
                    en="QuestMod placeholder drama. Replace this scenario with your own quest flow.",
                    actor=narrator,
                ),
                d.dialog(choices=[d.option("Continue", to="end")]),
            )
            d.node(
                "end",
                d.say("Template check complete.", en="Template check complete.", actor=narrator),
                d.transition(clear_bg=True, fade=0.3),
                d.end(),
            )
            wb = compile_xlsx(
                d.story(start="main", meta={"sheet_name": "quest_drama_replace_me"}),
                strict=False,
            )
            save_xlsx(wb, str(v2_path), sheet="quest_drama_replace_me")

            legacy_rows = _read_rows(legacy_path)
            v2_rows = _read_rows(v2_path)
            self.assertEqual(len(legacy_rows), len(v2_rows))

            mismatches = []
            for idx, (legacy_row, v2_row) in enumerate(zip(legacy_rows, v2_rows), start=1):
                legacy_cmp = {k: v for k, v in legacy_row.items() if k != "id"}
                v2_cmp = {k: v for k, v in v2_row.items() if k != "id"}
                if legacy_cmp != v2_cmp:
                    mismatches.append((idx, legacy_cmp, v2_cmp))

            self.assertEqual(
                [],
                mismatches,
                f"rows mismatch (except id): {mismatches}",
            )

            # v2 must use auto-generated ids (not legacy hand-written ids).
            legacy_ids = [row["id"] for row in legacy_rows if row["id"]]
            v2_ids = [row["id"] for row in v2_rows if row["id"]]
            self.assertNotEqual(legacy_ids, v2_ids)
            self.assertTrue(all(str(i).startswith("t_") for i in v2_ids))


if __name__ == "__main__":
    unittest.main()

