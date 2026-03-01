#!/usr/bin/env python3
"""
リファクタリング検証スクリプト

ゴールデンデータと現在の出力を比較し、リファクタリングが
既存の動作を壊していないことを確認する。

使用方法:
    cd tests
    uv run python verify_refactoring.py
"""

import json
import os
import sys
import difflib
import hashlib
import subprocess
from pathlib import Path

# パス設定
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
GOLDEN_DIR = SCRIPT_DIR / "golden"
TOOLS_DIR = PROJECT_ROOT / "tools"
TOOLS_BUILDER = PROJECT_ROOT / "tools" / "builder"
DRAMA_OUTPUT_DIR = PROJECT_ROOT / "LangMod" / "JP" / "Dialog" / "Drama"

# 比較対象ファイル
COMPARISONS = [
    ("Package/quest_definitions.json", "quest_definitions.json"),
    ("src/Generated/ArenaFlags.cs", "ArenaFlags.cs"),
    ("src/Generated/ArenaQuestData.cs", "ArenaQuestData.cs"),
]

def compare_files(current_path: Path, golden_path: Path) -> tuple[bool, str]:
    """2つのファイルを比較し、差分を返す"""
    if not current_path.exists():
        return False, f"Current file not found: {current_path}"
    if not golden_path.exists():
        return False, f"Golden file not found: {golden_path}"

    with open(current_path, 'r', encoding='utf-8') as f:
        current = f.readlines()
    with open(golden_path, 'r', encoding='utf-8') as f:
        golden = f.readlines()

    if current == golden:
        return True, "Files are identical"

    diff = list(difflib.unified_diff(
        golden, current,
        fromfile=f"golden/{golden_path.name}",
        tofile=str(current_path.relative_to(PROJECT_ROOT)),
        lineterm=''
    ))
    return False, '\n'.join(diff[:50])  # 最初の50行のみ表示


def compare_json(current_path: Path, golden_path: Path) -> tuple[bool, str]:
    """JSONファイルを構造的に比較"""
    if not current_path.exists():
        return False, f"Current file not found: {current_path}"
    if not golden_path.exists():
        return False, f"Golden file not found: {golden_path}"

    with open(current_path, 'r', encoding='utf-8') as f:
        current = json.load(f)
    with open(golden_path, 'r', encoding='utf-8') as f:
        golden = json.load(f)

    if current == golden:
        return True, "JSON content is identical"

    # 差分の詳細を生成
    current_str = json.dumps(current, indent=2, ensure_ascii=False, sort_keys=True)
    golden_str = json.dumps(golden, indent=2, ensure_ascii=False, sort_keys=True)

    diff = list(difflib.unified_diff(
        golden_str.splitlines(keepends=True),
        current_str.splitlines(keepends=True),
        fromfile=f"golden/{golden_path.name}",
        tofile=str(current_path.relative_to(PROJECT_ROOT)),
    ))
    return False, ''.join(diff[:100])


def extract_excel_content(excel_path: Path, skip_metadata_columns: bool = True) -> list[str]:
    """
    ExcelファイルからCSV形式の内容を抽出

    Args:
        excel_path: Excelファイルパス
        skip_metadata_columns: True の場合、version列などメタデータをスキップ

    Returns:
        行のリスト
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        content_parts = []

        # CWLドラマ形式のヘッダー列インデックス（0-based）
        # step, jump, if, if2, action, param, actor, version, id, text_JP, text_EN, text
        # version列（インデックス7）はメタデータなのでスキップ対象
        skip_columns = {7} if skip_metadata_columns else set()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            content_parts.append(f"=== SHEET: {sheet_name} ===")
            for row_idx, row in enumerate(ws.iter_rows(), 1):
                row_values = []
                for col_idx, cell in enumerate(row):
                    if col_idx in skip_columns:
                        continue
                    row_values.append(str(cell.value) if cell.value is not None else "")
                # 空行をスキップ
                if any(v for v in row_values):
                    content_parts.append(f"{row_idx}|" + "|".join(row_values))

        return content_parts
    except Exception as e:
        return [f"ERROR: {e}"]


def hash_excel_content(excel_path: Path) -> str:
    """Excelファイルの内容をハッシュ化"""
    content = "\n".join(extract_excel_content(excel_path))
    return hashlib.sha256(content.encode('utf-8')).hexdigest()[:16]


def compare_excel_content(current_path: Path, golden_lines: list[str]) -> tuple[bool, list[str]]:
    """
    Excelファイルの内容をゴールデンデータと比較

    Returns:
        (is_identical, diff_lines)
    """
    current_lines = extract_excel_content(current_path)

    if current_lines == golden_lines:
        return True, []

    # 差分を計算
    diff = list(difflib.unified_diff(
        golden_lines,
        current_lines,
        fromfile="golden",
        tofile="current",
        lineterm=""
    ))
    return False, diff


def generate_drama_excels() -> bool:
    """ドラマExcelファイルを生成"""
    create_script = TOOLS_BUILDER / "create_drama_excel.py"
    if not create_script.exists():
        print(f"  [SKIP] Drama generator not found: {create_script}")
        return False

    try:
        result = subprocess.run(
            ["uv", "run", "python", str(create_script)],
            cwd=str(PROJECT_ROOT / "tools"),
            capture_output=True,
            text=True,
            timeout=120
        )
        return result.returncode == 0
    except Exception as e:
        print(f"  [ERROR] Failed to generate drama files: {e}")
        return False


def verify_drama_excels() -> tuple[bool, list[str]]:
    """
    ドラマExcelファイルを検証（コンテンツ比較）
    全26ファイルを対象とし、詳細な比較結果をログファイルに出力

    Returns:
        (all_passed, messages)
    """
    golden_content_path = GOLDEN_DIR / "drama_content.json"
    messages = []
    log_lines = []

    if not golden_content_path.exists():
        messages.append(f"Golden content file not found: {golden_content_path}")
        messages.append("Run with --update-golden-drama to create it")
        return False, messages

    with open(golden_content_path, 'r', encoding='utf-8') as f:
        golden_content = json.load(f)

    # 全ドラマファイルを対象にする（ゴールデンデータに含まれる全ファイル）
    all_drama_files = sorted(golden_content.keys())
    log_lines.append(f"Drama Excel Verification Report")
    log_lines.append(f"=" * 60)
    log_lines.append(f"Total files to check: {len(all_drama_files)}")
    log_lines.append(f"")

    all_passed = True
    passed_count = 0
    diff_count = 0
    missing_count = 0

    for filename in all_drama_files:
        excel_path = DRAMA_OUTPUT_DIR / filename
        if not excel_path.exists():
            messages.append(f"[MISS] {filename}")
            log_lines.append(f"[MISS] {filename} - file not found")
            all_passed = False
            missing_count += 1
            continue

        golden_lines = golden_content.get(filename)
        if golden_lines is None:
            messages.append(f"[NEW] {filename}")
            log_lines.append(f"[NEW] {filename} - not in golden data")
            all_passed = False
            continue

        is_identical, diff = compare_excel_content(excel_path, golden_lines)

        if is_identical:
            messages.append(f"[PASS] {filename}")
            log_lines.append(f"[PASS] {filename}")
            passed_count += 1
        else:
            messages.append(f"[DIFF] {filename} ({len(diff)} lines)")
            log_lines.append(f"")
            log_lines.append(f"[DIFF] {filename}")
            log_lines.append(f"-" * 40)
            # 差分の全行を出力（制限なし）
            log_lines.extend(diff)
            log_lines.append(f"")
            diff_count += 1

    # サマリー
    log_lines.append(f"")
    log_lines.append(f"=" * 60)
    log_lines.append(f"Summary:")
    log_lines.append(f"  PASS: {passed_count}/{len(all_drama_files)}")
    log_lines.append(f"  DIFF: {diff_count}/{len(all_drama_files)}")
    log_lines.append(f"  MISS: {missing_count}/{len(all_drama_files)}")

    messages.append(f"Summary: PASS={passed_count}, DIFF={diff_count}, MISS={missing_count}")

    # 常にログファイルに出力
    log_path = SCRIPT_DIR / "drama_verification.log"
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("\n".join(log_lines))
    messages.append(f"Full report: {log_path}")

    return all_passed, messages


def update_golden_drama_content():
    """ゴールデンドラマコンテンツを更新（行単位で保存）"""
    drama_files = list(DRAMA_OUTPUT_DIR.glob("*.xlsx"))

    content = {}
    for excel_path in sorted(drama_files):
        content[excel_path.name] = extract_excel_content(excel_path)

    golden_content_path = GOLDEN_DIR / "drama_content.json"
    with open(golden_content_path, 'w', encoding='utf-8') as f:
        json.dump(content, f, indent=2, ensure_ascii=False)

    print(f"Updated: {golden_content_path}")
    print(f"  Total files: {len(content)}")


def regenerate_simulation() -> dict:
    """クエストシミュレーションを再生成"""
    sys.path.insert(0, str(TOOLS_DIR))

    from arena.builders import get_quest_graph
    from arena.data import Keys, Phase

    graph = get_quest_graph()

    results = {
        'quest_chains': {},
        'phase_availability': {},
        'npc_quests': {},
        'auto_trigger_quests': {},
    }

    # 1. 各クエストへのチェーン
    key_quests = ['18_last_battle', '17_vs_astaroth', '16_lily_real_name', '15_vs_balgas']
    for quest_id in key_quests:
        results['quest_chains'][quest_id] = graph.get_quest_chain(quest_id)

    # 2. 各フェーズでの利用可能クエスト
    phase_progression = [
        (Phase.PROLOGUE, set()),
        (Phase.INITIATION, {'01_opening', '02_rank_up_G'}),
        (Phase.RISING, {'01_opening', '02_rank_up_G', '04_rank_up_F', '03_zek_intro', '05_1_lily_experiment', '05_2_zek_steal_bottle', '05_2_zek_steal_bottle_refuse', '06_rank_up_E', '06_2_zek_steal_soulgem', '06_2_zek_steal_soulgem_return'}),
        (Phase.AWAKENING, {'01_opening', '02_rank_up_G', '04_rank_up_F', '03_zek_intro', '05_1_lily_experiment', '05_2_zek_steal_bottle', '05_2_zek_steal_bottle_refuse', '06_rank_up_E', '06_2_zek_steal_soulgem', '06_2_zek_steal_soulgem_return', '07_upper_existence', '10_rank_up_D', '08_lily_private', '09_balgas_training', '09_rank_up_C', '11_rank_up_B'}),
        (Phase.CONFRONTATION, {'01_opening', '02_rank_up_G', '04_rank_up_F', '03_zek_intro', '05_1_lily_experiment', '05_2_zek_steal_bottle', '05_2_zek_steal_bottle_refuse', '06_rank_up_E', '06_2_zek_steal_soulgem', '06_2_zek_steal_soulgem_return', '07_upper_existence', '10_rank_up_D', '08_lily_private', '09_balgas_training', '09_rank_up_C', '11_rank_up_B', '12_makuma', '13_makuma2', '12_rank_up_A', '15_vs_balgas', '15_vs_balgas_spared', '16_lily_real_name'}),
    ]

    for phase, completed_quests in phase_progression:
        flags = {Keys.CURRENT_PHASE: list(Phase).index(phase)}
        available = graph.get_available_quests(flags, completed_quests)
        results['phase_availability'][phase.value] = [q.quest_id for q in available]

    # 3. NPC別クエスト
    npc_test_completed = {'01_opening', '02_rank_up_G', '04_rank_up_F', '06_rank_up_E', '10_rank_up_D', '09_rank_up_C', '11_rank_up_B'}
    npc_flags = {Keys.CURRENT_PHASE: 4}
    npc_quests = graph.get_all_npc_quests(npc_flags, npc_test_completed)
    for npc_id, quests in npc_quests.items():
        results['npc_quests'][npc_id] = [q.quest_id for q in quests]

    # 4. 自動発動クエスト
    for phase, completed_quests in phase_progression:
        flags = {Keys.CURRENT_PHASE: list(Phase).index(phase)}
        auto = graph.get_auto_trigger_quests(flags, completed_quests)
        if auto:
            results['auto_trigger_quests'][phase.value] = [q.quest_id for q in auto]

    return results


def main():
    print("=" * 60)
    print("Quest System Refactoring Verification")
    print("=" * 60)

    all_passed = True

    # 1. ファイル比較
    print("\n[1] File Comparisons")
    print("-" * 40)

    for current_rel, golden_name in COMPARISONS:
        current_path = PROJECT_ROOT / current_rel
        golden_path = GOLDEN_DIR / golden_name

        if golden_name.endswith('.json'):
            passed, message = compare_json(current_path, golden_path)
        else:
            passed, message = compare_files(current_path, golden_path)

        status = "PASS" if passed else "FAIL"
        print(f"  [{status}] {current_rel}")

        if not passed:
            all_passed = False
            print(f"        {message[:200]}...")

    # 2. クエストシミュレーション比較
    print("\n[2] Quest Simulation Comparison")
    print("-" * 40)

    golden_sim_path = GOLDEN_DIR / "quest_simulation.json"
    if golden_sim_path.exists():
        with open(golden_sim_path, 'r', encoding='utf-8') as f:
            golden_sim = json.load(f)

        try:
            current_sim = regenerate_simulation()

            if current_sim == golden_sim:
                print("  [PASS] Quest simulation matches golden data")
            else:
                all_passed = False
                print("  [FAIL] Quest simulation differs from golden data")

                # 差分の詳細
                for key in set(golden_sim.keys()) | set(current_sim.keys()):
                    if golden_sim.get(key) != current_sim.get(key):
                        print(f"        Difference in '{key}':")
                        print(f"          Golden: {json.dumps(golden_sim.get(key), ensure_ascii=False)[:100]}")
                        print(f"          Current: {json.dumps(current_sim.get(key), ensure_ascii=False)[:100]}")
        except Exception as e:
            all_passed = False
            print(f"  [FAIL] Error running simulation: {e}")
    else:
        print("  [SKIP] Golden simulation file not found")

    # 3. 依存関係グラフ比較
    print("\n[3] Dependency Graph Comparison")
    print("-" * 40)

    golden_graph_path = GOLDEN_DIR / "quest_graph.dot"
    if golden_graph_path.exists():
        try:
            sys.path.insert(0, str(TOOLS_DIR))
            from arena.builders import get_quest_graph

            graph = get_quest_graph()
            current_graph = graph.generate_quest_graph_viz()

            with open(golden_graph_path, 'r', encoding='utf-8') as f:
                golden_graph = f.read()

            if current_graph.strip() == golden_graph.strip():
                print("  [PASS] Dependency graph matches golden data")
            else:
                all_passed = False
                print("  [FAIL] Dependency graph differs from golden data")
        except Exception as e:
            all_passed = False
            print(f"  [FAIL] Error generating graph: {e}")
    else:
        print("  [SKIP] Golden graph file not found")

    # 4. ドラマExcel検証
    print("\n[4] Drama Excel Verification")
    print("-" * 40)

    golden_drama_path = GOLDEN_DIR / "drama_hashes.json"
    if golden_drama_path.exists():
        passed, messages = verify_drama_excels()
        for msg in messages:
            print(f"  {msg}")
        if not passed:
            all_passed = False
    else:
        print("  [SKIP] Golden drama hashes not found")
        print("  Run: uv run python verify_refactoring.py --update-golden-drama")

    # 結果サマリー
    print("\n" + "=" * 60)
    if all_passed:
        print("RESULT: ALL TESTS PASSED")
        print("Refactoring is safe to proceed.")
        return 0
    else:
        print("RESULT: SOME TESTS FAILED")
        print("Please review the differences before proceeding.")
        return 1


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Verify refactoring against golden data")
    parser.add_argument("--update-golden-drama", action="store_true",
                        help="Update golden drama content from current files")

    args = parser.parse_args()

    if args.update_golden_drama:
        update_golden_drama_content()
        sys.exit(0)

    sys.exit(main())
