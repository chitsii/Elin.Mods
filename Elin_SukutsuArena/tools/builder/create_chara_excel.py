# -*- coding: utf-8 -*-
"""
create_chara_excel.py - SourceChara.xlsx 自動生成

カスタムキャラクター（ボス、NPC等）をCWL形式のExcelファイルとして生成する。
バニラと同じ50カラム構造 + CWL拡張カラム（Author, portrait）を使用。
EN版とCN版の両方を生成。
"""

import argparse
import csv
import os
import re
import sys

import openpyxl

# コマンドライン引数の解析
parser = argparse.ArgumentParser(description="Create Chara Excel for Elin_SukutsuArena")
parser.add_argument(
    "--debug", action="store_true", help="Debug mode: enable debug-only entries"
)
parser.add_argument(
    "--boss-lv1", action="store_true", help="Set all boss characters to LV=1"
)
args = parser.parse_args()

DEBUG_MODE = args.debug
if DEBUG_MODE:
    print("[DEBUG MODE] Debug-only entries enabled")

BOSS_LV1 = args.boss_lv1
if BOSS_LV1:
    print("[DEBUG MODE] All boss characters will have LV=1 (--boss-lv1)")

# パス設定
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)
OUTPUT_EN_TSV = os.path.join(PROJECT_ROOT, "LangMod", "EN", "Chara.tsv")
OUTPUT_CN_TSV = os.path.join(PROJECT_ROOT, "LangMod", "CN", "Chara.tsv")
VANILLA_SOURCE_GAME_XLSX = os.path.join(
    PROJECT_ROOT, "..", "SourceExcels", "SourceGame.xlsx"
)
MOD_SOURCE_ELEMENT_XLSX = os.path.join(
    PROJECT_ROOT, "LangMod", "EN", "SourceElement.xlsx"
)
VANILLA_SOURCE_CHARA_XLSX = os.path.join(
    PROJECT_ROOT, "..", "SourceExcels", "SourceChara.xlsx"
)
VANILLA_CHARA_TONE_XLSX = os.path.join(
    PROJECT_ROOT, "elin_link", "Package", "_Elona", "Lang", "JP", "Data", "chara_tone.xlsx"
)

# SourceChara カラム定義（バニラと同じ50カラム + CWL拡張2カラム）
HEADERS = [
    "id",  # 0
    "_id",  # 1
    "name_JP",  # 2
    "name",  # 3
    "aka_JP",  # 4
    "aka",  # 5
    "idActor",  # 6
    "sort",  # 7
    "size",  # 8
    "_idRenderData",  # 9
    "tiles",  # 10
    "tiles_snow",  # 11
    "colorMod",  # 12
    "components",  # 13
    "defMat",  # 14
    "LV",  # 15
    "chance",  # 16
    "quality",  # 17
    "hostility",  # 18
    "biome",  # 19
    "tag",  # 20
    "trait",  # 21
    "race",  # 22
    "job",  # 23
    "tactics",  # 24
    "aiIdle",  # 25
    "aiParam",  # 26
    "actCombat",  # 27
    "mainElement",  # 28
    "elements",  # 29
    "equip",  # 30
    "loot",  # 31
    "category",  # 32
    "filter",  # 33
    "gachaFilter",  # 34
    "tone",  # 35
    "actIdle",  # 36
    "lightData",  # 37
    "idExtra",  # 38
    "bio",  # 39
    "faith",  # 40
    "works",  # 41
    "hobbies",  # 42
    "idText",  # 43
    "moveAnime",  # 44
    "factory",  # 45
    "components",  # 46 (重複)
    "recruitItems",  # 47
    "detail_JP",  # 48
    "detail",  # 49
    # CWL拡張カラム
    "Author",  # 50
    "portrait",  # 51
]

# 型情報（2行目）- バニラに合わせる
TYPES = [
    "string",  # id
    "int",  # _id
    "string",  # name_JP
    "string",  # name
    "string",  # aka_JP
    "string",  # aka
    "string[]",  # idActor
    "int",  # sort
    "int[]",  # size
    "string",  # _idRenderData
    "int[]",  # tiles
    "int[]",  # tiles_snow
    "int",  # colorMod
    "string[]",  # components
    "string",  # defMat
    "int",  # LV
    "int",  # chance
    "int",  # quality
    "string",  # hostility
    "string",  # biome
    "string[]",  # tag
    "string[]",  # trait
    "string",  # race
    "string",  # job
    "string",  # tactics
    "string",  # aiIdle
    "int[]",  # aiParam
    "string[]",  # actCombat
    "string[]",  # mainElement
    "elements",  # elements
    "string",  # equip
    "string[]",  # loot
    "string",  # category
    "string[]",  # filter
    "string[]",  # gachaFilter
    "string",  # tone
    "string[]",  # actIdle
    "string",  # lightData
    "string",  # idExtra
    "string",  # bio
    "string",  # faith
    "string[]",  # works
    "string[]",  # hobbies
    "string",  # idText
    "string",  # moveAnime
    "string[]",  # factory
    "string[]",  # components (重複)
    "string[]",  # recruitItems
    "string",  # detail_JP
    "string",  # detail
    # CWL拡張カラム
    "string",  # Author
    "string",  # portrait
]

# デフォルト値（3行目）- バニラに合わせる
# 同名カラム(components)があるためリストで定義
DEFAULTS = [
    "",  # id (0)
    "552",  # _id (1)
    "",  # name_JP (2)
    "",  # name (3)
    "",  # aka_JP (4)
    "",  # aka (5)
    "",  # idActor (6)
    "",  # sort (7)
    "",  # size (8)
    "chara",  # _idRenderData (9)
    "0",  # tiles (10)
    "",  # tiles_snow (11)
    "0",  # colorMod (12)
    "log/1",  # components (13)
    "",  # defMat (14)
    "1",  # LV (15)
    "100",  # chance (16)
    "",  # quality (17)
    "",  # hostility (18)
    "",  # biome (19)
    "",  # tag (20)
    "",  # trait (21)
    "norland",  # race (22)
    "none",  # job (23)
    "",  # tactics (24)
    "",  # aiIdle (25)
    "",  # aiParam (26)
    "",  # actCombat (27)
    "",  # mainElement (28)
    "",  # elements (29)
    "",  # equip (30)
    "",  # loot (31)
    "chara",  # category (32)
    "",  # filter (33)
    "",  # gachaFilter (34)
    "",  # tone (35)
    "",  # actIdle (36)
    "",  # lightData (37)
    "",  # idExtra (38)
    "",  # bio (39)
    "",  # faith (40)
    "",  # works (41)
    "",  # hobbies (42)
    "",  # idText (43)
    "",  # moveAnime (44)
    "",  # factory (45)
    "",  # components (46) - 空（バニラと同じ）
    "",  # recruitItems (47)
    "",  # detail_JP (48)
    "",  # detail (49)
    # CWL拡張カラム
    "",  # Author (50)
    "",  # portrait (51)
]

# カラム数
NUM_COLUMNS = len(HEADERS)

# header_map を生成
header_map = {i: h for i, h in enumerate(HEADERS)}


def load_element_aliases(xlsx_path):
    """SourceGame/SourceElement の Element シートから alias 一覧を読み込む"""
    aliases = set()
    if not os.path.exists(xlsx_path):
        return aliases

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "Element" not in wb.sheetnames:
            return aliases

        ws = wb["Element"]
        headers = list(
            next(ws.iter_rows(min_row=1, max_row=1, max_col=80, values_only=True))
        )
        while headers and headers[-1] is None:
            headers.pop()
        if "id" not in headers or "alias" not in headers:
            return aliases

        id_idx = headers.index("id")
        alias_idx = headers.index("alias")

        # 無限行走査を避けるため、連続空行で早期終了する
        empty_run = 0
        for row in ws.iter_rows(min_row=4, max_col=alias_idx + 1, values_only=True):
            row_id = row[id_idx] if id_idx < len(row) else None
            alias = row[alias_idx] if alias_idx < len(row) else None

            if row_id is None and alias in (None, ""):
                empty_run += 1
                if empty_run >= 200:
                    break
                continue

            empty_run = 0
            if isinstance(alias, str) and alias:
                aliases.add(alias)
    finally:
        wb.close()

    return aliases


def load_main_element_names(xlsx_path):
    """SourceGame の Element(910-926) から mainElement の有効名一覧を読み込む"""
    names = set()
    if not os.path.exists(xlsx_path):
        return names

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "Element" not in wb.sheetnames:
            return names

        ws = wb["Element"]
        headers = list(
            next(ws.iter_rows(min_row=1, max_row=1, max_col=80, values_only=True))
        )
        while headers and headers[-1] is None:
            headers.pop()
        if "id" not in headers or "name" not in headers:
            return names

        id_idx = headers.index("id")
        name_idx = headers.index("name")
        valid_ids = set(range(910, 927))  # eleFire .. eleVoid

        empty_run = 0
        for row in ws.iter_rows(
            min_row=4, max_col=max(id_idx, name_idx) + 1, values_only=True
        ):
            row_id = row[id_idx] if id_idx < len(row) else None
            name = row[name_idx] if name_idx < len(row) else None

            if row_id is None and name in (None, ""):
                empty_run += 1
                if empty_run >= 200:
                    break
                continue

            empty_run = 0
            if (
                isinstance(row_id, (int, float))
                and int(row_id) in valid_ids
                and isinstance(name, str)
                and name
            ):
                names.add(name)
    finally:
        wb.close()

    return names


def load_tone_ids(xlsx_path):
    """chara_tone.xlsx から有効な tone id 一覧を抽出"""
    tone_ids = set()
    if not os.path.exists(xlsx_path):
        return tone_ids

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, max_col=80, values_only=True)))
        while headers and headers[-1] is None:
            headers.pop()
        if "id" not in headers:
            return tone_ids

        id_idx = headers.index("id")

        empty_run = 0
        for row in ws.iter_rows(min_row=2, max_col=id_idx + 1, values_only=True):
            row_id = row[id_idx] if id_idx < len(row) else None

            if row_id in (None, ""):
                empty_run += 1
                if empty_run >= 200:
                    break
                continue

            empty_run = 0
            tone_ids.add(str(row_id).strip())
    finally:
        wb.close()

    return tone_ids


def load_bio_tones_from_source_chara(xlsx_path):
    """SourceChara.xlsx の bio 出現値から tone 候補を抽出"""
    tone_ids = set()
    if not os.path.exists(xlsx_path):
        return tone_ids

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = list(
            next(ws.iter_rows(min_row=1, max_row=1, max_col=80, values_only=True))
        )
        while headers and headers[-1] is None:
            headers.pop()
        if "bio" not in headers:
            return tone_ids

        bio_idx = headers.index("bio")
        empty_run = 0
        for row in ws.iter_rows(min_row=4, max_col=bio_idx + 1, values_only=True):
            value = row[bio_idx] if bio_idx < len(row) else None
            if value in (None, ""):
                empty_run += 1
                if empty_run >= 200:
                    break
                continue

            empty_run = 0
            bio = str(value).strip()
            parts = [p.strip() for p in bio.split("/")]
            if len(parts) == 5 and parts[4]:
                tone_ids.add(parts[4])
    finally:
        wb.close()

    return tone_ids


def validate_act_combat_entry(chara_id, act_combat, known_aliases):
    """actCombat 文字列の構文・IDを検証してエラー一覧を返す"""
    errors = []
    if not act_combat:
        return errors

    entries = [e.strip() for e in act_combat.split(",") if e.strip()]
    for idx, entry in enumerate(entries):
        if "/" not in entry:
            errors.append(
                f"{chara_id}: actCombat[{idx}] '{entry}' is invalid (expected '<Action>/<Weight>')"
            )
            continue

        action, weight_str = entry.split("/", 1)
        action = action.strip()
        weight_str = weight_str.strip()

        if not action:
            errors.append(f"{chara_id}: actCombat[{idx}] has empty action in '{entry}'")
            continue

        if not weight_str.isdigit() or int(weight_str) <= 0:
            errors.append(
                f"{chara_id}: actCombat[{idx}] has invalid weight in '{entry}'"
            )

        # Hallucination/typo 検知: 既知の Act*/Sp* alias と突合
        if action.startswith(("Act", "Sp")):
            if action not in known_aliases:
                errors.append(
                    f"{chara_id}: actCombat[{idx}] unknown action '{action}' "
                    "(not found in Element aliases)"
                )
            continue

        # パラメータ化された攻撃IDは形式のみ検証
        if action.startswith(("breathe_", "hand_", "arrow_")):
            if not re.match(r"^[A-Za-z0-9_]+$", action):
                errors.append(f"{chara_id}: actCombat[{idx}] invalid token '{action}'")
            continue

        # それ以外は明示的にエラー（未知フォーマット）
        errors.append(
            f"{chara_id}: actCombat[{idx}] unsupported action format '{action}'"
        )

    return errors


def validate_elements_entry(chara_id, elements, known_aliases):
    """elements 文字列の構文・IDを検証してエラー一覧を返す"""
    errors = []
    if not elements:
        return errors

    entries = [e.strip() for e in elements.split(",") if e.strip()]
    for idx, entry in enumerate(entries):
        if "/" not in entry:
            errors.append(
                f"{chara_id}: elements[{idx}] '{entry}' is invalid (expected '<ElementAlias>/<Value>')"
            )
            continue

        alias, value_str = entry.split("/", 1)
        alias = alias.strip()
        value_str = value_str.strip()

        if not alias:
            errors.append(f"{chara_id}: elements[{idx}] has empty alias in '{entry}'")
            continue

        if alias not in known_aliases:
            errors.append(
                f"{chara_id}: elements[{idx}] unknown alias '{alias}' "
                "(not found in Element aliases)"
            )

        if not re.match(r"^-?\d+$", value_str):
            errors.append(
                f"{chara_id}: elements[{idx}] has invalid value '{value_str}' in '{entry}'"
            )

    return errors


def validate_bio_entry(chara_id, bio, known_tone_ids):
    """bio 文字列の形式を検証してエラー一覧を返す"""
    errors = []
    if not bio:
        return errors

    gender_set = {"m", "f", "n"}
    value = bio.strip()

    # バニラ互換: "f" のような短縮形式を許容
    if value in gender_set:
        return errors

    parts = value.split("/")
    if len(parts) != 5:
        errors.append(
            f"{chara_id}: bio '{bio}' is invalid (expected 'gender/version/height/weight/personality' or short gender)"
        )
        return errors

    gender, version, height, weight, personality = [p.strip() for p in parts]
    if gender not in gender_set:
        errors.append(f"{chara_id}: bio has invalid gender '{gender}' in '{bio}'")
    if not version.isdigit():
        errors.append(f"{chara_id}: bio has invalid version '{version}' in '{bio}'")
    if not re.match(r"^-?\d+$", height):
        errors.append(f"{chara_id}: bio has invalid height '{height}' in '{bio}'")
    if not re.match(r"^-?\d+$", weight):
        errors.append(f"{chara_id}: bio has invalid weight '{weight}' in '{bio}'")
    if personality == "":
        errors.append(f"{chara_id}: bio has empty personality segment in '{bio}'")
    elif personality not in known_tone_ids:
        errors.append(
            f"{chara_id}: bio has unknown tone id '{personality}' in '{bio}'"
        )

    return errors


def validate_main_element_entry(chara_id, main_element, valid_names):
    """mainElement 文字列の構文・値を検証してエラー一覧を返す"""
    errors = []
    if not main_element:
        return errors

    entries = [e.strip() for e in main_element.split(",") if e.strip()]
    for idx, entry in enumerate(entries):
        token = entry
        weight = None
        if "/" in entry:
            token, weight = [x.strip() for x in entry.split("/", 1)]
            if not re.match(r"^-?\d+$", weight):
                errors.append(
                    f"{chara_id}: mainElement[{idx}] has invalid weight '{weight}' in '{entry}'"
                )

        if token not in valid_names:
            errors.append(
                f"{chara_id}: mainElement[{idx}] unknown element '{token}' "
                f"(valid: {', '.join(sorted(valid_names))})"
            )

    return errors


def validate_all_chara_fields(npc_defs):
    """全NPCの actCombat/elements/bio/mainElement を検証。エラーがあれば終了"""
    known_aliases = set()
    known_aliases.update(load_element_aliases(VANILLA_SOURCE_GAME_XLSX))
    known_aliases.update(load_element_aliases(MOD_SOURCE_ELEMENT_XLSX))
    main_element_names = load_main_element_names(VANILLA_SOURCE_GAME_XLSX)
    known_tone_ids = load_tone_ids(VANILLA_CHARA_TONE_XLSX)
    if not known_tone_ids:
        # フォールバック: SourceChara の bio 出現値から抽出
        known_tone_ids = load_bio_tones_from_source_chara(VANILLA_SOURCE_CHARA_XLSX)
    if not main_element_names:
        # フォールバック（バニラ既定の主要属性）
        main_element_names = {
            "Fire",
            "Cold",
            "Lightning",
            "Darkness",
            "Mind",
            "Poison",
            "Nether",
            "Sound",
            "Nerve",
            "Holy",
            "Chaos",
            "Magic",
            "Ether",
            "Acid",
            "Cut",
            "Impact",
            "Void",
        }

    if not known_aliases:
        print(
            "WARNING: actCombat/elements validation skipped (no Element alias sources found)"
        )
        return
    if not known_tone_ids:
        print("WARNING: bio validation skipped (no chara_tone source found)")

    errors = []
    for npc in npc_defs:
        chara_id = npc.get("id", "<unknown>")
        act_combat = npc.get("actCombat", "")
        elements = npc.get("elements", "")
        bio = npc.get("bio", "")
        main_element = npc.get("mainElement", "")
        errors.extend(validate_act_combat_entry(chara_id, act_combat, known_aliases))
        errors.extend(validate_elements_entry(chara_id, elements, known_aliases))
        if known_tone_ids:
            errors.extend(validate_bio_entry(chara_id, bio, known_tone_ids))
        errors.extend(
            validate_main_element_entry(chara_id, main_element, main_element_names)
        )

    if errors:
        print("\n[ERROR] chara field validation failed:")
        for e in errors:
            print(f"  - {e}")
        print(
            "\nFix invalid actCombat/elements/bio/mainElement entries in tools/builder/create_chara_excel.py"
        )
        sys.exit(1)


def create_npc_row(npc_def):
    """NPC定義から行データを作成"""
    row = [""] * NUM_COLUMNS
    for k, v in npc_def.items():
        # name_CN, aka_CN はExcelカラムには含めない（ビルド時変換用）
        if k in ["name_CN", "aka_CN"]:
            continue
        found = False
        for col_idx, col_name in header_map.items():
            if col_name == k:
                row[col_idx] = v
                found = True
                break
        if not found:
            print(f"WARNING: Field '{k}' not found in headers!")
    return row


def create_npc_row_for_lang(npc_def, lang="en"):
    """言語に応じたNPC行を作成"""
    row = [""] * NUM_COLUMNS
    for k, v in npc_def.items():
        # CN言語の場合、nameとakaとdetailカラムを中国語に置き換え
        if lang == "cn":
            if k == "name" and "name_CN" in npc_def and npc_def["name_CN"]:
                v = npc_def["name_CN"]
            elif k == "aka" and "aka_CN" in npc_def and npc_def["aka_CN"]:
                v = npc_def["aka_CN"]
            elif k == "detail" and "detail_CN" in npc_def and npc_def["detail_CN"]:
                v = npc_def["detail_CN"]

        # name_CN, aka_CN, detail_CN はExcelカラムには含めない
        if k in ["name_CN", "aka_CN", "detail_CN"]:
            continue

        found = False
        for col_idx, col_name in header_map.items():
            if col_name == k:
                row[col_idx] = v
                found = True
                break
        if not found:
            print(f"WARNING: Field '{k}' not found in headers!")
    return row


# ===== NPC定義 =====
# bio format: 性別(m/f) / バージョン(固定ID) / 身長 / 体重 / 性格 | 髪色 | 肌色
# _idRenderData: @chara (PCCシステムを使用、Texture/ID.pngが自動ロードされる)
# tiles: バニラのフォールバックタイルID (カスタム画像が見つからない時に使用)
# idText: テキストID (同じIDのテキストファイルが参照される)
#
# CWL タグ仕様:
# - addZone_ゾーンID: 指定ゾーンにキャラクターを生成
# - addFlag_StayHomeZone: ランダム移動を無効化（初期ゾーンに留まる）
# - addDrama_テーブル名: ドラマシートをリンク
# - humanSpeak: 人間らしい会話表示（括弧なし）
# - addStock: 商人の在庫を追加
#
# 描画メモ:
#   256x256スプライトは下部に64px空白を設けること（地面めり込み防止）

ZONE_ID = "sukutsu_arena"  # カスタムゾーンID

npcs = []

# 1. リリィ (サキュバス / 女性 / 受付嬢)
npcs.append(
    {
        "id": "sukutsu_receptionist",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "リリシエル",
        "name": "Lilithiel",
        "name_CN": "莉莉西尔",
        "aka_JP": "絡繰り番",
        "aka": "Clockwork Keeper",
        "aka_CN": "机关守护者",
        "race": "succubus",
        "job": "shopkeeper",
        "_idRenderData": "@chara",
        "tiles": 340,
        "LV": 500,
        "hostility": "Friend",
        "bio": "f/1001/165/52/friendly",
        "idText": "sukutsu_receptionist",
        "tag": f"neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addStock_sukutsu_receptionist,addDrama_drama_sukutsu_receptionist,humanSpeak",
        "trait": "SukutsuMerchant",
        "quality": 4,
        "chance": 0,
    }
)

# 2. バルガス (人間 / 男性 / アリーナマスター)
_sukutsu_balgas_base = {
    "id": "sukutsu_arena_master",
    "Author": "tishi.elin.sukutsu_arena",
    "name_JP": "バルガス",
    "name": "Vargus",
    "name_CN": "巴尔加斯",
    "aka_JP": "アリーナマスター",
    "aka": "Champion of Hundred Battles",
    "aka_CN": "百战霸主",
    "race": "juere",
    "job": "warrior",
    "_idRenderData": "@chara",
    "tiles": 0,
    "LV": 800,
    "hostility": "Friend",
    "bio": "m/1002/185/90/rude",
    "idText": "sukutsu_arena_master",
    "tag": f"neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addDrama_drama_sukutsu_arena_master,humanSpeak",
    "trait": "SukutsuNPC",
    "quality": 4,
    "chance": 0,
}
npcs.append(_sukutsu_balgas_base)

# 3. アスタロト (ドラゴン / 男性 / アリーナ創設者 / 最終ボス)
_sukutsu_astaroth_base = {
    "id": "sukutsu_astaroth",
    "Author": "tishi.elin.sukutsu_arena",
    "name_JP": "アスタロト",
    "name": "Astaroth",
    "name_CN": "阿斯塔罗特",
    "aka_JP": "うつろいし竜神",
    "aka": "The Hollow Dragon God",
    "aka_CN": "虚幻龙神",
    "race": "dragon",
    "job": "warrior",
    "_idRenderData": "@chara",
    "tiles": 168,
    "LV": 3000,
    "hostility": "Friend",
    "bio": "m/1003/350/500/VIP",
    "idText": "sukutsu_astaroth",
    "portrait": "UN_sukutsu_astaroth",
    "mainElement": "Chaos",
    # P1弱点: 毒（古代の竜神は毒への免疫がない）
    "elements": "featElder/1,resChaos/20,resNether/20,resMagic/20,resDarkness/20,resFire/20,resNerve/20,resMind/20,resHoly/20,resCut/20,resImpact/20,featBoost/1,featBloodBond/1,resPoison/-10",
    "aiParam": "2,70,60",
    "actCombat": "breathe_Void/35,breathe_Chaos/25,breathe_Nether/20,SpGravity/10,SpBane/5,ActGazeInsane/15,ActGazeMutation/10,SpHeal/5,SpSummonDragon/5,SpEarthquake/5,SpShutterHex/5,SpSpeedDown/5,SpSilence/5,SpWeakness/5,SpSummonTentacle/5,SpGate/5,ActTouchDrown/5,ActNeckHunt/5",
    "tag": f"boss,undead,addZone_{ZONE_ID},addFlag_StayHomeZone,addDrama_drama_sukutsu_astaroth,humanSpeak,phase_line=astaroth,phase2=sukutsu_astaroth_p2,phase2_hp=70,phase3=sukutsu_astaroth_p3,phase3_hp=40,phase4=sukutsu_astaroth_p4,phase4_hp=20",
    "trait": "SukutsuNPC",
    "quality": 4,  # Legendary（デバッグ用）
    "chance": 0,
}
npcs.append(_sukutsu_astaroth_base)

_sukutsu_astaroth_p2 = _sukutsu_astaroth_base.copy()
_sukutsu_astaroth_p2.update(
    {
        "id": "sukutsu_astaroth_p2",
        "name_JP": "アスタロト・反撃形態",
        "name": "Astaroth (Counter)",
        "name_CN": "阿斯塔罗特・反击形态",
        "aiParam": "1,60,50",
        # P2弱点: 魔法（物理特化形態→魔法が弱点）+ 冷気ブレス（対メタル）
        "elements": "featElder/1,resChaos/20,resNether/20,resMagic/-10,resDarkness/20,resFire/20,resNerve/20,resMind/20,resHoly/20,resCut/20,resImpact/20,featBoost/1,featBloodBond/1",
        "actCombat": "breathe_Void/25,breathe_Chaos/20,breathe_Cold/30,SpGravity/10,ActGazeInsane/20,ActGazeMutation/15,ActTouchDrown/15,ActNeckHunt/15,SpHeal/5,SpShutterHex/5,SpSilence/5",
        "tag": "boss,undead,phase_line=astaroth,phase3=sukutsu_astaroth_p3,phase3_hp=40,phase4=sukutsu_astaroth_p4,phase4_hp=20",
    }
)
npcs.append(_sukutsu_astaroth_p2)

_sukutsu_astaroth_p3 = _sukutsu_astaroth_base.copy()
_sukutsu_astaroth_p3.update(
    {
        "id": "sukutsu_astaroth_p3",
        "name_JP": "アスタロト・召喚形態",
        "name": "Astaroth (Summoner)",
        "name_CN": "阿斯塔罗特・召唤形态",
        "aiParam": "3,90,90",
        # P3弱点: 冷気（召喚中は集中が必要→冷気で乱される）+ マナ吸収（対イモーロナク）
        "elements": "featElder/1,resChaos/20,resNether/20,resMagic/20,resDarkness/20,resFire/20,resNerve/20,resMind/20,resHoly/20,resCut/20,resImpact/20,featBoost/1,featBloodBond/1,resCold/-10",
        "actCombat": "SpSummonDragon/50,SpSummonDragon/45,SpSummonDragon/40,SpSummonTentacle/20,SpGate/10,SpSpeedDown/10,SpWeakness/10,SpSilence/10,breathe_Void/10,SpHeal/20,SpHeal/15",
        "tag": "boss,undead,phase_line=astaroth,phase4=sukutsu_astaroth_p4,phase4_hp=20",
    }
)
npcs.append(_sukutsu_astaroth_p3)

_sukutsu_astaroth_p4 = _sukutsu_astaroth_base.copy()
_sukutsu_astaroth_p4.update(
    {
        "id": "sukutsu_astaroth_p4",
        "name_JP": "アスタロト・異形",
        "name": "Astaroth (Aberrant)",
        "name_CN": "阿斯塔罗特・异形",
        "aiParam": "4,95,100",
        "size": "2,2",
        # P4弱点: 混沌（異形化で秩序を失い、混沌に脆くなる）+ 音ブレス・マナ吸収（対チートペット）
        "elements": "featElder/1,resChaos/-12,resNether/20,resMagic/20,resDarkness/20,resFire/20,resNerve/20,resMind/20,resHoly/20,resCut/20,resImpact/20,featBoost/1,featBloodBond/1",
        "actCombat": "breathe_Chaos/25,breathe_Nether/20,breathe_Sound/25,SpMeteor/35,SpEarthquake/35,SpGravity/15,SpShutterHex/15,SpSummonTentacle/20,SpGate/15,ActGazeInsane/20,ActGazeMutation/20,SpHeal/10,ActTouchDrown/15",
        "tag": "boss,undead,phase_line=astaroth",
    }
)
npcs.append(_sukutsu_astaroth_p4)

# 4. 怪しい商人 (ミュータント / 性別不明 / 商人)
npcs.append(
    {
        "id": "sukutsu_shady_merchant",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "ゼク",
        "name": "Ezekiel",
        "name_CN": "泽克",
        "aka_JP": "剥製師",
        "aka": "The Taxidermist",
        "aka_CN": "标本师",
        "race": "mutant",
        "job": "merchant",
        "LV": 666,
        "hostility": "Friend",
        "tiles": 807,
        "_idRenderData": "@chara",
        "bio": "m/1004/170/65/shy",
        "idText": "sukutsu_shady_merchant",
        "tag": f"neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addStock_sukutsu_shady_merchant,addDrama_drama_sukutsu_shady_merchant,humanSpeak",
        "trait": "SukutsuShadyMerchant",
        "quality": 4,
        "chance": 0,
    }
)

# 5. アイリス (トレーナー / Yith族)
npcs.append(
    {
        "id": "sukutsu_trainer",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "アイリス",
        "name": "Iris",
        "name_CN": "艾莉丝",
        "aka_JP": "仮座の漂客",
        "aka": "The Drifting Guest",
        "aka_CN": "假座的漂客",
        "race": "yith",
        "job": "warrior",
        "_idRenderData": "@chara",
        "tiles": 340,
        "LV": 500,
        "hostility": "Friend",
        "bio": "f/1010/162/52/friendly",
        "idText": "sukutsu_trainer",
        # 精神系フィート・耐性（Yith族: 精神/神経/冥界/混沌耐性、高回避+カウンター）
        "elements": "featElder/1,resMind/50,resNerve/40,resNether/30,resChaos/20,evasionPerfect/85,counter/40",
        # 精神系スキル（狂気の眼差し、悪夢、首狩り、テレポート、透明化）
        "actCombat": "ActGazeInsane/30,SpNightmare/25,ActNeckHunt/20,SpTeleport/15,SpInvisibility/10",
        "tag": f"neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addDrama_drama_sukutsu_trainer,humanSpeak",
        "trait": "SukutsuTrainer,combat",
        "quality": 4,
        "chance": 0,
    }
)

# 6. デバッグマスター (開発テスト用)
if DEBUG_MODE:
    npcs.append(
        {
            "id": "sukutsu_debug_master",
            "Author": "tishi.elin.sukutsu_arena",
            "name_JP": "観測者",
            "name": "Observer",
            "name_CN": "观测者",
            "aka_JP": "次元の記録者",
            "aka": "Dimensional Recorder",
            "aka_CN": "次元记录者",
            "race": "spirit",
            "job": "wizard",
            "_idRenderData": "@chara",
            "tiles": 478,
            "LV": 1,
            "hostility": "Friend",
            "bio": "n/1005/160/0/shy",
            "idText": "sukutsu_debug_master",
            "tag": f"neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addDrama_drama_debug_menu,humanSpeak",
            "quality": 5,
            "chance": 0,
        }
    )

# ===== ストーリーバトル用敵キャラクター =====

# 6. ヴォイド・プチ (Rank G昇格試合)
npcs.append(
    {
        "id": "sukutsu_void_ooze",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "ヴォイド・プチ",
        "name": "Void Putty",
        "name_CN": "虚空普提",
        "aka_JP": "混沌の落とし子",
        "aka": "Child of Chaos",
        "aka_CN": "混沌之子",
        "race": "slime",
        "job": "warrior",
        "_idRenderData": "@chara",
        "tiles": 296,
        "LV": 15,
        "hostility": "Enemy",
        "bio": "n/2001/50/30/nano",
        "idText": "sukutsu_void_ooze",
        "mainElement": "Chaos",
        "elements": "resChaos/100,featMeatCushion/4",
        "actCombat": "breathe_Chaos/30,ActStealFood/15,ActStealMoney/15,ActDraw/10,ActTouchDrown/10",
        "tag": "boss",
        "quality": 3,
        "chance": 0,
    }
)

# 6b. メタルプチ (観客講義クエスト)
npcs.append(
    {
        "id": "sukutsu_metal_putty",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "メタルプチ？",
        "name": "Metal Putty?",
        "name_CN": "金属普提？",
        "aka_JP": "鋼の塊",
        "aka": "Steel Blob",
        "aka_CN": "钢铁之块",
        "race": "slime",
        "job": "warrior",
        "_idRenderData": "@chara",
        "tiles": 296,
        "LV": 30,
        "hostility": "Enemy",
        "bio": "n/2007/50/30/nano",
        "idText": "sukutsu_metal_putty",
        "mainElement": "Cut",
        "elements": "featMetal/995,resCut/60,resImpact/60",
        "actCombat": "ActStealFood/15,ActStealMoney/15,ActDraw/10",
        "tag": "boss",
        "quality": 3,
        "chance": 0,
    }
)

# 7. ヴォイド・アイスハウンド (Rank F昇格試合)
npcs.append(
    {
        "id": "sukutsu_frost_hound",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "ヴォイド・アイスハウンド",
        "name": "Void Ice Hound",
        "name_CN": "虚空冰犬",
        "aka_JP": "凍える牙",
        "aka": "Freezing Fang",
        "aka_CN": "冰冻之牙",
        "race": "hound",
        "job": "warrior",
        "_idRenderData": "@chara",
        "tiles": 254,
        "LV": 45,
        "hostility": "Enemy",
        "bio": "n/2002/120/80/rude",
        "idText": "sukutsu_frost_hound",
        "mainElement": "Cold",
        "elements": "invisibility/1,featElder/1,resCold/100",
        "actCombat": "breathe_Cold/40,SpSpeedDown/20",
        "tag": "boss",
        "quality": 4,
        "chance": 0,
    }
)

# 8. 訓練用バルガス
_sukutsu_balgas_training = _sukutsu_balgas_base.copy()
_sukutsu_balgas_training.update(
    {
        "id": "sukutsu_balgas_training",
        "LV": 200,
        "hostility": "Enemy",
        "mainElement": "Cut",
        "elements": "featElder/1,resCut/60,resImpact/60",
        "actCombat": "",
        "tag": "boss",
        "trait": "",
    }
)
npcs.append(_sukutsu_balgas_training)

# 9. 全盛期バルガス (Rank S昇格試合)
_sukutsu_balgas_prime = _sukutsu_balgas_base.copy()
_sukutsu_balgas_prime.update(
    {
        "id": "sukutsu_balgas_prime",
        "name_JP": "全盛期のバルガス",
        "name": "Vargus at His Prime",
        "name_CN": "全盛期的巴尔加斯",
        "aka_JP": "鉄血の覇者",
        "aka": "Iron-Blooded Champion",
        "aka_CN": "铁血霸主",
        "LV": 2000,
        "hostility": "Enemy",
        "bio": "m/1002/188/95/rude",
        "mainElement": "Cut",
        # 雷弱点（鉄血の覇者→金属鎧→導電性）+ 音ブレス追加（対メタル）
        "elements": "featElder/1,resCut/80,resImpact/80,resFire/40,resCold/40,resLightning/-12,resMind/60,resNerve/60,antiMagic/60,vopal/60,mod_flurry/65,mod_chaser/30,mod_cleave/10,redirect_blaser/23,counter/20",
        # aiParam: [1,90,80] = 距離1維持、移動90%、再移動80% → 積極的に近接を狙う
        "aiParam": "1,90,80",
        "actCombat": "breathe_Cut/40,breathe_Impact/30,breathe_Sound/20,SpHero/10,ActRush/25,ActBash/20",
        "tag": "boss",
        "trait": "",
        "quality": 4,  # Legendary（有精卵からの増殖を防ぐ）
    }
)
npcs.append(_sukutsu_balgas_prime)

# 10. カイン (復活後NPC / バルガスの弟弟子)
npcs.append(
    {
        "id": "sukutsu_cain",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "カイン",
        "name": "Cain",
        "name_CN": "凯恩",
        "aka_JP": "鉄血団の俊英",
        "aka": "Prodigy of Iron Blood",
        "aka_CN": "铁血团的俊杰",
        "race": "juere",  # バルガスと同じ種族
        "job": "warrior",
        "_idRenderData": "@chara",
        "tiles": 0,
        "LV": 500,  # バルガスより少し低め
        "hostility": "Friend",
        "bio": "m/1012/180/85/friendly",
        "idText": "sukutsu_cain",
        "tag": f"neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,humanSpeak",
        "trait": "SukutsuNPC",
        "quality": 4,
        "chance": 0,
    }
)

# 10b. カインの亡霊 (Rank E昇格試合)
npcs.append(
    {
        "id": "sukutsu_kain_ghost",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "錆びついた英雄カイン",
        "name": "Rusted Hero Cain",
        "name_CN": "锈蚀的英雄凯恩",
        "aka_JP": "忘れられた副団長",
        "aka": "Forgotten Vice-Captain",
        "aka_CN": "被遗忘的副团长",
        "race": "wraith",
        "job": "warrior",
        "_idRenderData": "@chara",
        "tiles": 458,
        "LV": 150,
        "hostility": "Enemy",
        "bio": "m/1006/180/0/shy",
        "idText": "sukutsu_kain_ghost",
        "mainElement": "Nether",
        "actCombat": "breathe_Acid/25,breathe_Cut/25,hand_Nether/30,SpSummonShadow/15",
        "tag": "boss,undead",
        "quality": 4,
        "chance": 0,
    }
)

# 11. ヌル (Rank B昇格試合 / NPC版)
_sukutsu_null_base = {
    "id": "sukutsu_null",
    "Author": "tishi.elin.sukutsu_arena",
    "name_JP": "Nul",
    "name": "Nul",
    "name_CN": "Nul",
    "aka_JP": "虚無の処刑人",
    "aka": "Void Executioner",
    "aka_CN": "虚空行刑者",
    "race": "machine",
    "job": "thief",
    "_idRenderData": "@chara",
    "tiles": 536,
    "LV": 1000,
    "hostility": "Friend",
    "bio": "f/1007/165/45/princess",
    "idText": "sukutsu_null",
    "portrait": "UN_sukutsu_null",
    "mainElement": "Void",
    "elements": "featSplit/1,featElder/1,resChaos/80,resNether/60,resMagic/40,resNerve/100,resMind/100,featGolem/1,featReboot/1,featBoost/1,featEarthStrength/1,featRapidArrow/3,featGeneSlot/10,featMiscreation/1,featMetal/120,featManaMeat/1,featRoran/1,evasionPerfect/60",
    "actCombat": "hand_Void/40,SpInvisibility/30,SpSilence/15,ActGazeInsane/15,ActRush/10,ActInsult/10,SpEarthquake/10,arrow_Void/20",
    "tag": f"neutral,addZone_{ZONE_ID},addFlag_StayHomeZone,addDrama_drama_sukutsu_null,humanSpeak",
    "trait": "SukutsuNPC",
    "quality": 4,
    "chance": 0,
}
npcs.append(_sukutsu_null_base)

# 11b. ヌル (敵バージョン)
_sukutsu_null_enemy = _sukutsu_null_base.copy()
_sukutsu_null_enemy.update(
    {
        "id": "sukutsu_null_enemy",
        "hostility": "Enemy",
        "tag": "",  # ボス属性なし（HP倍率を適用しない）
        "trait": "",
    }
)
npcs.append(_sukutsu_null_enemy)

# 12. 影の自己 (Rank A戦 / NPC)
npcs.append(
    {
        "id": "sukutsu_shadow_self",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "影の自己",
        "name": "Shadow Self",
        "name_CN": "影之自我",
        "aka_JP": "影の分身",
        "aka": "Mirror Image",
        "aka_CN": "影之分身",
        "detail_JP": "",
        "detail": "",
        "detail_CN": "",
        "race": "phantom",
        "job": "warmage",
        "_idRenderData": "@chara",
        "tiles": 460,
        "LV": 1500,
        "hostility": "Enemy",
        "bio": "n/1008/170/60/rude",
        "idText": "sukutsu_shadow_self",
        "elements": "featDemon/1,featElder/1,resDarkness/60,resNether/50,resMind/80,resNerve/80,featBoost/1,featRapidArrow/2,evasionPerfect/40,mod_flurry/40,counter/30",
        "actCombat": "SpSummonTentacle/40,SpSummonTentacle/35,SpGate/20,SpHeal/20,SpTeleport/15,SpSilence/10,SpWeakness/10,arrow_Darkness/10",
        "tag": "boss,undead,phase_line=shadow_self,phase2=sukutsu_shadow_self_p2,phase2_hp=70,phase3=sukutsu_shadow_self_p3,phase3_hp=40,phase4=sukutsu_shadow_self_p4,phase4_hp=20",
        "quality": 4,
        "chance": 0,
    }
)

_sukutsu_shadow_self_p2 = {
    "id": "sukutsu_shadow_self_p2",
    "Author": "tishi.elin.sukutsu_arena",
    "name_JP": "影の自己・召喚形態",
    "name": "Shadow Self (Summoner)",
    "name_CN": "影之自我・召唤形态",
    "detail_JP": "・召喚形態",
    "detail": " (Summoner)",
    "detail_CN": "・召唤形态",
    "race": "phantom",
    "job": "warmage",
    "_idRenderData": "@chara",
    "tiles": 460,
    "LV": 1500,
    "hostility": "Enemy",
    "bio": "n/1008/170/60/rude",
    "idText": "sukutsu_shadow_self",
    "elements": "featDemon/1,featElder/1,resDarkness/60,resNether/50,resMind/80,resNerve/80,featBoost/1,featRapidArrow/2,evasionPerfect/40,mod_flurry/40,counter/30",
    "aiParam": "3,90,90",
    "actCombat": "SpSummonTentacle/50,SpSummonTentacle/45,SpGate/25,SpHeal/25,SpTeleport/15,SpSilence/10,SpWeakness/10",
    "tag": "boss,undead,phase_line=shadow_self,phase3=sukutsu_shadow_self_p3,phase3_hp=40,phase4=sukutsu_shadow_self_p4,phase4_hp=20",
    "quality": 4,
    "chance": 0,
}
npcs.append(_sukutsu_shadow_self_p2)

_sukutsu_shadow_self_p3 = _sukutsu_shadow_self_p2.copy()
_sukutsu_shadow_self_p3.update(
    {
        "id": "sukutsu_shadow_self_p3",
        "name_JP": "影の自己・召喚形態+",
        "name": "Shadow Self (Summoner+)",
        "name_CN": "影之自我・召唤形态+",
        "detail_JP": "・召喚形態+",
        "detail": " (Summoner+)",
        "detail_CN": "・召唤形态+",
        "aiParam": "3,95,95",
        "actCombat": "SpSummonTentacle/60,SpSummonTentacle/55,SpGate/30,SpHeal/25,SpTeleport/15,SpSilence/10",
        "tag": "boss,undead,phase_line=shadow_self,phase4=sukutsu_shadow_self_p4,phase4_hp=20",
    }
)
npcs.append(_sukutsu_shadow_self_p3)

_sukutsu_shadow_self_p4 = _sukutsu_shadow_self_p2.copy()
_sukutsu_shadow_self_p4.update(
    {
        "id": "sukutsu_shadow_self_p4",
        "name_JP": "影の自己・暴走形態",
        "name": "Shadow Self (Rampage)",
        "name_CN": "影之自我・暴走形态",
        "detail_JP": "・暴走形態",
        "detail": " (Rampage)",
        "detail_CN": "・暴走形态",
        "aiParam": "4,95,100",
        "actCombat": "SpSummonTentacle/55,SpSummonTentacle/50,SpGate/30,SpHeal/30,SpTeleport/15,SpSilence/15,SpWeakness/15",
        "tag": "boss,undead,phase_line=shadow_self",
    }
)
npcs.append(_sukutsu_shadow_self_p4)

# 13. グリード (Rank D昇格試合)
npcs.append(
    {
        "id": "sukutsu_greed",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "グリード",
        "name": "Greed",
        "name_CN": "贪婪",
        "aka_JP": "観客の代弁者",
        "aka": "Voice of the Audience",
        "aka_CN": "观众的代言人",
        "race": "wraith",
        "job": "warrior",
        "_idRenderData": "@chara",
        "tiles": 0,
        "LV": 280,
        "hostility": "Enemy",
        "bio": "m/2003/175/70/rude",
        "idText": "sukutsu_greed",
        "mainElement": "Sound",
        "elements": "resSound/50,resChaos/30,resMagic/60,featElder/1,featCosmicHorror/1",
        "actCombat": "breathe_Sound/35,breathe_Chaos/25,SpWeakness/15,ActInsult/30,SpHeal/10",
        "tag": "boss",
        "quality": 4,
        "chance": 0,
    }
)

# ===== Cランク: 朱砂食い（3体ボス）=====

# 14. クロウ（影）
npcs.append(
    {
        "id": "sukutsu_crow_shadow",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "クロウ",
        "name": "Crow",
        "name_CN": "克罗",
        "aka_JP": "影の鴉",
        "aka": "Crow of Shadows",
        "aka_CN": "暗影之鸦",
        "race": "wraith",
        "job": "thief",
        "_idRenderData": "@chara",
        "tiles": 2,
        "LV": 350,
        "hostility": "Enemy",
        "bio": "m/2004/170/60/shy",
        "idText": "sukutsu_crow_shadow",
        "mainElement": "Darkness",
        "elements": "invisibility/1,resDarkness/60,resNether/40",
        "actCombat": "hand_Darkness/40,SpInvisibility/20,hand_Nether/30",
        "tag": "boss",
        "quality": 4,
        "chance": 0,
    }
)

# 15. レイヴン（刃）
npcs.append(
    {
        "id": "sukutsu_raven_blade",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "レイヴン",
        "name": "Raven",
        "name_CN": "雷文",
        "aka_JP": "刃の鴉",
        "aka": "Raven of Blades",
        "aka_CN": "利刃之鸦",
        "race": "wraith",
        "job": "warrior",
        "_idRenderData": "@chara",
        "tiles": 0,
        "LV": 400,
        "hostility": "Enemy",
        "bio": "m/2005/185/90/rude",
        "idText": "sukutsu_raven_blade",
        "mainElement": "Cut",
        "elements": "featElder/1,resCut/50,resImpact/40",
        "actCombat": "breathe_Cut/40,breathe_Impact/30",
        "tag": "boss",
        "quality": 4,
        "chance": 0,
    }
)

# 16. カラス（毒）
npcs.append(
    {
        "id": "sukutsu_karasu_venom",
        "Author": "tishi.elin.sukutsu_arena",
        "name_JP": "カラス",
        "name": "Karasu",
        "name_CN": "卡拉斯",
        "aka_JP": "毒の鴉",
        "aka": "Crow of Venom",
        "aka_CN": "剧毒之鸦",
        "race": "mutant",
        "job": "wizard",
        "_idRenderData": "@chara",
        "tiles": 807,
        "LV": 450,
        "hostility": "Enemy",
        "bio": "f/2006/160/50/shy",
        "idText": "sukutsu_karasu_venom",
        "mainElement": "Poison",
        "elements": "resPoison/100,resAcid/60",
        "actCombat": "breathe_Poison/50,breathe_Acid/30,SpWeakResEle/15",
        "tag": "boss",
        "quality": 4,
        "chance": 0,
    }
)

# デバッグモードの場合、全キャラのLVを1に設定
if BOSS_LV1:
    for npc in npcs:
        tag = npc.get("tag", "")
        if "boss" not in tag:
            continue
        original_lv = npc.get("LV", "N/A")
        npc["LV"] = 1
        print(f"  {npc['id']}: LV {original_lv} -> 1")


def write_tsv(path, row_data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
        writer.writerows(row_data)
    print(f"Created TSV: {path}")


def main():
    print(f"Generating Chara TSV from {len(npcs)} NPC definition(s)...")

    # 生成前に actCombat/elements の検証を実施（typo/hallucination対策）
    validate_all_chara_fields(npcs)

    # デフォルト行（DEFAULTSはリストなのでそのまま使用）
    default_row = DEFAULTS

    # EN版
    rows_en = [HEADERS, TYPES, default_row]
    for npc in npcs:
        rows_en.append(create_npc_row(npc))
    write_tsv(OUTPUT_EN_TSV, rows_en)

    # CN版
    rows_cn = [HEADERS, TYPES, default_row]
    for npc in npcs:
        rows_cn.append(create_npc_row_for_lang(npc, lang="cn"))
    write_tsv(OUTPUT_CN_TSV, rows_cn)

    print(f"Generated {len(npcs)} NPC(s) (EN + CN)")


if __name__ == "__main__":
    main()

