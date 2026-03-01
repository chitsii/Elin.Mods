# -*- coding: utf-8 -*-
"""
dialog.xlsx 生成スクリプト

CWLのdialog.xlsx（Uniqueシート）を生成する。
各NPCに「話す」選択肢でランダム表示されるテキストを定義。

使用方法:
    cd tools
    uv run python builder/create_dialog_excel.py
"""

import os
import sys
from dataclasses import dataclass
from pathlib import Path

# toolsディレクトリをパスに追加
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
sys.path.insert(0, TOOLS_DIR)

from openpyxl import Workbook
from openpyxl.styles import Font


@dataclass
class UniqueDialogEntry:
    """Uniqueシートのエントリ定義"""

    chara_id: str
    texts_jp: list[str]
    texts_en: list[str]
    texts_cn: list[str]


# ============================================================================
# NPC会話定義 (v6ベース)
# ============================================================================

# 各NPCの「話す」選択肢で表示されるランダムテキスト
# 改行で区切られた複数テキストからランダムに選ばれる
# POSTをベースに、違和感のないPREも追加

UNIQUE_DIALOGS = [
    # バルガス（アリーナマスター）- 普通のおっさん枠、世話焼き
    UniqueDialogEntry(
        chara_id="sukutsu_arena_master",
        texts_jp=[
            # POST
            "35年ぶりに朝日を見た。……眩しいな、ちくしょう。",
            "熱いシャワーを浴びたい。闘技場は水浴びだったからな。",
            "メシを選べるってすげえ事だよな、まったく。",
            "さっき子供に道聞かれた。……俺が怖くなかったらしい。",
            "お前と旅してると、35年が嘘みてえだ。",
            "酒場で喧嘩売られた。……買わなかったぞ。丸くなったもんだ。",
            "リリィに『太った』って言われた。……余計なお世話だ。",
            "お前、冒険者としては変わってるよな。……褒めてねえぞ。",
            # PRE（違和感なし）
            "腹減ったな。お前、飯食ったか？",
            "怪我してんなら言え。黙ってると死ぬぞ。",
            "俺の必殺技に名前つけようと思ってな。……いや、やっぱいい。忘れろ。",
            "アイリスのやつ、また変なトレーニング考えてたぞ。逃げとけ。",
            "ゼクから物買うなよ。あいつの『お買い得』は信用すんな。",
            "お前、また死んできただろ。その顔見りゃわかる。",
            "なんだその荷物。お前、引っ越しでもすんのか？",
        ],
        texts_en=[
            "First sunrise I've seen in 35 years. ...Damn, it's bright.",
            "I want a hot shower. We only had cold water in the arena.",
            "Being able to choose what to eat is amazing, really.",
            "A kid asked me for directions earlier. ...Guess I'm not that scary.",
            "Traveling with you makes those 35 years feel like a dream.",
            "Got challenged to a fight at the tavern. ...Didn't take it. I've mellowed out.",
            "Lily told me I've gained weight. ...Mind your own business.",
            "You're a weird one for an adventurer. ...That's not a compliment.",
            "I'm hungry. You eaten yet?",
            "If you're hurt, say so. You'll die if you keep quiet.",
            "I was thinking of naming my special move. ...Nah, forget it.",
            "Iris is cooking up some weird training again. Better run.",
            "Don't buy from Zek. His 'bargains' are lies.",
            "You died again, didn't you? I can tell by your face.",
            "What's with all that luggage? You moving house?",
        ],
        texts_cn=[
            "35年来第一次看到日出。……真刺眼啊，该死。",
            "想洗个热水澡。斗技场只有冷水浴。",
            "能选择吃什么真是太棒了。",
            "刚才有小孩向我问路。……看来我没那么吓人。",
            "和你一起旅行，那35年感觉像做梦一样。",
            "在酒馆被人挑衅了。……没理他。我变温和了。",
            "莉莉说我胖了。……多管闲事。",
            "作为冒险者，你真是个怪人。……不是在夸你。",
            "饿了。你吃了吗？",
            "受伤了就说。憋着会死的。",
            "我在想给必杀技起个名字。……算了，忘掉吧。",
            "艾莉丝又在想奇怪的训练了。快跑。",
            "别从泽克那买东西。他的『划算』都是骗人的。",
            "你又死了吧。看你表情就知道。",
            "那些行李是怎么回事。你搬家吗？",
        ],
    ),
    # アイリス（トレーナー）- 陽キャ枠、異文化オタク
    UniqueDialogEntry(
        chara_id="sukutsu_trainer",
        texts_jp=[
            # POST
            "明日何しよっか。……こういうの考えるの、楽しいね。",
            "『観光』って概念すごい。目的なく歩くのが目的なの。哲学じゃん。",
            "猫って温かくて良い。もふもふ。……ヌルちゃんも触りたがってた。",
            "『二度寝』って贅沢の極みじゃない？罪悪感が最高。",
            "オパートス様の信者って筋肉すごいらしいね。入信しようかな。",
            "アスタロトに『エモい』って言ったら、30分くらい意味聞かれた。",
            "バルガスが私のトレーニング褒めてくれた。……たぶん褒めてた。たぶん。",
            "リリィと友達になった！……反応薄かったけど、嬉しかったと思う。たぶん。",
            "君の『ちょっとダンジョン行ってくる』、全然ちょっとじゃないよね。",
            "ゼっさん……あ、ゼクさんが『略すな』ってうるさいの。面白いから続ける。",
            # PRE（違和感なし）
            "知ってる？人間の筋肉、3日で衰え始めるんだよ。だから毎日鍛えよ？",
            "ね、『推し』って概念あるじゃん？私の推し、君かも。",
            "人間の寿命って短すぎない？焦るよね、いろいろ。",
            "バルガスって、怖い顔してるけど優しいよね。バレバレだけど。",
            "ヌルちゃんにトレーニング教えようとしたら、逆に私が倒されちゃった。",
            "君、荷物重すぎない？人間の背骨って意外と脆いんだよ？",
            "また変なもの食べた？君の胃腸、研究対象にしたいレベル。",
        ],
        texts_en=[
            "What should we do tomorrow? ...Thinking about this is fun.",
            "'Sightseeing' is wild. Walking with no purpose IS the purpose. Philosophy!",
            "Cats are so warm and fluffy. ...Null wanted to pet one too.",
            "Isn't sleeping in the ultimate luxury? The guilt feels great.",
            "Opatos followers have amazing muscles apparently. Maybe I should join.",
            "I told Astaroth something was 'lit' and he asked what it meant for 30 minutes.",
            "Vargus complimented my training. ...Probably. I think.",
            "I became friends with Lily! ...She didn't react much, but I think she was happy.",
            "Your 'quick dungeon run' is never quick, is it?",
            "Ze-san... oh, Zek keeps telling me not to shorten his name. It's fun so I'll keep doing it.",
            "Did you know? Human muscles start weakening after 3 days. Train every day!",
            "You know the concept of having a 'fave'? Mine might be you.",
            "Human lifespans are way too short, right? Makes you anxious.",
            "Vargus looks scary but he's kind, right? So obvious.",
            "I tried to teach Null some training, but she knocked me down instead.",
            "Isn't your bag too heavy? Human spines are surprisingly fragile.",
            "Did you eat something weird again? Your stomach is research-worthy.",
        ],
        texts_cn=[
            "明天做什么好呢。……想这些真开心。",
            "『观光』这个概念好厉害。漫无目的地走就是目的。太哲学了！",
            "猫好温暖好舒服。毛茸茸的。……努尔也想摸来着。",
            "『赖床』不是终极奢侈吗？罪恶感超棒。",
            "听说欧帕托斯的信徒肌肉超厉害。要不要入教呢。",
            "我跟阿斯塔罗特说『很酷』，结果他问了我30分钟什么意思。",
            "巴尔加斯夸了我的训练。……大概是在夸吧。大概。",
            "我和莉莉成为朋友了！……虽然她反应平淡，但应该很高兴。大概。",
            "你说的『去一下迷宫』，根本不是一下吧？",
            "泽桑……啊，泽克一直说不要缩写他的名字。很有趣所以继续叫。",
            "你知道吗？人类的肌肉3天就开始退化。所以每天都要锻炼哦？",
            "你知道『推』这个概念吧？我的推，可能是你哦。",
            "人类的寿命太短了吧？各种事情都很着急呢。",
            "巴尔加斯虽然看起来很凶但很温柔吧。太明显了。",
            "我想教努尔训练，结果反被她打倒了。",
            "你的行李太重了吧？人类的脊椎意外地脆弱哦？",
            "又吃了奇怪的东西？你的肠胃值得研究。",
        ],
    ),
    # リリィ（受付嬢）- 毒舌観察者、控えめだが全部見てる
    UniqueDialogEntry(
        chara_id="sukutsu_receptionist",
        texts_jp=[
            # POST
            "初めて自分で行き先を決めました。……怖いけど、嬉しいです。",
            "夢、見ました。内容は覚えてないけど……嬉しかったです。",
            "ヌルさんと一緒にいると、言葉がなくても大丈夫な気がします。",
            "バルガスさん、自由になってから少し太りましたよね。……言わない方がいいですか？",
            "アスタロト様、『退屈』って言わなくなりました。……良いことだと思います。",
            "アイリスさんに『友達』って言われました。……嬉しいけど、疲れそうです。",
            "ゼクさんに『商売の才能ある』って言われました。……褒められてないですよね。",
            "また死んでしまったんですか？……心配で胸が痛いんです。本当に。",
            "あなたの死に方、パターンが読めてきました。……次は止めますね。",
            "呪われた装備、集めてどうするんですか……？趣味ですか……？",
            # PRE（違和感なし）
            "あなたの足音は、遠くからでもわかります。",
            "アイリスさん、元気すぎて見てるだけで疲れます。……少しだけ。",
            "バルガスさん、優しいの隠すの下手ですよね。……バレバレです。",
            "ゼクさんの商売、成功率低いです。……見てればわかります。",
            "ヌルさんに『似てる』って言われます。……私、あんなに正直じゃないです。",
            "あなた、死ぬ直前にいつも同じ顔しますよね。……覚えました。",
        ],
        texts_en=[
            "I decided where to go by myself for the first time. ...Scary, but happy.",
            "I had a dream. Can't remember what, but...I was happy.",
            "Being with Null, I feel okay even without words.",
            "Vargus has gained a bit of weight since becoming free. ...Should I not mention it?",
            "Astaroth stopped saying 'boring'. ...I think that's good.",
            "Iris called me a 'friend'. ...Happy, but sounds tiring.",
            "Zek said I have 'talent for business'. ...That's not a compliment, is it?",
            "You died again? ...My heart hurts with worry. Really.",
            "I'm starting to recognize your death patterns. ...I'll stop you next time.",
            "Why collect cursed equipment...? Is it a hobby...?",
            "I can recognize your footsteps from far away.",
            "Iris is too energetic. Just watching makes me tired. ...A little.",
            "Vargus is bad at hiding his kindness. ...It's so obvious.",
            "Zek's sales success rate is low. ...You can tell by watching.",
            "Null says we're 'similar'. ...I'm not that honest.",
            "You always make the same face right before dying. ...I've memorized it.",
        ],
        texts_cn=[
            "第一次自己决定去哪里。……虽然害怕，但很高兴。",
            "做梦了。虽然不记得内容……但很高兴。",
            "和努尔在一起，即使不说话也感觉没关系。",
            "巴尔加斯先生自由后胖了一点呢。……不说比较好吗？",
            "阿斯塔罗特大人不再说『无聊』了。……我觉得这是好事。",
            "艾莉丝小姐说我是『朋友』。……很高兴，但感觉会累。",
            "泽克先生说我『有做生意的才能』。……这不是在夸我吧。",
            "又死了吗？……我真的很担心，心都痛了。",
            "你的死法，我开始能预测了。……下次我会阻止你的。",
            "收集诅咒装备干什么……？是爱好吗……？",
            "你的脚步声，远远的我就能认出来。",
            "艾莉丝小姐太有活力了。光看着就累。……只是一点点。",
            "巴尔加斯先生不擅长隐藏温柔呢。……太明显了。",
            "泽克先生的生意成功率很低。……看着就知道。",
            "努尔说我们『很像』。……我没那么坦诚。",
            "你死前总是做同一个表情呢。……我记住了。",
        ],
    ),
    # ゼク（怪しい商人）- 調子いいペテン師
    UniqueDialogEntry(
        chara_id="sukutsu_shady_merchant",
        texts_jp=[
            # POST
            "この町の人々、全員の剥製を……いえ、何でもありません。",
            "アイリスさん、私を『ゼっさん』と呼ぶんですよ。……略すな。",
            "マニの信者は盗みが上手いとか。……商売敵ですね。",
            "ネフィアのアーティファクト、仕入れたいものです。見つけたら売ってください。",
            "『友人価格』でどうです？……通常価格の1.2倍です。気持ちの問題です。",
            "貴方には特別に教えますが、この情報自体が商品です。",
            "ヌルさんに『嘘つき』と言われました。……商売人ですよ、私は。",
            "貴方の死亡回数、何桁になりました？……記念に何か買いませんか？",
            # PRE（違和感なし）
            "いい買い物をすると、人は3日間幸せになれます。私の統計です。",
            "『呪われた品』と『祝福された品』、違いは持ち主が生き残ったかどうかです。",
            "貴方は面白い。死なないでくださいね。記録が途切れてしまう。",
            "お買い得ですよ！……何がって？全部ですよ、全部。",
            "返品？もちろん不可能です。",
            "今日だけの特別価格です。",
            "リリィさんに商売の才能があります。……あの目、完全に値踏みしてますよね。",
            "貴方の死に方のバリエーション、記録が追いつきません。図鑑作れますよ。",
        ],
        texts_en=[
            "I want to collect stuffed specimens of everyone in this town... No, nothing.",
            "Iris calls me 'Ze-san'. ...Don't shorten it.",
            "Mani's followers are skilled thieves, apparently. ...Business rivals.",
            "I'd like to stock Nefia artifacts. Sell them to me if you find any.",
            "'Friend price' how about it? ...1.2 times the normal price. It's about feelings.",
            "I'll tell you specially, but this information itself is merchandise.",
            "Null called me a 'liar'. ...I'm a merchant, not a liar.",
            "How many digits is your death count now? ...Want to buy something to celebrate?",
            "A good purchase makes one happy for 3 days. My statistics.",
            "'Cursed items' and 'blessed items' - the difference is whether the owner survived.",
            "You're interesting. Please don't die. My records would be incomplete.",
            "Great deal! ...What's great? Everything, everything.",
            "Returns? Absolutely impossible.",
            "Special price for today only.",
            "Lily has talent for business. ...Those eyes are definitely appraising things.",
            "Your death variations - I can't keep up with recording them. Could make an encyclopedia.",
        ],
        texts_cn=[
            "这个镇上的人，全部做成标本……不，没什么。",
            "艾莉丝叫我『泽桑』。……不要缩写。",
            "听说玛尼的信徒擅长偷盗。……是商业对手呢。",
            "想进一些奈菲亚的神器。找到的话卖给我吧。",
            "『朋友价』怎么样？……是正常价格的1.2倍。是心意的问题。",
            "特别告诉你，但这个情报本身就是商品。",
            "努尔说我是『骗子』。……我是商人，不是骗子。",
            "你的死亡次数有几位数了？……要不要买点什么纪念一下？",
            "好的购物能让人幸福3天。我的统计。",
            "『被诅咒的物品』和『被祝福的物品』，区别在于主人有没有活下来。",
            "你很有趣。请不要死。记录会中断的。",
            "超划算！……什么划算？全部都划算。",
            "退货？当然不可能。",
            "仅限今天的特别价格。",
            "莉莉小姐有做生意的才能。……那眼神，完全在估价呢。",
            "你死法的变化，记录都跟不上。能做成图鉴了。",
        ],
    ),
    # ヌル（暗殺人形）- 真実砲、淡々と本質を突く
    UniqueDialogEntry(
        chara_id="sukutsu_null",
        texts_jp=[
            # POST
            "空。……広い。……綺麗。",
            "お前の隣。……落ち着く。",
            "甘いもの。……好き。……もう一個、いい？",
            "猫、触りたい。……また逃げられた。",
            "お前の荷物。……重すぎ。……なぜ捨てない。",
            "お前と一緒にいると力が抜ける。……比喩じゃない。本当に抜けた。",
            "敵の時は強かったのに、味方になると弱くなるやつ。……いる。……ここに。",
            # POST（p2性格：アイリス経由の機能紹介）
            "足湯、バカにしてた。スタミナが回復する。……見直した。",
            "アイリスの暗闘訓練、受けてみた。暗闇でも遠くが見える。……やり方は荒いけど、効く。",
            "アイリスの足腰訓練、きつい。でも完全回避できる攻撃が増えた。……認める。",
            "ボスが強すぎるなら、アイリスに相談してみて。恥ずかしいことじゃない。……アイリスもそう言ってた。",
            "怪我ならアイリスに診せろ。呪いも毒も全部治る。……ただしバフも消える。注意。",
            "アイリスが言ってた。『神罰も治せるけど、すっごくお金かかるよ？』って。……どんな値段だろう。",
            "アイリスが『訓練サボるな』ってうるさい。……でも効果はある。3つとも受けるべき。",
            "アイリスと『特別な訓練』ができるらしい。……詳しくは知らない。知りたくもない。",
            "アイリスはうるさい。でも便利。治療も訓練も全部できる。……感謝はしてる。本人には言わないで。",
            "アイリスは目がいい。お前の怪我、一目でわかるって。……私もわかるけど。",
        ],
        texts_en=[
            "Sky. ...Wide. ...Beautiful.",
            "Next to you. ...Calm.",
            "Sweets. ...Like. ...One more, okay?",
            "Cat, want to touch. ...It ran away again.",
            "Your luggage. ...Too heavy. ...Why not drop some.",
            "Being with you drains my strength. ...Not a metaphor. It literally did.",
            "Bosses who get weak once they join you. ...Know any? ...You're looking at one.",
            "I underestimated the hot spring. It restores stamina. ...I stand corrected.",
            "Tried Iris's darkness training. Can see far even in the dark. ...Her methods are rough, but it works.",
            "Iris's leg training is brutal. But I'm fully dodging more attacks now. ...Acknowledged.",
            "If bosses are too tough, try talking to Iris. Nothing to be ashamed of. ...She said so too.",
            "If you're hurt, see Iris. Cures curses, poison, everything. ...But buffs vanish too. Fair warning.",
            "Iris mentioned she 'can even cure divine punishment, but it'll cost a LOT.' ...I wonder how much.",
            "Iris nags about 'don't skip training.' ...But it works. Do all three.",
            "Apparently you can do 'special training' with Iris. ...Don't know the details. Don't want to.",
            "Iris is loud. But useful. Healing, training, everything. ...I'm grateful. Don't tell her.",
            "Iris has good eyes. Says she can spot your injuries at a glance. ...So can I, though.",
        ],
        texts_cn=[
            "天空。……广阔。……漂亮。",
            "你的身边。……安心。",
            "甜食。……喜欢。……再来一个，可以吗？",
            "猫，想摸。……又跑掉了。",
            "你的行李。……太重。……为什么不扔掉。",
            "和你在一起会失去力量。……不是比喻。真的失去了。",
            "当敌人时很强，成为同伴就变弱的家伙。……有。……就在这里。",
            "小看了足汤。会恢复体力。……刮目相看。",
            "试了艾莉丝的暗斗训练。黑暗中也能看到远处。……方法虽然粗暴，但有效。",
            "艾莉丝的腿脚训练很辛苦。但能完全闪避的攻击增加了。……认可。",
            "Boss太强的话，试着去找艾莉丝吧。不丢人。……她也这么说的。",
            "受伤了就让艾莉丝看看。诅咒、毒什么都能治。……但增益也会消失。注意。",
            "艾莉丝说过『神罚也能治，不过要花好多钱哦？』……到底多少钱。",
            "艾莉丝一直唠叨『不要偷懒不训练』。……但确实有效。三种都做吧。",
            "好像可以和艾莉丝做『特别训练』。……不知道细节。也不想知道。",
            "艾莉丝很吵。但有用。治疗、训练什么都行。……很感激。别告诉她。",
            "艾莉丝眼力好。说一眼就能看出你的伤。……我也看得出来就是了。",
        ],
    ),
    # アスタロト（グランドマスター）- 時代錯誤おじいちゃん
    UniqueDialogEntry(
        chara_id="sukutsu_astaroth",
        texts_jp=[
            # POST
            "私を倒したお前と、こうして歩いている。……滑稽だが、悪くない。",
            "カラドリウスの風と、この風は違う。……だが、どちらも風だ。",
            "お前の死に様、私が見てきた中でもトップクラスだ。褒めていないぞ。",
            "アイリスに『エモい』と言われた。……30分かけて意味を聞いた。わからん。",
            # PRE（違和感なし）
            "最近の若者は礼儀がなっていない。……5万年間ずっとそう思っている。",
            "『草』とは何だ。……なぜ笑いを草と呼ぶ。植物だろう。",
            "お前たちの『1日』は短い。私の昼寝より短い。",
            "名前を覚えるのが苦手でな。……5万年で会った人間が多すぎる。",
        ],
        texts_en=[
            "Walking alongside you, who defeated me. ...Absurd, but not bad.",
            "The wind of Caladrius and this wind are different. ...But both are wind.",
            "Your death scenes are top-tier among all I've witnessed. Not a compliment.",
            "Iris said I was 'lit'. ...I asked what it meant for 30 minutes. Still don't get it.",
            "Young people these days have no manners. ...I've thought this for 50,000 years.",
            "What is 'lol'? ...Why call laughter a plant? It's vegetation.",
            "Your 'one day' is short. Shorter than my nap.",
            "I'm bad with names. ...Met too many people in 50,000 years.",
        ],
        texts_cn=[
            "和打败我的你一起走着。……荒唐，但不坏。",
            "卡拉德里乌斯的风和这里的风不同。……但都是风。",
            "你的死相，在我见过的里面也是顶级的。不是在夸你。",
            "艾莉丝说我『很酷』。……我问了30分钟什么意思。还是不懂。",
            "最近的年轻人没有礼貌。……我这样想了5万年。",
            "『草』是什么。……为什么把笑叫做草。是植物吧。",
            "你们的『一天』太短了。比我的午睡还短。",
            "我不擅长记名字。……5万年遇到的人太多了。",
        ],
    ),
]


def create_unique_sheet(wb: Workbook, lang: str = "en") -> None:
    """uniqueシートを作成（小文字）

    Args:
        wb: 対象Workbook
        lang: 言語コード。"en" or "cn"。
              textカラムにこの言語のテキストが入る。
    """
    ws = wb.create_sheet("unique")

    # 行1: カラムヘッダー（CWL形式）
    # CWLのdialog.xlsxではtextカラムが実際に表示されるテキスト。
    # 言語フォルダごとにtextの内容を切り替える。
    headers = ["id", "text", "text_JP", "text_EN"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # 行2-4: 空行（CWL形式）
    # 行5以降: データ
    row = 5
    for entry in UNIQUE_DIALOGS:
        text_jp = "\n".join(entry.texts_jp)
        text_en = "\n".join(entry.texts_en)
        if lang == "cn":
            text_display = "\n".join(entry.texts_cn)
        else:
            text_display = text_en

        ws.cell(row=row, column=1, value=entry.chara_id)
        ws.cell(row=row, column=2, value=text_display)
        ws.cell(row=row, column=3, value=text_jp)
        ws.cell(row=row, column=4, value=text_en)

        row += 1

    # カラム幅調整
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50


def create_dialog_excel(output_dir: Path, lang: str = "en") -> Path:
    """dialog.xlsxを生成

    Args:
        output_dir: 出力ディレクトリ
        lang: 言語コード ("en" or "cn")
    """
    wb = Workbook()

    # デフォルトシートを削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Uniqueシートを作成
    create_unique_sheet(wb, lang=lang)

    # 出力パス
    output_path = output_dir / "dialog.xlsx"
    wb.save(output_path)

    return output_path


def main():
    """メインエントリポイント"""
    # プロジェクトルートを取得
    project_root = Path(TOOLS_DIR).parent

    # LangMod/EN/Dialog
    en_dialog_dir = project_root / "LangMod" / "EN" / "Dialog"
    en_dialog_dir.mkdir(parents=True, exist_ok=True)
    en_output_path = create_dialog_excel(en_dialog_dir, lang="en")
    print(f"Generated: {en_output_path}")

    # LangMod/CN/Dialog（textカラムに中国語を入れて別途生成）
    cn_dialog_dir = project_root / "LangMod" / "CN" / "Dialog"
    cn_dialog_dir.mkdir(parents=True, exist_ok=True)
    cn_output_path = create_dialog_excel(cn_dialog_dir, lang="cn")
    print(f"Generated: {cn_output_path}")


if __name__ == "__main__":
    main()
