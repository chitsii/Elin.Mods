import openpyxl
import os
import csv

# パス設定
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)

# CWLサンプルファイルのパス（環境変数で上書き可能）
DEFAULT_SAMPLE_PATH = r"c:\Users\tishi\programming\elin_modding\CWL_AddLocation_Example\LangMod\EN\SourceSSS.xlsx"
SAMPLE_PATH = os.environ.get("CWL_SAMPLE_PATH", DEFAULT_SAMPLE_PATH)
OUTPUT_EN_TSV = os.path.join(PROJECT_ROOT, "LangMod", "EN", "Zone.tsv")
OUTPUT_CN_TSV = os.path.join(PROJECT_ROOT, "LangMod", "CN", "Zone.tsv")

print(f"Reading sample file from: {SAMPLE_PATH}")

# サンプル読み込み
sample_wb = openpyxl.load_workbook(SAMPLE_PATH)
sample_ws = sample_wb["Zone"]

# Row 1-3 (Header, Type, Default) Copy
header_rows = []
for r in range(1, 4):
    row_data = []
    for c in range(1, sample_ws.max_column + 1):
        cell_val = sample_ws.cell(row=r, column=c).value
        row_data.append(cell_val if cell_val is not None else "")
    header_rows.append(row_data)


# Zone定義（CNテキスト含む）
ZONE_DEFINITIONS = [
    {
        "id": "sukutsu_arena",
        "parent": "ntyris",
        "name_JP": "狭間の闘技場",
        "name": "The Rift Arena",
        "name_CN": "裂隙斗技场",
        "type": "Elin_SukutsuArena.Zone_SukutsuArena",
        "LV": 50,
        "chance": 100,
        "value": 100,
        "idFile": "sukutsu_arena",
        "idBiome": "Plain",
        "tag": "addMap,light",
        "image": "default",
        "pos": "2,-31,323",
        "textFlavor_JP": "時空の揺らぎを感じる。",
        "textFlavor": "You sense a ripple in spacetime.",
        "textFlavor_CN": "你感受到时空的波动。",
    },
    {
        "id": "field_fine",
        "parent": "sukutsu_arena",
        "name_JP": "闘技場フィールド",
        "name": "Arena Field",
        "name_CN": "斗技场战场",
        "type": "Elin_SukutsuArena.Zone_FieldFine",
        "LV": 1,
        "chance": 100,
        "idFile": "chitsii_battle_field_fine",
        "idBiome": "Plain",
        "tag": "addMap",
    },
    {
        "id": "field_snow",
        "parent": "sukutsu_arena",
        "name_JP": "雪原フィールド",
        "name": "Snow Field",
        "name_CN": "雪原战场",
        "type": "Elin_SukutsuArena.Zone_FieldSnow",
        "LV": 1,
        "chance": 100,
        "idFile": "chitsii_battle_field_snow",
        "idBiome": "Snow",
        "tag": "addMap",
    },
]

# カラムインデックスマッピング（サンプルから取得）
COLUMN_MAP = {
    "id": 0,
    "parent": 1,
    "name_JP": 2,
    "name": 3,
    "type": 4,
    "LV": 5,
    "chance": 6,
    "value": 8,
    "idFile": 10,
    "idBiome": 11,
    "idPlaylist": 13,
    "tag": 14,
    "image": 17,
    "pos": 18,
    "textFlavor_JP": 20,
    "textFlavor": 21,
}


def create_zone_row(zone_def, lang="en"):
    """Zone定義から行データを生成"""
    row = [""] * sample_ws.max_column

    for key, col_idx in COLUMN_MAP.items():
        value = zone_def.get(key, "")

        # CN版: サフィックスなしの列（name, textFlavor）に中国語を入れる
        if lang == "cn":
            if key == "name":
                value = zone_def.get("name_CN", zone_def.get("name", ""))
            elif key == "textFlavor":
                value = zone_def.get("textFlavor_CN", zone_def.get("textFlavor", ""))

        if value:
            row[col_idx] = value

    return row


def write_tsv(path, row_data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
        writer.writerows(row_data)
    print(f"Created TSV: {path}")


# EN版
rows_en = header_rows.copy()
for zone_def in ZONE_DEFINITIONS:
    rows_en.append(create_zone_row(zone_def, lang="en"))
write_tsv(OUTPUT_EN_TSV, rows_en)

# CN版
rows_cn = header_rows.copy()
for zone_def in ZONE_DEFINITIONS:
    rows_cn.append(create_zone_row(zone_def, lang="cn"))
write_tsv(OUTPUT_CN_TSV, rows_cn)

print(f"Generated {len(ZONE_DEFINITIONS)} zone(s) (EN + CN)")
