"""
diff_excel.py - Excel Diff Tool for Drama and Source Files

Compares two Excel files (or directories) and reports differences.
Uses smart line-based diff (difflib) to handle row insertions/deletions gracefully.

Features:
- Column names displayed instead of column letters (e.g., "action" instead of "Col D")
- Source Excel files show id/name format for better readability
- Drama Excel files show step/action/param format

Usage:
    # Compare two files
    python diff_excel.py before.xlsx after.xlsx

    # Compare two directories
    python diff_excel.py --dir before_dir after_dir

    # Compare specific drama files
    python diff_excel.py --drama drama_sukutsu_arena_master.xlsx

Exit codes:
    0 - Files are identical
    1 - Files differ
    2 - Error occurred
"""

import sys
import argparse
import difflib
from pathlib import Path
from typing import List, Tuple, Optional, Dict
import openpyxl


# File type detection
def is_source_excel(filepath: Path) -> bool:
    """Check if the file is a Source*.xlsx file."""
    return filepath.name.startswith("Source")


def is_drama_excel(filepath: Path) -> bool:
    """Check if the file is a drama Excel file."""
    return filepath.name.startswith("drama_")


def load_excel_data(filepath: Path) -> Tuple[dict, dict]:
    """
    Load Excel file and return all sheet data and headers as dicts.

    Returns:
        (data, headers) where:
        - data: {sheet_name: [[cell_values...]]}
        - headers: {sheet_name: [header_names...]}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {}
    headers = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        sheet_headers = []

        for row_idx, row in enumerate(ws.iter_rows()):
            row_data = [cell.value for cell in row]
            # Skip completely empty rows at the end
            if any(v is not None for v in row_data):
                rows.append(row_data)
                # First row is header
                if row_idx == 0:
                    sheet_headers = [normalize_value(cell.value) for cell in row]

        data[sheet_name] = rows
        headers[sheet_name] = sheet_headers

    wb.close()
    return data, headers


def normalize_value(value) -> str:
    """Normalize cell value for comparison."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Handle float/int comparison
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def row_to_string(row: list) -> str:
    """Convert a row to a comparable string representation."""
    return "\t".join(normalize_value(cell) for cell in row)


def get_row_key(row: list) -> str:
    """
    Get a key for identifying a row.
    For drama files, use step column (A) if present, otherwise the whole row.
    """
    if row and row[0]:
        return f"step:{normalize_value(row[0])}"
    return row_to_string(row)


def summarize_row_drama(row: list) -> str:
    """
    Create a brief summary of a row's content for drama files.
    Format: step, action, param (truncated)
    """
    parts = []

    # Column A: step
    step = normalize_value(row[0]) if len(row) > 0 else ""
    if step:
        parts.append(f"step:{step}")

    # Column D: action
    action = normalize_value(row[3]) if len(row) > 3 else ""
    if action:
        parts.append(f"action:{action}")

    # Column E: param (truncated)
    param = normalize_value(row[4]) if len(row) > 4 else ""
    if param:
        if len(param) > 50:
            param = param[:47] + "..."
        parts.append(f"param:{param}")

    return " | ".join(parts) if parts else "(empty row)"


def summarize_row_source(row: list, headers: List[str]) -> str:
    """
    Create a brief summary of a row's content for Source files.
    Format: id, name (using actual header names)
    """
    parts = []

    # Find id column (usually first column or named 'id')
    id_val = normalize_value(row[0]) if len(row) > 0 else ""
    id_name = headers[0] if headers else "id"
    if id_val:
        parts.append(f"{id_name}:{id_val}")

    # Find name column (look for 'name', 'name_JP', etc.)
    name_cols = ["name", "name_JP", "alias"]
    for col_name in name_cols:
        if col_name in headers:
            col_idx = headers.index(col_name)
            if col_idx < len(row):
                name_val = normalize_value(row[col_idx])
                if name_val:
                    if len(name_val) > 30:
                        name_val = name_val[:27] + "..."
                    parts.append(f"{col_name}:{name_val}")
                    break

    return " | ".join(parts) if parts else "(empty row)"


def summarize_row(row: list, headers: List[str] = None, is_source: bool = False) -> str:
    """
    Create a brief summary of a row's content.
    Dispatches to appropriate format based on file type.
    """
    if is_source and headers:
        return summarize_row_source(row, headers)
    else:
        return summarize_row_drama(row)


def compare_sheets(
    sheet1: List[list],
    sheet2: List[list],
    sheet_name: str,
    headers: List[str] = None,
    is_source: bool = False,
) -> List[str]:
    """
    Compare two sheets using smart line-based diff.
    Handles row insertions/deletions gracefully using difflib.
    """
    differences = []

    # Convert rows to strings for comparison (skip header row for Source files)
    start_row = 1 if is_source and len(sheet1) > 1 else 0
    lines1 = [row_to_string(row) for row in sheet1[start_row:]]
    lines2 = [row_to_string(row) for row in sheet2[start_row:]]

    # Use SequenceMatcher for smart diff
    matcher = difflib.SequenceMatcher(None, lines1, lines2)

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        # Adjust row numbers to account for header row
        actual_i1 = i1 + start_row
        actual_i2 = i2 + start_row
        actual_j1 = j1 + start_row
        actual_j2 = j2 + start_row

        if tag == "equal":
            continue
        elif tag == "delete":
            # Rows removed - show content
            for i in range(i1, i2):
                actual_row = actual_i1 + (i - i1)
                summary = summarize_row(sheet1[actual_row], headers, is_source)
                differences.append(f"    [{sheet_name}] Row {actual_row + 1} DELETED: {summary}")
        elif tag == "insert":
            # Rows added - show content
            for j in range(j1, j2):
                actual_row = actual_j1 + (j - j1)
                summary = summarize_row(sheet2[actual_row], headers, is_source)
                differences.append(
                    f"    [{sheet_name}] Row {actual_row + 1} INSERTED: {summary}"
                )
        elif tag == "replace":
            # Rows changed - show detailed diff for small changes
            old_count = i2 - i1
            new_count = j2 - j1

            if old_count == new_count and old_count <= 3:
                # Same number of rows, show cell-level changes
                for idx in range(old_count):
                    row1 = sheet1[actual_i1 + idx]
                    row2 = sheet2[actual_j1 + idx]
                    row_diffs = compare_row_cells(
                        row1, row2, actual_i1 + idx + 1, sheet_name, headers, is_source
                    )
                    differences.extend(row_diffs)
            else:
                # Different number of rows, show summary with content preview
                id1 = normalize_value(sheet1[actual_i1][0]) if sheet1[actual_i1] else ""
                id2 = normalize_value(sheet2[actual_j1][0]) if sheet2[actual_j1] else ""
                id_label = "id" if is_source else "step"
                id_info = ""
                if id1 or id2:
                    if id1 and id2 and id1 != id2:
                        id_info = f" ({id_label}: {id1}..{id2})"
                    elif id1:
                        id_info = f" ({id_label}: {id1})"
                    elif id2:
                        id_info = f" ({id_label}: {id2})"
                differences.append(
                    f"    [{sheet_name}] Rows {actual_i1 + 1}-{actual_i2} replaced with rows {actual_j1 + 1}-{actual_j2}{id_info}"
                )
                # Show first new row as preview
                if new_count > 0:
                    preview = summarize_row(sheet2[actual_j1], headers, is_source)
                    differences.append(f"      First new row: {preview}")

    return differences


def compare_row_cells(
    row1: list,
    row2: list,
    row_num: int,
    sheet_name: str,
    headers: List[str] = None,
    is_source: bool = False,
) -> List[str]:
    """Compare individual cells between two rows, using header names when available."""
    differences = []
    max_cols = max(len(row1), len(row2))

    changed_cells = []
    for col_idx in range(max_cols):
        val1 = row1[col_idx] if col_idx < len(row1) else None
        val2 = row2[col_idx] if col_idx < len(row2) else None

        norm1 = normalize_value(val1)
        norm2 = normalize_value(val2)

        if norm1 != norm2:
            # Use header name if available, otherwise fall back to column letter
            if headers and col_idx < len(headers) and headers[col_idx]:
                col_name = headers[col_idx]
            else:
                col_name = f"Col {openpyxl.utils.get_column_letter(col_idx + 1)}"

            # Truncate long values for readability
            disp1 = norm1 if len(norm1) <= 30 else norm1[:27] + "..."
            disp2 = norm2 if len(norm2) <= 30 else norm2[:27] + "..."
            changed_cells.append(f"{col_name}: '{disp1}' -> '{disp2}'")

    if changed_cells:
        # Get row identifier (id for Source, step for Drama)
        id_val = normalize_value(row1[0]) if row1 else ""
        if is_source:
            id_info = f" (id: {id_val})" if id_val else ""
        else:
            id_info = f" (step: {id_val})" if id_val else ""

        if len(changed_cells) <= 3:
            differences.append(
                f"  [{sheet_name}] Row {row_num}{id_info}: {'; '.join(changed_cells)}"
            )
        else:
            # Show first 2 changes and count
            shown = "; ".join(changed_cells[:2])
            differences.append(
                f"  [{sheet_name}] Row {row_num}{id_info}: {shown}; +{len(changed_cells) - 2} more"
            )

    return differences


def compare_excel_files(file1: Path, file2: Path) -> Tuple[bool, List[str]]:
    """
    Compare two Excel files.

    Returns:
        (identical: bool, differences: List[str])
    """
    try:
        data1, headers1 = load_excel_data(file1)
        data2, headers2 = load_excel_data(file2)
    except Exception as e:
        return False, [f"Error loading files: {e}"]

    # Detect file type
    is_source = is_source_excel(file1) or is_source_excel(file2)

    differences = []

    # Check sheet names
    sheets1 = set(data1.keys())
    sheets2 = set(data2.keys())

    only_in_1 = sheets1 - sheets2
    only_in_2 = sheets2 - sheets1

    if only_in_1:
        differences.append(f"  Sheets removed: {only_in_1}")
    if only_in_2:
        differences.append(f"  Sheets added: {only_in_2}")

    # Compare common sheets
    common_sheets = sheets1 & sheets2
    for sheet_name in sorted(common_sheets):
        # Use headers from file2 (current) as they may have been updated
        headers = headers2.get(sheet_name, headers1.get(sheet_name, []))
        sheet_diffs = compare_sheets(
            data1[sheet_name], data2[sheet_name], sheet_name, headers, is_source
        )
        differences.extend(sheet_diffs)

    return len(differences) == 0, differences


def compare_directories(
    dir1: Path, dir2: Path, pattern: str = "*.xlsx"
) -> Tuple[bool, dict]:
    """
    Compare all Excel files in two directories.

    Returns:
        (all_identical: bool, {filename: (identical, differences)})
    """
    files1 = {f.name: f for f in dir1.glob(pattern)}
    files2 = {f.name: f for f in dir2.glob(pattern)}

    results = {}
    all_identical = True

    # Files only in dir1
    for name in sorted(files1.keys() - files2.keys()):
        results[name] = (False, [f"Only exists in {dir1}"])
        all_identical = False

    # Files only in dir2
    for name in sorted(files2.keys() - files1.keys()):
        results[name] = (False, [f"Only exists in {dir2}"])
        all_identical = False

    # Compare common files
    for name in sorted(files1.keys() & files2.keys()):
        identical, diffs = compare_excel_files(files1[name], files2[name])
        results[name] = (identical, diffs)
        if not identical:
            all_identical = False

    return all_identical, results


def main():
    parser = argparse.ArgumentParser(
        description="Compare Excel files for drama migration verification"
    )
    parser.add_argument("path1", help="First file or directory")
    parser.add_argument("path2", help="Second file or directory")
    parser.add_argument(
        "--dir", action="store_true", help="Compare directories instead of files"
    )
    parser.add_argument(
        "--drama",
        metavar="NAME",
        help="Compare specific drama file in LangMod/EN/Dialog/Drama/",
    )
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="Show all comparisons, not just differences",
    )

    args = parser.parse_args()

    if args.drama:
        # Compare specific drama file between two base directories
        base1 = Path(args.path1)
        base2 = Path(args.path2)
        drama_path = Path("LangMod/EN/Dialog/Drama") / args.drama

        file1 = base1 / drama_path
        file2 = base2 / drama_path

        if not file1.exists():
            print(f"Error: {file1} does not exist")
            sys.exit(2)
        if not file2.exists():
            print(f"Error: {file2} does not exist")
            sys.exit(2)

        identical, diffs = compare_excel_files(file1, file2)

        if identical:
            print(f"[OK] {args.drama}: Files are identical")
            sys.exit(0)
        else:
            print(f"[DIFF] {args.drama}: Files differ")
            for diff in diffs[:20]:  # Limit output
                print(diff)
            if len(diffs) > 20:
                print(f"  ... and {len(diffs) - 20} more differences")
            sys.exit(1)

    elif args.dir:
        # Compare directories
        dir1 = Path(args.path1)
        dir2 = Path(args.path2)

        if not dir1.is_dir():
            print(f"Error: {dir1} is not a directory")
            sys.exit(2)
        if not dir2.is_dir():
            print(f"Error: {dir2} is not a directory")
            sys.exit(2)

        all_identical, results = compare_directories(dir1, dir2)

        for name, (identical, diffs) in results.items():
            if identical:
                if args.verbose:
                    print(f"[OK] {name}")
            else:
                print(f"[DIFF] {name}")
                for diff in diffs[:10]:
                    print(diff)
                if len(diffs) > 10:
                    print(f"  ... and {len(diffs) - 10} more differences")

        if all_identical:
            print(f"\nAll {len(results)} files are identical")
            sys.exit(0)
        else:
            diff_count = sum(
                1 for _, (identical, _) in results.items() if not identical
            )
            print(f"\n{diff_count}/{len(results)} files differ")
            sys.exit(1)

    else:
        # Compare two files directly
        file1 = Path(args.path1)
        file2 = Path(args.path2)

        if not file1.exists():
            print(f"Error: {file1} does not exist")
            sys.exit(2)
        if not file2.exists():
            print(f"Error: {file2} does not exist")
            sys.exit(2)

        identical, diffs = compare_excel_files(file1, file2)

        if identical:
            print(f"Files are identical")
            sys.exit(0)
        else:
            print(f"Files differ:")
            for diff in diffs:
                print(diff)
            sys.exit(1)


if __name__ == "__main__":
    main()
