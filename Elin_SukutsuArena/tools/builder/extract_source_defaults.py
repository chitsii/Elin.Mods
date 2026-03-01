# -*- coding: utf-8 -*-
"""
extract_source_defaults.py - バニラSource Excelのデフォルト値抽出・比較ツール

MODのビルダースクリプトで定義しているデフォルト値が、
バニラのSource Excelと一致しているか機械的に検証する。

Usage:
    # MODビルダーとバニラを比較（JSONキャッシュ使用、高速）
    uv run python builder/extract_source_defaults.py compare

    # 特定シートのみ比較
    uv run python builder/extract_source_defaults.py compare --sheets Thing,Element

    # Google DocsからExcelをダウンロードしてキャッシュを再生成
    uv run python builder/extract_source_defaults.py sync

    # バニラデータをJSONにキャッシュ（ローカルExcelから）
    uv run python builder/extract_source_defaults.py extract

    # 特定シートのみ抽出
    uv run python builder/extract_source_defaults.py extract --sheets Thing,Element,Stat

Note:
    compare コマンドはまずJSONキャッシュ（vanilla_defaults.json）を使用します。
    キャッシュがない場合はExcelから直接読み込みます（遅い）。
    ゲームアップデート後は sync コマンドでキャッシュを再生成してください。
"""

import argparse
import importlib.util
import json
import os
import sys
import urllib.request
from typing import Any

import openpyxl

# パス設定
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)

# バニラSourceExcelsのパス
SOURCE_EXCELS_PATH = os.path.join(PROJECT_ROOT, "..", "SourceExcels")

# 出力先
OUTPUT_JSON = os.path.join(BUILDER_DIR, "vanilla_defaults.json")

# Excelファイルとシートの対応
EXCEL_SHEETS = {
    "SourceCard.xlsx": [
        "Thing",
        "ThingV",
        "Food",
        "Recipe",
        "SpawnList",
        "Category",
        "Collectible",
        "KeyItem",
    ],
    "SourceGame.xlsx": [
        "Element",
        "Calc",
        "Stat",
        "Check",
        "Faction",
        "Religion",
        "Zone",
        "ZoneAffix",
        "Quest",
        "Area",
        "HomeResource",
        "Research",
        "Person",
    ],
    "SourceChara.xlsx": ["Chara", "CharaText", "Tactics", "Race", "Job", "Hobby"],
    "SourceBlock.xlsx": [
        "GlobalTile",
        "Block",
        "Floor",
        "Obj",
        "CellEffect",
        "Material",
    ],
    "Lang.xlsx": ["General", "Game", "Note", "List", "Word"],
}

# MODビルダーとシートの対応
MOD_BUILDERS = {
    "Thing": "create_thing_excel.py",
    "Element": "create_element_excel.py",
    "Stat": "create_stat_excel.py",
    "Quest": "create_quest_excel.py",
    "Zone": "create_zone_excel.py",
    "Chara": "create_chara_excel.py",
}

# Google Docs File IDs（ExcelファイルとGoogle Docs IDの対応）
# URL形式: https://docs.google.com/spreadsheets/d/{FILE_ID}/edit
GOOGLE_DOCS_IDS = {
    "SourceCard.xlsx": "175DaEeB-8qU3N4iBTnaal1ZcP5SU6S_Z",
    "Lang.xlsx": "1cje2GHiKwjBd_YLYWqWlddm2YLsYnRiB",
    "SourceBlock.xlsx": "13oxL_cQEqoTUlcWsjKZyNuAaITFGK56v",
    "SourceChara.xlsx": "1CJqsXFF2FLlpPz710oCpNFYF4W_5yoVn",
    "SourceGame.xlsx": "16-LkHtVqjuN9U0rripjBn-nYwyqqSGg_",
}


# ============================================================
# 共通ユーティリティ
# ============================================================

def _trim_trailing_empty(lst: list[str]) -> list[str]:
    """末尾の空文字列を除去"""
    while lst and lst[-1] == "":
        lst.pop()
    return lst


def _load_vanilla_cache() -> dict[str, dict] | None:
    """JSONキャッシュからバニラデータを読み込む"""
    if not os.path.exists(OUTPUT_JSON):
        return None
    try:
        with open(OUTPUT_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _get_vanilla_from_cache(cache: dict, sheet_name: str) -> list[dict[str, str]] | None:
    """キャッシュからシートデータを辞書リスト形式で取得"""
    if sheet_name not in cache:
        return None
    data = cache[sheet_name]
    columns = data.get("columns", [])
    types = data.get("types", [])
    defaults = data.get("defaults", [])
    return _to_column_dicts(columns, types, defaults)


def _to_column_dicts(
    columns: list[str], types: list[str], defaults: list[str]
) -> list[dict[str, str]]:
    """ヘッダー・型・デフォルトのリストを辞書リストに正規化"""
    result = []
    for i, name in enumerate(columns):
        result.append({
            "name": name,
            "type": types[i] if i < len(types) else "",
            "default": defaults[i] if i < len(defaults) else "",
        })
    return result


# ============================================================
# バニラExcel読み込み
# ============================================================

def find_source_file(sheet_name: str) -> str | None:
    """シート名から対応するExcelファイルを検索"""
    for excel_file, sheets in EXCEL_SHEETS.items():
        if sheet_name in sheets:
            return excel_file
    return None


def read_vanilla_sheet(sheet_name: str) -> list[dict[str, str]] | None:
    """バニラExcelからシートを読み込み、辞書リストを返す"""
    excel_file = find_source_file(sheet_name)
    if not excel_file:
        print(f"  Warning: No source file found for sheet '{sheet_name}'")
        return None

    excel_path = os.path.join(SOURCE_EXCELS_PATH, excel_file)
    if not os.path.exists(excel_path):
        print(f"  Warning: Excel file not found: {excel_path}")
        return None

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: Sheet '{sheet_name}' not found in {excel_file}")
            wb.close()
            return None

        ws = wb[sheet_name]
        max_col = ws.max_column

        def read_row(row_num: int) -> list[str]:
            return [
                str(ws.cell(row=row_num, column=c).value)
                if ws.cell(row=row_num, column=c).value is not None
                else ""
                for c in range(1, max_col + 1)
            ]

        columns = _trim_trailing_empty(read_row(1))
        types = read_row(2)[: len(columns)]
        defaults = read_row(3)[: len(columns)]

        wb.close()
        return _to_column_dicts(columns, types, defaults)

    except Exception as e:
        print(f"  Error reading {excel_path}/{sheet_name}: {e}")
        return None


# ============================================================
# MODビルダー読み込み
# ============================================================

def read_mod_builder(builder_file: str) -> list[dict[str, str]] | None:
    """MODビルダースクリプトからカラム定義を読み込み、辞書リストを返す

    パターンA (Thing, Element, Stat, Quest): HEADERS/TYPES/DEFAULTS 変数
    パターンB (Zone, Chara): モジュール実行後の header_rows / rows 変数
    """
    builder_path = os.path.join(BUILDER_DIR, builder_file)
    if not os.path.exists(builder_path):
        print(f"  Warning: Builder not found: {builder_path}")
        return None

    try:
        spec = importlib.util.spec_from_file_location("builder_module", builder_path)
        if spec is None or spec.loader is None:
            return None

        module = importlib.util.module_from_spec(spec)

        # 副作用を抑制: argv をリセットし、stdout を抑制
        original_argv = sys.argv
        sys.argv = [builder_path]

        original_stdout = sys.stdout
        sys.stdout = open(os.devnull, "w")

        try:
            spec.loader.exec_module(module)
        except SystemExit:
            pass
        finally:
            sys.stdout.close()
            sys.stdout = original_stdout
            sys.argv = original_argv

        # パターンA: HEADERS/TYPES/DEFAULTS をモジュールレベル変数として定義
        if hasattr(module, "HEADERS"):
            headers = module.HEADERS
            if isinstance(headers, list):
                columns = [str(h) for h in headers]
            elif isinstance(headers, dict):
                columns = list(headers.keys())
            else:
                columns = []

            # TYPES
            types: list[str] = []
            if hasattr(module, "TYPES"):
                raw_types = module.TYPES
                if isinstance(raw_types, list):
                    types = [str(t) for t in raw_types]
                elif isinstance(raw_types, dict):
                    types = [str(v) for v in raw_types.values()]

            # DEFAULTS
            defaults: list[str] = []
            if hasattr(module, "DEFAULTS"):
                raw_defaults = module.DEFAULTS
                if isinstance(raw_defaults, list):
                    defaults = [str(d) if d is not None else "" for d in raw_defaults]
                elif isinstance(raw_defaults, dict):
                    defaults = [
                        str(raw_defaults.get(col, "")) for col in columns
                    ]

            return _to_column_dicts(columns, types, defaults)

        # パターンB: header_rows / rows 変数から取得
        header_rows_data = None
        if hasattr(module, "header_rows"):
            header_rows_data = module.header_rows
        elif hasattr(module, "rows"):
            rows_data = module.rows
            if isinstance(rows_data, list) and len(rows_data) >= 3:
                header_rows_data = rows_data[:3]

        if header_rows_data and len(header_rows_data) >= 3:
            columns = _trim_trailing_empty(
                [str(v) if v is not None else "" for v in header_rows_data[0]]
            )
            types = [str(v) if v is not None else "" for v in header_rows_data[1]][
                : len(columns)
            ]
            defaults = [str(v) if v is not None else "" for v in header_rows_data[2]][
                : len(columns)
            ]
            return _to_column_dicts(columns, types, defaults)

        print(f"  Warning: No HEADERS or header_rows found in {builder_file}")
        return None

    except Exception as e:
        print(f"  Error loading {builder_file}: {e}")
        import traceback
        traceback.print_exc()
        return None


# ============================================================
# 比較ロジック
# ============================================================

def compare_columns(
    sheet_name: str,
    vanilla: list[dict[str, str]],
    mod: list[dict[str, str]],
) -> list[str]:
    """バニラとMODの辞書リストを厳密比較してレポートを生成"""
    issues = []

    # 1. カラム数の比較
    if len(vanilla) != len(mod):
        issues.append(
            f"[カラム数] バニラ: {len(vanilla)}, MOD: {len(mod)}"
        )

    # 2. 各位置で name, type, default を比較
    min_cols = min(len(vanilla), len(mod))
    name_mismatches = []
    type_mismatches = []
    default_mismatches = []

    for i in range(min_cols):
        v = vanilla[i]
        m = mod[i]

        if v["name"] != m["name"]:
            name_mismatches.append(
                f"  Col {i+1}: バニラ='{v['name']}', MOD='{m['name']}'"
            )

        if v["type"] != m["type"]:
            type_mismatches.append(
                f"  Col {i+1} ({v['name']}): バニラ='{v['type']}', MOD='{m['type']}'"
            )

        if v["default"] != m["default"]:
            default_mismatches.append(
                f"  Col {i+1} ({v['name']}): バニラ='{v['default']}', MOD='{m['default']}'"
            )

    if name_mismatches:
        issues.append("[カラム名の差異]")
        issues.extend(name_mismatches)

    if type_mismatches:
        issues.append("[型の差異]")
        issues.extend(type_mismatches)

    if default_mismatches:
        issues.append("[デフォルト値の差異]")
        issues.extend(default_mismatches)

    return issues


# ============================================================
# コマンド: compare
# ============================================================

def cmd_compare(sheets_filter: list[str] | None = None, quiet: bool = False) -> bool:
    """MODビルダーとバニラを比較（JSONキャッシュ優先）

    Returns:
        bool: 全シートで差異がなければTrue
    """
    # JSONキャッシュを試みる
    cache = _load_vanilla_cache()
    use_cache = cache is not None

    if not quiet:
        if use_cache:
            print("\n=== MODビルダー vs バニラ（キャッシュ）比較 ===\n")
        else:
            print("\n=== MODビルダー vs バニラ Source Excel 比較 ===\n")
            print("  Note: キャッシュなし。'extract' コマンドで高速化可能\n")

    targets = sheets_filter if sheets_filter else list(MOD_BUILDERS.keys())
    all_ok = True

    for sheet_name in targets:
        if sheet_name not in MOD_BUILDERS:
            if not quiet:
                print(f"--- {sheet_name}: MODビルダーなし、スキップ ---\n")
            continue

        builder_file = MOD_BUILDERS[sheet_name]
        if not quiet:
            print(f"--- {sheet_name} ({builder_file}) ---")

        # バニラ読み込み（キャッシュ優先）
        vanilla = None
        if use_cache:
            vanilla = _get_vanilla_from_cache(cache, sheet_name)
        if vanilla is None:
            # キャッシュになければExcelから読み込み
            vanilla = read_vanilla_sheet(sheet_name)

        if vanilla is None:
            if not quiet:
                print("  バニラデータの読み込みに失敗\n")
            all_ok = False
            continue

        # MODビルダー読み込み
        mod = read_mod_builder(builder_file)
        if mod is None:
            if not quiet:
                print("  MODビルダーの読み込みに失敗\n")
            all_ok = False
            continue

        if not quiet:
            print(f"  バニラ: {len(vanilla)} カラム, MOD: {len(mod)} カラム")

        # 比較
        issues = compare_columns(sheet_name, vanilla, mod)

        # Chara: CWL拡張カラム（Author, portrait）の差異は意図的
        if sheet_name == "Chara" and len(mod) > len(vanilla):
            extra_cols = [m["name"] for m in mod[len(vanilla):]]
            if set(extra_cols) == {"Author", "portrait"}:
                # カラム数差異のみの場合はOKとみなす
                issues = [i for i in issues if not i.startswith("[カラム数]")]
                if not issues:
                    if not quiet:
                        print("  OK (CWL拡張カラム Author, portrait を追加)")
                else:
                    for issue in issues:
                        print(f"  {issue}")
                    if not quiet:
                        print("  ※ CWL拡張カラム (Author, portrait) は意図的な追加")
                if not quiet:
                    print()
                continue

        if issues:
            all_ok = False
            for issue in issues:
                print(f"  {issue}")
        else:
            if not quiet:
                print("  OK")

        if not quiet:
            print()

    if not quiet:
        if all_ok:
            print("全シートで差異なし")
        else:
            print("差異が検出されました（上記参照）")

    return all_ok


# ============================================================
# コマンド: sync（Google Docsからダウンロード＋キャッシュ再生成）
# ============================================================

def download_from_google_docs(file_id: str, output_path: str) -> bool:
    """Google DocsからExcelファイルをダウンロード"""
    export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    try:
        # User-Agentを設定（Google Docsはブラウザ以外からのアクセスをブロックすることがある）
        request = urllib.request.Request(
            export_url,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        )
        with urllib.request.urlopen(request, timeout=60) as response:
            with open(output_path, "wb") as f:
                f.write(response.read())
        return True
    except Exception as e:
        print(f"  Error downloading: {e}")
        return False


def cmd_sync(files_filter: list[str] | None = None) -> bool:
    """Google DocsからExcelをダウンロードしてキャッシュを再生成

    Returns:
        bool: 全てのダウンロードと抽出が成功したらTrue
    """
    print("\n=== Google Docsからバニラ Source Excel を同期 ===\n")

    # ダウンロード対象を決定（デフォルトは全ファイル）
    if files_filter:
        targets = {f: fid for f, fid in GOOGLE_DOCS_IDS.items() if f in files_filter}
    else:
        targets = GOOGLE_DOCS_IDS

    if not targets:
        print("  ダウンロード対象のファイルがありません")
        return False

    # SourceExcelsディレクトリを作成
    os.makedirs(SOURCE_EXCELS_PATH, exist_ok=True)

    # ダウンロード
    print(f"Downloading {len(targets)} file(s) from Google Docs...")
    all_ok = True
    for excel_file, file_id in targets.items():
        output_path = os.path.join(SOURCE_EXCELS_PATH, excel_file)
        print(f"  {excel_file}...", end=" ", flush=True)
        if download_from_google_docs(file_id, output_path):
            print("OK")
        else:
            print("FAILED")
            all_ok = False

    if not all_ok:
        print("\n一部のダウンロードに失敗しました")
        return False

    # キャッシュを再生成（MODで使用するシートのみ）
    print("\nRegenerating cache...")
    target_sheets = list(MOD_BUILDERS.keys())
    cmd_extract(target_sheets)

    return True


# ============================================================
# コマンド: extract（サブ機能）
# ============================================================

def cmd_extract(sheets_filter: list[str] | None = None) -> None:
    """バニラデータをJSONに保存"""
    result = {}

    if sheets_filter:
        target_sheets = sheets_filter
    else:
        target_sheets = []
        for sheets in EXCEL_SHEETS.values():
            target_sheets.extend(sheets)

    print(f"Extracting defaults from {len(target_sheets)} sheets...")

    for sheet_name in target_sheets:
        excel_file = find_source_file(sheet_name)
        if not excel_file:
            print(f"  Warning: No source file found for sheet '{sheet_name}'")
            continue

        excel_path = os.path.join(SOURCE_EXCELS_PATH, excel_file)
        if not os.path.exists(excel_path):
            print(f"  Warning: Excel file not found: {excel_path}")
            continue

        print(f"  Extracting {sheet_name} from {excel_file}...")
        column_dicts = read_vanilla_sheet(sheet_name)
        if column_dicts:
            # JSON保存用にフラット形式に変換
            result[sheet_name] = {
                "source_file": excel_file,
                "columns": [c["name"] for c in column_dicts],
                "types": [c["type"] for c in column_dicts],
                "defaults": [c["default"] for c in column_dicts],
            }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\nSaved to: {OUTPUT_JSON}")
    print(f"Extracted {len(result)} sheet(s)")


# ============================================================
# メイン
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="バニラSource Excelのデフォルト値抽出・比較ツール"
    )
    subparsers = parser.add_subparsers(dest="command", help="コマンド")

    # compare コマンド（デフォルト）
    compare_parser = subparsers.add_parser("compare", help="バニラExcelとMODビルダーを直接比較")
    compare_parser.add_argument(
        "--sheets",
        type=str,
        help="比較するシート名（カンマ区切り、例: Thing,Element）",
    )
    compare_parser.add_argument(
        "-q", "--quiet",
        action="store_true",
        help="成功時は出力を抑制（ビルド組み込み用）",
    )

    # extract コマンド
    extract_parser = subparsers.add_parser("extract", help="バニラデータをJSONに保存")
    extract_parser.add_argument(
        "--sheets",
        type=str,
        help="抽出するシート名（カンマ区切り、例: Thing,Element,Stat）",
    )

    # sync コマンド
    sync_parser = subparsers.add_parser(
        "sync",
        help="Google DocsからExcelをダウンロードしてキャッシュを再生成"
    )
    sync_parser.add_argument(
        "--files",
        type=str,
        help="ダウンロードするファイル名（カンマ区切り、例: SourceGame.xlsx,SourceChara.xlsx）",
    )

    args = parser.parse_args()

    # デフォルトは compare
    command = args.command or "compare"

    sheets_filter = None
    if hasattr(args, "sheets") and args.sheets:
        sheets_filter = [s.strip() for s in args.sheets.split(",")]

    files_filter = None
    if hasattr(args, "files") and args.files:
        files_filter = [f.strip() for f in args.files.split(",")]

    quiet = getattr(args, "quiet", False)

    if command == "compare":
        success = cmd_compare(sheets_filter, quiet=quiet)
        sys.exit(0 if success else 1)
    elif command == "extract":
        cmd_extract(sheets_filter)
        sys.exit(0)
    elif command == "sync":
        success = cmd_sync(files_filter)
        sys.exit(0 if success else 1)
    else:
        parser.print_help()
        sys.exit(0)


if __name__ == "__main__":
    main()
