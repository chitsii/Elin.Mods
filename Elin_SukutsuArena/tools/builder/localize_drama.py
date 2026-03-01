# -*- coding: utf-8 -*-
"""
ドラマExcelのtext列を言語に応じて設定するスクリプト

ElinのDramaManagerは非組み込み言語でdictLocalizeを構築する際、
言語フォルダからtext列の値を読み込む。このスクリプトはEN版を
CN/JPにコピーし、text列を該当言語に設定する。
"""

import shutil
from pathlib import Path

import openpyxl


def localize_drama_file(src_path: Path, dest_path: Path, lang: str) -> int:
    """
    ドラマExcelファイルをコピーし、text列を指定言語に設定する

    Returns:
        更新された行数
    """
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src_path, dest_path)

    wb = openpyxl.load_workbook(dest_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    try:
        text_col = headers.index("text") + 1
        text_lang_col = headers.index(f"text_{lang}") + 1
    except ValueError:
        wb.close()
        return 0

    updated = 0
    for row in ws.iter_rows(min_row=2):
        text_lang_value = row[text_lang_col - 1].value
        if text_lang_value:
            row[text_col - 1].value = text_lang_value
            updated += 1

    wb.save(dest_path)
    wb.close()
    return updated


def main():
    project_root = Path(__file__).parent.parent.parent
    en_drama_dir = project_root / "LangMod" / "EN" / "Dialog" / "Drama"

    if not en_drama_dir.exists():
        print(f"Error: EN Drama directory not found")
        return

    for lang in ["CN", "JP"]:
        lang_drama_dir = project_root / "LangMod" / lang / "Dialog" / "Drama"
        total = 0
        for xlsx_file in en_drama_dir.glob("*.xlsx"):
            updated = localize_drama_file(xlsx_file, lang_drama_dir / xlsx_file.name, lang)
            total += updated
        print(f"  {lang}: {total} rows localized")


if __name__ == "__main__":
    main()
