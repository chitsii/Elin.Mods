# -*- coding: utf-8 -*-
"""
validate_source_excel.py - 生成されたSource Excelのヘッダー検証

生成されたSource*.xlsxの最初の3行（ヘッダー、型、デフォルト）が
バニラのSource Excelと一致するか検証する。

Usage:
    # 全Source Excelを検証
    uv run python builder/validate_source_excel.py

    # 詳細表示
    uv run python builder/validate_source_excel.py -v

Exit codes:
    0 - 全てパス
    1 - 検証エラー
"""

import argparse
import os
import sys

import openpyxl

# パス設定
BUILDER_DIR = os.path.dirname(os.path.abspath(__file__))
TOOLS_DIR = os.path.dirname(BUILDER_DIR)
PROJECT_ROOT = os.path.dirname(TOOLS_DIR)

# バニラSourceExcelsのパス
SOURCE_EXCELS_PATH = os.path.join(PROJECT_ROOT, "..", "SourceExcels")

# MOD生成Excelのパス
MOD_LANGMOD_PATH = os.path.join(PROJECT_ROOT, "LangMod", "EN")

# 検証対象のマッピング: (MODファイル名, バニラファイル名, シート名)
VALIDATION_TARGETS = [
    ("SourceChara.xlsx", "SourceChara.xlsx", "Chara"),
    ("SourceThing.xlsx", "SourceCard.xlsx", "Thing"),
    ("SourceElement.xlsx", "SourceGame.xlsx", "Element"),
    ("SourceStat.xlsx", "SourceGame.xlsx", "Stat"),
    ("SourceQuest.xlsx", "SourceGame.xlsx", "Quest"),
    ("SourceSukutsu.xlsx", "SourceGame.xlsx", "Zone"),
]


def read_header_rows(filepath: str, sheet_name: str, num_rows: int = 3) -> list[list[str]]:
    """Excelファイルからヘッダー行を読み取る"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found in {filepath}")

    ws = wb[sheet_name]
    max_col = ws.max_column

    rows = []
    for row_num in range(1, num_rows + 1):
        row_data = []
        for col in range(1, max_col + 1):
            val = ws.cell(row=row_num, column=col).value
            row_data.append(str(val) if val is not None else "")
        # 末尾の空文字を除去
        while row_data and row_data[-1] == "":
            row_data.pop()
        rows.append(row_data)

    wb.close()
    return rows


def compare_header_rows(
    mod_rows: list[list[str]],
    vanilla_rows: list[list[str]],
    mod_file: str,
    vanilla_file: str,
    sheet_name: str,
) -> list[str]:
    """ヘッダー行を比較してエラーメッセージを返す"""
    issues = []

    # Chara: CWL拡張カラム（Author, portrait）は許容
    is_chara = sheet_name == "Chara"

    for row_idx in range(min(len(mod_rows), len(vanilla_rows))):
        mod_row = mod_rows[row_idx]
        vanilla_row = vanilla_rows[row_idx]

        # Charaの場合、バニラの列数までを比較
        compare_len = len(vanilla_row)

        for col_idx in range(compare_len):
            mod_val = mod_row[col_idx] if col_idx < len(mod_row) else ""
            vanilla_val = vanilla_row[col_idx] if col_idx < len(vanilla_row) else ""

            if mod_val != vanilla_val:
                row_name = ["Header", "Type", "Default"][row_idx] if row_idx < 3 else f"Row{row_idx+1}"
                col_name = vanilla_row[col_idx] if row_idx > 0 and col_idx < len(vanilla_rows[0]) else f"Col{col_idx+1}"
                issues.append(
                    f"  {row_name} row, {col_name}: MOD='{mod_val}' vs Vanilla='{vanilla_val}'"
                )

    # カラム数チェック（Chara以外）
    if not is_chara:
        if len(mod_rows[0]) != len(vanilla_rows[0]):
            issues.append(
                f"  Column count mismatch: MOD={len(mod_rows[0])}, Vanilla={len(vanilla_rows[0])}"
            )

    return issues


def validate_all(verbose: bool = False) -> bool:
    """全てのSource Excelを検証"""
    all_ok = True

    for mod_file, vanilla_file, sheet_name in VALIDATION_TARGETS:
        mod_path = os.path.join(MOD_LANGMOD_PATH, mod_file)
        vanilla_path = os.path.join(SOURCE_EXCELS_PATH, vanilla_file)

        if not os.path.exists(mod_path):
            if verbose:
                print(f"[SKIP] {mod_file}: MOD file not found")
            continue

        if not os.path.exists(vanilla_path):
            if verbose:
                print(f"[SKIP] {mod_file}: Vanilla file not found ({vanilla_file})")
            continue

        try:
            mod_rows = read_header_rows(mod_path, sheet_name)
            vanilla_rows = read_header_rows(vanilla_path, sheet_name)

            issues = compare_header_rows(mod_rows, vanilla_rows, mod_file, vanilla_file, sheet_name)

            if issues:
                all_ok = False
                print(f"[FAIL] {mod_file} ({sheet_name}):")
                for issue in issues:
                    print(issue)
            elif verbose:
                print(f"[OK] {mod_file} ({sheet_name})")

        except Exception as e:
            all_ok = False
            print(f"[ERROR] {mod_file}: {e}")

    return all_ok


def main():
    parser = argparse.ArgumentParser(description="生成されたSource Excelのヘッダー検証")
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="詳細表示",
    )
    args = parser.parse_args()

    success = validate_all(verbose=args.verbose)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
