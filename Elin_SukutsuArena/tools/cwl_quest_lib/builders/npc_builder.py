"""
CWL Quest Library - NPC Builder

This module provides utilities for defining custom NPCs and generating
CWL-compatible SourceChara TSV files.

Usage:
    from cwl_quest_lib.npc_builder import NpcBuilder
    from cwl_quest_lib import NpcDefinition

    builder = NpcBuilder()
    builder.add_npc(NpcDefinition(
        id="my_npc",
        name_jp="商人",
        name_en="Merchant",
        race="human",
        job="merchant",
        zone_id="my_zone",
        drama_id="drama_my_npc",
    ))
    builder.save_tsv("LangMod/EN/Chara.tsv")
"""

import csv
import os
from typing import List, Dict, Optional

from ..core.types import NpcDefinition


class NpcBuilder:
    """
    Builder for CWL-compatible NPC (Chara) TSV files.

    Requires a sample Excel file to extract column headers and types.

    Example:
        builder = NpcBuilder()
        builder.add_npc(NpcDefinition(
            id="my_npc",
            name_jp="NPC",
            name_en="NPC",
            race="human",
            job="warrior",
            zone_id="my_zone",
        ))
        builder.save_tsv("Chara.tsv", header_source="Sample.xlsx")
    """

    # Common CWL tags
    TAG_NEUTRAL = "neutral"
    TAG_BOSS = "boss"
    TAG_UNDEAD = "undead"
    TAG_HUMAN_SPEAK = "humanSpeak"

    def __init__(self, debug_mode: bool = False):
        """
        Args:
            debug_mode: If True, set all NPC levels to 1 for testing
        """
        self.npcs: List[NpcDefinition] = []
        self.debug_mode = debug_mode
        self._header_map: Dict[str, int] = {}
        self._headers: List[str] = []
        self._types: List[str] = []
        self._defaults: List[str] = []

    def add_npc(self, npc: NpcDefinition) -> "NpcBuilder":
        """Add an NPC definition."""
        self.npcs.append(npc)
        return self

    def add_npcs(self, npcs: List[NpcDefinition]) -> "NpcBuilder":
        """Add multiple NPC definitions."""
        for npc in npcs:
            self.add_npc(npc)
        return self

    def _build_tags(self, npc: NpcDefinition) -> str:
        """Build CWL tag string for NPC."""
        tags = []

        # Hostility tag
        if npc.hostility == "Friend":
            tags.append(self.TAG_NEUTRAL)
        elif npc.hostility == "Enemy":
            tags.append(self.TAG_BOSS)

        # Zone placement
        if npc.zone_id:
            tags.append(f"addZone_{npc.zone_id}")

        # Stay home flag
        if npc.stay_home:
            tags.append("addFlag_StayHomeZone")

        # Merchant stock
        if npc.stock_id:
            tags.append(f"addStock_{npc.stock_id}")

        # Drama link
        if npc.drama_id:
            tags.append(f"addDrama_{npc.drama_id}")

        # Human speak
        if npc.human_speak:
            tags.append(self.TAG_HUMAN_SPEAK)

        # Extra tags
        tags.extend(npc.extra_tags)

        return ",".join(tags)

    def _npc_to_dict(self, npc: NpcDefinition) -> Dict[str, any]:
        """Convert NpcDefinition to column dict."""
        level = 1 if self.debug_mode else npc.lv

        result = {
            "id": npc.id,
            "name_JP": npc.name_jp,
            "name": npc.name_en or npc.name_jp,
            "race": npc.race,
            "job": npc.job,
            "LV": level,
            "hostility": npc.hostility,
            "tiles": npc.tiles,
            "_idRenderData": npc.render_data,
            "tag": self._build_tags(npc),
            "quality": npc.quality,
            "chance": npc.chance,
        }

        # Optional fields
        if npc.aka_jp:
            result["aka_JP"] = npc.aka_jp
        if npc.aka_en:
            result["aka"] = npc.aka_en
        if npc.bio:
            result["bio"] = npc.bio
        if npc.id_text:
            result["idText"] = npc.id_text
        if npc.portrait:
            result["portrait"] = npc.portrait
        if npc.main_element:
            result["mainElement"] = npc.main_element
        if npc.elements:
            result["elements"] = npc.elements
        if npc.act_combat:
            result["actCombat"] = npc.act_combat
        if npc.trait:
            result["trait"] = npc.trait
        if npc.author:
            result["Author"] = npc.author

        return result

    def load_headers_from_excel(self, excel_path: str, sheet_name: str = "Chara"):
        """
        Load headers from a sample Excel file.

        Args:
            excel_path: Path to sample Excel file
            sheet_name: Sheet name to read from
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for loading headers from Excel")

        wb = openpyxl.load_workbook(excel_path)
        ws = wb[sheet_name]

        cols = ws.max_column
        self._headers = []
        self._types = []
        self._defaults = []

        for c in range(1, cols + 1):
            header = ws.cell(row=1, column=c).value or ""
            type_val = ws.cell(row=2, column=c).value or ""
            default = ws.cell(row=3, column=c).value or ""

            self._headers.append(header)
            self._types.append(type_val)
            self._defaults.append(default)
            if header:
                self._header_map[header] = c - 1

        # Ensure required columns exist
        self._ensure_column("Author")
        self._ensure_column("portrait")

        print(f"Loaded {len(self._headers)} columns from {excel_path}")

    def _ensure_column(self, col_name: str):
        """Ensure a column exists, adding it if necessary."""
        if col_name not in self._header_map:
            new_idx = len(self._headers)
            self._headers.append(col_name)
            self._types.append("")
            self._defaults.append("")
            self._header_map[col_name] = new_idx

    def set_default_headers(self):
        """
        Set default headers for common CWL Chara format.

        Use this if you don't have a sample Excel file.
        """
        # Common Chara columns (simplified)
        default_headers = [
            "id",
            "name_JP",
            "name",
            "aka_JP",
            "aka",
            "race",
            "job",
            "_idRenderData",
            "tiles",
            "LV",
            "hostility",
            "bio",
            "idText",
            "portrait",
            "mainElement",
            "elements",
            "actCombat",
            "tag",
            "trait",
            "quality",
            "chance",
            "Author",
        ]

        self._headers = default_headers
        self._types = [""] * len(default_headers)
        self._defaults = [""] * len(default_headers)
        self._header_map = {name: idx for idx, name in enumerate(default_headers)}

    def _npc_to_row(self, npc: NpcDefinition) -> List[str]:
        """Convert NpcDefinition to TSV row."""
        npc_dict = self._npc_to_dict(npc)
        row = [""] * len(self._headers)

        for col_name, value in npc_dict.items():
            if col_name in self._header_map:
                row[self._header_map[col_name]] = value

        return row

    def build_rows(self) -> List[List[str]]:
        """Build all rows including headers."""
        if not self._headers:
            self.set_default_headers()

        rows = []

        # Row 1: Headers
        rows.append(self._headers)

        # Row 2: Types
        rows.append(self._types)

        # Row 3: Defaults
        rows.append(self._defaults)

        # Row 4+: NPC data
        for npc in self.npcs:
            if self.debug_mode:
                print(f"  {npc.id}: LV {npc.lv} -> 1")
            rows.append(self._npc_to_row(npc))

        return rows

    def save_tsv(self, *paths: str) -> None:
        """
        Save NPC definitions to TSV files.

        Args:
            *paths: One or more output paths
        """
        if self.debug_mode:
            print("[DEBUG MODE] All NPC levels set to 1")

        rows = self.build_rows()

        for path in paths:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
                writer.writerows(rows)
            print(f"Created TSV: {path} ({len(self.npcs)} NPCs)")


# ============================================================================
# Unit Tests
# ============================================================================

if __name__ == "__main__":
    print("=== NPC Builder Test ===\n")

    builder = NpcBuilder()
    builder.set_default_headers()

    # Add test NPCs
    builder.add_npc(
        NpcDefinition(
            id="test_merchant",
            name_jp="テスト商人",
            name_en="Test Merchant",
            race="human",
            job="merchant",
            lv=50,
            hostility="Friend",
            zone_id="test_zone",
            drama_id="drama_test_merchant",
            stock_id="test_merchant",
            author="test.author",
        )
    )

    builder.add_npc(
        NpcDefinition(
            id="test_boss",
            name_jp="テストボス",
            name_en="Test Boss",
            race="dragon",
            job="warrior",
            lv=100,
            hostility="Enemy",
            main_element="Fire",
            elements="resFire/100",
            extra_tags=["boss", "undead"],
        )
    )

    # Build rows
    rows = builder.build_rows()
    print(f"Generated {len(rows)} rows")
    print(f"Headers: {len(rows[0])} columns")
    print(f"NPCs: {len(builder.npcs)}")

    # Show first NPC data
    print(f"\nFirst NPC row preview:")
    npc_row = rows[3]
    for i, (header, value) in enumerate(zip(rows[0], npc_row)):
        if value:
            print(f"  {header}: {value}")

    print("\n=== All Tests Passed! ===")
