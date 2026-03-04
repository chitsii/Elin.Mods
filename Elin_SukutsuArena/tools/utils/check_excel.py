import openpyxl
import os

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(TOOLS_DIR))
DEFAULT_SOURCE_GAME_PATH = os.path.join(PROJECT_ROOT, "..", "SourceExcels", "SourceGame.xlsx")
SOURCE_GAME_PATH = os.environ.get("SOURCE_GAME_XLSX", DEFAULT_SOURCE_GAME_PATH)

wb = openpyxl.load_workbook(SOURCE_GAME_PATH)
ws = wb["Chara"]

# Header (Row 1)
headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    if val:  # Only non-None
        headers.append((col, val))

# Data Row 4 - full dump
print("=== Sample NPC Full Data ===")
for col, header in headers:
    val = ws.cell(row=4, column=col).value
    if val is not None and val != "":
        print(f"{header}: {val}")
