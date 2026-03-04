import openpyxl
import os

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(os.path.dirname(TOOLS_DIR))
DEFAULT_SOURCE_GAME_PATH = os.path.join(PROJECT_ROOT, "..", "SourceExcels", "SourceGame.xlsx")
SOURCE_GAME_PATH = os.environ.get("SOURCE_GAME_XLSX", DEFAULT_SOURCE_GAME_PATH)

if os.path.exists(SOURCE_GAME_PATH):
    wb = openpyxl.load_workbook(SOURCE_GAME_PATH)
    if "Chara" in wb.sheetnames:
        ws = wb["Chara"]
        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                headers.append(val)
        print("Chara sheet headers:", headers)
    else:
        print("Chara sheet not found.")
else:
    print(f"SourceGame file not found at {SOURCE_GAME_PATH}")
